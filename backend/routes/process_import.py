"""
Process Intelligence Import Routes - Upload process diagrams and build asset hierarchy.

Converts process flow diagrams (PFDs), schematics, and engineering documents
into ISO 14224-aligned asset hierarchies for AssetIQ.
"""

import asyncio
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import Optional, List, Dict, Any
from datetime import datetime, timezone
import logging
import io

from database import db
from auth import get_current_user
from services.process_import_service import ProcessImportService

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/process-import", tags=["Process Intelligence Import"])


# ============== MODELS ==============

class ProcessImportOptions(BaseModel):
    """Options for process import."""
    generate_subunits: bool = True
    generate_maintainable_items: bool = False
    estimate_criticality: bool = True


class ItemUpdateRequest(BaseModel):
    """Request model for updating a hierarchy item."""
    tag: Optional[str] = None
    name: Optional[str] = None
    level: Optional[str] = None
    equipment_type: Optional[str] = None
    description: Optional[str] = None
    parent_id: Optional[str] = None
    criticality: Optional[Dict[str, int]] = None


class AddItemRequest(BaseModel):
    """Request model for adding a new item."""
    tag: str
    name: str
    level: str
    equipment_type: Optional[str] = ""
    description: Optional[str] = ""
    parent_id: Optional[str] = None
    criticality: Optional[Dict[str, int]] = None


class ImportRequest(BaseModel):
    """Request model for final import."""
    installation_id: str


# Background task for processing
async def process_diagram_background(
    session_id: str,
    file_name: str,
    file_type: str,
    file_content: bytes,
    options: Dict[str, Any]
):
    """Process diagram file in background."""
    service = ProcessImportService(db)
    try:
        await service.process_session(
            session_id=session_id,
            file_name=file_name,
            file_type=file_type,
            file_content=file_content,
            options=options
        )
    except Exception as e:
        logger.error(f"Background processing error for session {session_id}: {e}", exc_info=True)
        await db.process_import_sessions.update_one(
            {"session_id": session_id},
            {"$set": {
                "status": "error",
                "error_message": str(e),
                "progress_message": f"Error: {str(e)[:200]}",
                "updated_at": datetime.now(timezone.utc)
            }}
        )


# ============== ROUTES ==============

@router.post("/upload")
async def upload_process_diagram(
    file: UploadFile = File(...),
    generate_subunits: bool = Query(True, description="Auto-generate subunits from templates"),
    generate_maintainable_items: bool = Query(False, description="Auto-generate maintainable items"),
    estimate_criticality: bool = Query(True, description="AI-estimate criticality scores"),
    current_user: dict = Depends(get_current_user)
):
    """
    Upload a process diagram for analysis.
    
    Supported file types:
    - PDF (PFDs, process schematics)
    - Images (.png, .jpg, .jpeg, .webp)
    
    Returns a session ID to track processing progress.
    """
    
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")
    
    filename_lower = file.filename.lower()
    
    if filename_lower.endswith('.pdf'):
        file_type = 'pdf'
    elif filename_lower.endswith('.png'):
        file_type = 'png'
    elif filename_lower.endswith(('.jpg', '.jpeg')):
        file_type = 'jpg'
    elif filename_lower.endswith('.webp'):
        file_type = 'webp'
    else:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file type. Supported: PDF, Images (.png, .jpg, .jpeg, .webp)"
        )
    
    content = await file.read()
    
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="Empty file uploaded")
    
    if len(content) > 50 * 1024 * 1024:  # 50MB limit for diagrams
        raise HTTPException(status_code=400, detail="File too large. Maximum size is 50MB")
    
    service = ProcessImportService(db)
    user_id = current_user.get("id", current_user.get("email", "unknown"))
    
    options = {
        "generate_subunits": generate_subunits,
        "generate_maintainable_items": generate_maintainable_items,
        "estimate_criticality": estimate_criticality
    }
    
    try:
        session_id = await service.create_session_placeholder(
            file_name=file.filename,
            file_type=file_type,
            created_by=user_id,
            options=options
        )
        
        # Process in background
        asyncio.create_task(process_diagram_background(
            session_id=session_id,
            file_name=file.filename,
            file_type=file_type,
            file_content=content,
            options=options
        ))
        
        return {
            "session_id": session_id,
            "status": "processing",
            "message": "File uploaded. Processing started in background.",
            "options": options
        }
        
    except Exception as e:
        logger.error(f"Upload error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/session/{session_id}")
async def get_session(
    session_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get session status and results."""
    
    service = ProcessImportService(db)
    session = await service.get_session(session_id)
    
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session.pop("_id", None)
    return session


@router.patch("/session/{session_id}/item/{item_id}")
async def update_item(
    session_id: str,
    item_id: str,
    updates: ItemUpdateRequest,
    current_user: dict = Depends(get_current_user)
):
    """Update a hierarchy item."""
    
    service = ProcessImportService(db)
    
    update_data = updates.model_dump(exclude_none=True)
    if not update_data:
        raise HTTPException(status_code=400, detail="No updates provided")
    
    result = await service.update_item(session_id, item_id, update_data)
    
    if not result:
        raise HTTPException(status_code=404, detail="Session or item not found")
    
    return {"success": True, "stats": result["stats"]}


@router.delete("/session/{session_id}/item/{item_id}")
async def delete_item(
    session_id: str,
    item_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a hierarchy item and its children."""
    
    service = ProcessImportService(db)
    result = await service.delete_item(session_id, item_id)
    
    if not result:
        raise HTTPException(status_code=404, detail="Session or item not found")
    
    return {"success": True, "stats": result["stats"]}


@router.post("/session/{session_id}/item")
async def add_item(
    session_id: str,
    item_data: AddItemRequest,
    current_user: dict = Depends(get_current_user)
):
    """Add a new hierarchy item manually."""
    
    service = ProcessImportService(db)
    result = await service.add_item(session_id, item_data.model_dump())
    
    if not result:
        raise HTTPException(status_code=404, detail="Session not found")
    
    return {"success": True, "item": result["item"], "stats": result["stats"]}


@router.post("/session/{session_id}/item/{item_id}/accept")
async def accept_item(
    session_id: str,
    item_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Mark an item as accepted."""
    
    service = ProcessImportService(db)
    result = await service.update_item(session_id, item_id, {"review_status": "accepted"})
    
    if not result:
        raise HTTPException(status_code=404, detail="Session or item not found")
    
    return {"success": True, "stats": result["stats"]}


@router.post("/session/{session_id}/item/{item_id}/reject")
async def reject_item(
    session_id: str,
    item_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Mark an item as rejected."""
    
    service = ProcessImportService(db)
    result = await service.update_item(session_id, item_id, {"review_status": "rejected"})
    
    if not result:
        raise HTTPException(status_code=404, detail="Session or item not found")
    
    return {"success": True, "stats": result["stats"]}


@router.post("/session/{session_id}/accept-all")
async def accept_all_items(
    session_id: str,
    min_confidence: int = Query(70, description="Minimum confidence to accept"),
    current_user: dict = Depends(get_current_user)
):
    """Accept all items with confidence >= threshold."""
    
    session = await db.process_import_sessions.find_one({"session_id": session_id})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    items = session.get("hierarchy_items", [])
    accepted_count = 0
    
    for item in items:
        if item.get("review_status") == "pending" and item.get("confidence", 0) >= min_confidence:
            item["review_status"] = "accepted"
            accepted_count += 1
    
    service = ProcessImportService(db)
    stats = service._calculate_stats(items, session.get("exceptions", []))
    
    await db.process_import_sessions.update_one(
        {"session_id": session_id},
        {"$set": {
            "hierarchy_items": items,
            "stats": stats,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {"success": True, "accepted_count": accepted_count, "stats": stats}


@router.post("/session/{session_id}/import")
async def import_to_assetiq(
    session_id: str,
    request: ImportRequest,
    current_user: dict = Depends(get_current_user)
):
    """Import hierarchy to AssetIQ equipment register."""
    
    service = ProcessImportService(db)
    
    try:
        result = await service.import_to_assetiq(
            session_id=session_id,
            installation_id=request.installation_id,
            created_by=current_user.get("id", current_user.get("email", "unknown"))
        )
        return result
        
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.error(f"Import error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/session/{session_id}/export")
async def export_csv(
    session_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Export hierarchy as CSV in AssetIQ import format."""
    import csv
    
    session = await db.process_import_sessions.find_one({"session_id": session_id})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    items = session.get("hierarchy_items", [])
    
    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header row matching AssetIQ import format
    writer.writerow([
        "ID or Tag",
        "Name",
        "Level",
        "Equipment Type",
        "Description",
        "Safety",
        "Production",
        "Environmental",
        "Reputation"
    ])
    
    # Data rows
    for item in items:
        if item.get("review_status") == "rejected":
            continue
        
        crit = item.get("criticality", {})
        writer.writerow([
            item.get("tag", ""),
            item.get("name", ""),
            item.get("level", ""),
            item.get("equipment_type", ""),
            item.get("description", ""),
            crit.get("safety", 0),
            crit.get("production", 0),
            crit.get("environmental", 0),
            crit.get("reputation", 0)
        ])
    
    # Return as file
    output.seek(0)
    content = output.getvalue().encode('utf-8')
    
    filename = f"hierarchy_export_{session_id[:8]}.csv"
    
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/session/{session_id}/export-excel")
async def export_excel(
    session_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Export hierarchy as Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    session = await db.process_import_sessions.find_one({"session_id": session_id})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    items = session.get("hierarchy_items", [])
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Asset Hierarchy"
    
    # Headers
    headers = [
        "ID or Tag", "Name", "Level", "Equipment Type", "Description",
        "Safety", "Production", "Environmental", "Reputation",
        "Confidence", "AI Reasoning"
    ]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # Data rows
    row_idx = 2
    for item in items:
        if item.get("review_status") == "rejected":
            continue
        
        crit = item.get("criticality", {})
        row_data = [
            item.get("tag", ""),
            item.get("name", ""),
            item.get("level", ""),
            item.get("equipment_type", ""),
            item.get("description", ""),
            crit.get("safety", 0),
            crit.get("production", 0),
            crit.get("environmental", 0),
            crit.get("reputation", 0),
            item.get("confidence", 0),
            item.get("ai_reasoning", "")
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            
            # Color code confidence
            if col == 10:  # Confidence
                if value >= 90:
                    cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                elif value >= 70:
                    cell.fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        
        row_idx += 1
    
    # Adjust column widths
    column_widths = [15, 30, 18, 18, 40, 10, 12, 14, 12, 12, 50]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    ws.freeze_panes = "A2"
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"hierarchy_export_{session_id[:8]}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.delete("/session/{session_id}")
async def delete_session(
    session_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a session."""
    
    result = await db.process_import_sessions.delete_one({"session_id": session_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Session not found")
    
    return {"success": True}


@router.get("/sessions")
async def list_sessions(
    limit: int = Query(20, le=100),
    skip: int = Query(0, ge=0),
    current_user: dict = Depends(get_current_user)
):
    """List user's import sessions."""
    
    user_id = current_user.get("id", current_user.get("email", "unknown"))
    
    cursor = db.process_import_sessions.find(
        {"created_by": user_id}
    ).sort("created_at", -1).skip(skip).limit(limit)
    
    sessions = []
    async for session in cursor:
        sessions.append({
            "session_id": session["session_id"],
            "file_name": session["file_name"],
            "status": session["status"],
            "stats": session.get("stats", {}),
            "created_at": session["created_at"].isoformat() if session.get("created_at") else None
        })
    
    total = await db.process_import_sessions.count_documents({"created_by": user_id})
    
    return {"sessions": sessions, "total": total}
