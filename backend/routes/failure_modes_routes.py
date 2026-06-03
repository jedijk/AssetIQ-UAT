"""
Failure Modes routes.
"""
from fastapi import APIRouter, Depends, HTTPException, Body, BackgroundTasks
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import List, Optional, Any
from datetime import datetime, timezone
import uuid
import logging
import io
from database import db, failure_modes_service, efm_service
from auth import get_current_user
from services.threat_score_service import recalculate_threat_scores_for_failure_mode
from services.translation_service import TranslationService
from models.translation import EntityType
from failure_modes import FAILURE_MODES_LIBRARY
logger = logging.getLogger(__name__)

router = APIRouter(tags=["Failure Modes"])


# Helper functions for fallback data
def get_all_categories():
    """Get unique categories from static library."""
    categories = set()
    for fm in FAILURE_MODES_LIBRARY:
        if fm.get("category"):
            categories.add(fm["category"])
    return sorted(list(categories))


def get_all_equipment_types():
    """Get unique equipment types from static library."""
    equipment_types = set()
    for fm in FAILURE_MODES_LIBRARY:
        if fm.get("equipment"):
            equipment_types.add(fm["equipment"])
    return sorted(list(equipment_types))


@router.get("/failure-modes")
async def get_failure_modes(
    category: Optional[str] = None,
    equipment: Optional[str] = None,
    search: Optional[str] = None,
    min_rpn: Optional[int] = None,
    equipment_type_id: Optional[str] = None,
    mechanism: Optional[str] = None,
    is_validated: Optional[bool] = None,
    failure_mode_type: Optional[str] = None,  # "generic", "customer_specific", or "recently_added"
    recently_added_days: Optional[int] = 30,  # For recently_added filter, default 30 days
    skip: int = 0,
    limit: int = 500
):
    """Get failure modes from MongoDB with optional filters."""
    # Try MongoDB first, fall back to static library if empty
    try:
        result = await failure_modes_service.get_all(
            category=category,
            equipment=equipment,
            search=search,
            min_rpn=min_rpn,
            equipment_type_id=equipment_type_id,
            mechanism=mechanism,
            is_validated=is_validated,
            failure_mode_type=failure_mode_type,
            recently_added_days=recently_added_days if failure_mode_type == "recently_added" else None,
            skip=skip,
            limit=limit
        )
        
        # If MongoDB is empty, use static library as fallback
        if result["total"] == 0 and not any([category, equipment, search, min_rpn, equipment_type_id, mechanism]):
            logger.info("MongoDB failure_modes empty, using static library fallback")
            results = FAILURE_MODES_LIBRARY.copy()
            results.sort(key=lambda x: -x["rpn"])
            return {"total": len(results), "failure_modes": results, "source": "static"}
        
        return result
    except Exception as e:
        logger.error(f"Error fetching from MongoDB, falling back to static: {e}")
        # Fallback to static library
        results = FAILURE_MODES_LIBRARY.copy()
        
        if search:
            search_lower = search.lower()
            results = [fm for fm in results if (
                search_lower in fm["failure_mode"].lower() or
                search_lower in fm["equipment"].lower() or
                search_lower in fm["category"].lower() or
                any(search_lower in kw.lower() for kw in fm.get("keywords", []))
            )]
        
        if category and category.lower() != "all":
            results = [fm for fm in results if fm["category"].lower() == category.lower()]
        
        if equipment:
            results = [fm for fm in results if fm["equipment"].lower() == equipment.lower()]
        
        if min_rpn:
            results = [fm for fm in results if fm["rpn"] >= min_rpn]
        
        results.sort(key=lambda x: -x["rpn"])
        return {"total": len(results), "failure_modes": results, "source": "static"}

@router.get("/failure-modes/categories")
async def get_categories():
    """Get all unique categories."""
    try:
        categories = await failure_modes_service.get_categories()
        if categories:
            return {"categories": categories}
    except Exception as e:
        logger.error(f"Error fetching categories from MongoDB: {e}")
    # Fallback to static
    return {"categories": get_all_categories()}

@router.get("/failure-modes/equipment-types")
async def get_equipment_types():
    """Get all unique equipment types."""
    try:
        types = await failure_modes_service.get_equipment_types()
        if types:
            return {"equipment_types": types}
    except Exception as e:
        logger.error(f"Error fetching equipment types from MongoDB: {e}")
    # Fallback to static
    return {"equipment_types": get_all_equipment_types()}

@router.get("/failure-modes/mechanisms")
async def get_mechanisms():
    """Get all unique ISO 14224 mechanisms."""
    try:
        mechanisms = await failure_modes_service.get_mechanisms()
        return {"mechanisms": mechanisms}
    except Exception as e:
        logger.error(f"Error fetching mechanisms: {e}")
        return {"mechanisms": []}


@router.get("/failure-modes/counts-by-equipment-type")
async def get_failure_mode_counts_by_equipment_type(
    current_user: dict = Depends(get_current_user)
):
    """
    Get failure mode counts grouped by equipment_type_id.
    Returns a dict mapping equipment_type_id to count of failure modes.
    """
    try:
        # Aggregate to count failure modes per equipment type
        pipeline = [
            {"$unwind": "$equipment_type_ids"},
            {"$group": {
                "_id": "$equipment_type_ids",
                "count": {"$sum": 1}
            }},
            {"$sort": {"count": -1}}
        ]
        
        results = await db.failure_modes.aggregate(pipeline).to_list(500)
        
        counts = {item["_id"]: item["count"] for item in results if item["_id"]}
        
        # Also get total count
        total = await db.failure_modes.count_documents({})
        
        return {
            "counts_by_type": counts,
            "total_failure_modes": total
        }
    except Exception as e:
        logger.error(f"Error fetching failure mode counts: {e}")
        return {"counts_by_type": {}, "total_failure_modes": 0}


@router.get("/failure-modes/export")
async def export_failure_modes_excel(
    current_user: dict = Depends(get_current_user)
):
    """Export all failure modes to an Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    try:
        # Fetch all failure modes
        result = await failure_modes_service.get_all(skip=0, limit=10000)
        failure_modes = result.get("failure_modes", []) if isinstance(result, dict) else []
        if not failure_modes:
            # Fallback to static library
            failure_modes = FAILURE_MODES_LIBRARY.copy()
    except Exception as e:
        logger.error(f"Error fetching failure modes for export: {e}")
        failure_modes = FAILURE_MODES_LIBRARY.copy()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Failure Modes"
    
    # Define headers
    headers = [
        "ID", "Category", "Equipment", "Failure Mode", "Process",
        "Potential Effects", "Potential Causes", "ISO 14224 Mechanism",
        "Severity", "Occurrence", "Detectability", "RPN",
        "Keywords", "Recommended Actions", "Validated", "Source"
    ]
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    for row_idx, fm in enumerate(failure_modes, 2):
        # Format keywords as comma-separated string
        keywords = fm.get("keywords", [])
        keywords_str = ", ".join(keywords) if isinstance(keywords, list) else str(keywords or "")
        
        # Format recommended actions
        actions = fm.get("recommended_actions", [])
        if isinstance(actions, list):
            if actions and isinstance(actions[0], dict):
                actions_str = "\n".join([
                    (
                        f"• {a.get('action') or a.get('description') or ''}".strip()
                        + (f" ({a.get('action_type', '')})" if a.get('action_type') else "")
                        + (f" — {a.get('estimated_minutes')} min" if a.get('estimated_minutes') not in (None, '', 0) else "")
                    )
                    for a in actions
                ])
            else:
                actions_str = "\n".join([f"• {a}" for a in actions])
        else:
            actions_str = str(actions or "")
        
        # Format potential_effects (could be list or string)
        potential_effects = fm.get("potential_effects", "")
        if isinstance(potential_effects, list):
            potential_effects_str = ", ".join(potential_effects)
        else:
            potential_effects_str = str(potential_effects or "")
        
        # Format potential_causes (could be list or string)
        potential_causes = fm.get("potential_causes", "")
        if isinstance(potential_causes, list):
            potential_causes_str = ", ".join(potential_causes)
        else:
            potential_causes_str = str(potential_causes or "")
        
        row_data = [
            str(fm.get("id", "")),
            fm.get("category", ""),
            fm.get("equipment", ""),
            fm.get("failure_mode", ""),
            fm.get("process", ""),
            potential_effects_str,
            potential_causes_str,
            fm.get("iso14224_mechanism", ""),
            fm.get("severity", 0),
            fm.get("occurrence", 0),
            fm.get("detectability", 0),
            fm.get("rpn", 0),
            keywords_str,
            actions_str,
            "Yes" if fm.get("is_validated") else "No",
            fm.get("source", "library")
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    # Adjust column widths
    column_widths = [10, 15, 18, 30, 20, 30, 30, 20, 10, 10, 12, 8, 25, 40, 10, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"failure_modes_{timestamp}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/failure-modes/high-risk")
async def get_high_risk_modes(threshold: int = 150):
    """Get failure modes with RPN above threshold."""
    try:
        high_risk = await failure_modes_service.get_high_risk(threshold)
        return {
            "threshold": threshold,
            "total": len(high_risk),
            "failure_modes": high_risk
        }
    except Exception as e:
        logger.error(f"Error fetching high-risk modes: {e}")
        # Fallback
        high_risk = [fm for fm in FAILURE_MODES_LIBRARY if fm["rpn"] >= threshold]
        high_risk.sort(key=lambda x: -x["rpn"])
        return {"threshold": threshold, "total": len(high_risk), "failure_modes": high_risk}

@router.get("/failure-modes/{mode_id}")
async def get_failure_mode_by_id(mode_id: str):
    """Get a specific failure mode by ID (MongoDB _id or legacy_id)."""
    try:
        fm = await failure_modes_service.get_by_id(mode_id)
        if fm:
            return fm
    except Exception as e:
        logger.error(f"Error fetching failure mode {mode_id}: {e}")
    
    # Fallback to static library with legacy_id
    try:
        legacy_id = int(mode_id)
        for fm in FAILURE_MODES_LIBRARY:
            if fm["id"] == legacy_id:
                return fm
    except ValueError:
        pass
    
    raise HTTPException(status_code=404, detail="Failure mode not found")


# Failure Mode CRUD Models
class FailureModeCreate(BaseModel):
    category: str
    equipment: str
    failure_mode: str
    keywords: List[str] = []
    severity: int = Field(ge=1, le=10)
    occurrence: int = Field(ge=1, le=10)
    detectability: int = Field(ge=1, le=10)
    recommended_actions: List[Any] = []
    equipment_type_ids: List[str] = []
    description: Optional[str] = None
    source: Optional[str] = None  # e.g., "observation", "manual", "import"
    linked_threat_id: Optional[str] = None
    failure_mode_type: Optional[str] = None  # "generic" or "customer_specific"
    # New fields
    process: Optional[str] = None  # Process area
    potential_effects: Optional[Any] = None  # Potential effects of failure mode (string or list)
    potential_causes: Optional[Any] = None  # Potential cause of failure mode (string or list)
    iso14224_mechanism: Optional[str] = None  # ISO 14224 failure mechanism

    model_config = {"extra": "ignore"}


class FailureModeUpdate(BaseModel):
    category: Optional[str] = None
    equipment: Optional[str] = None
    failure_mode: Optional[str] = None
    keywords: Optional[List[str]] = None
    severity: Optional[int] = Field(None, ge=1, le=10)
    occurrence: Optional[int] = Field(None, ge=1, le=10)
    detectability: Optional[int] = Field(None, ge=1, le=10)
    recommended_actions: Optional[List[Any]] = None
    equipment_type_ids: Optional[List[str]] = None
    # Validation fields
    is_validated: Optional[bool] = None
    validated_by_name: Optional[str] = None
    validated_by_position: Optional[str] = None
    validated_at: Optional[str] = None
    # New fields
    process: Optional[str] = None
    potential_effects: Optional[Any] = None
    potential_causes: Optional[Any] = None
    iso14224_mechanism: Optional[str] = None
    failure_mode_type: Optional[str] = None
    change_reason: Optional[str] = None
    # AI provenance — set when "Improve with AI" is applied.
    ai_improved_at: Optional[str] = None

    model_config = {"extra": "ignore"}


class FailureModeValidation(BaseModel):
    """Model for validating a failure mode"""
    validated_by_name: str
    validated_by_position: str
    validated_by_id: Optional[str] = None


def auto_link_equipment_types(equipment_name: str) -> List[str]:
    """Auto-detect equipment types based on equipment name. Returns list of matching type IDs."""
    equipment_lower = equipment_name.lower()
    
    # Map common equipment names to equipment type IDs
    equipment_mapping = {
        "pump": ["pump_centrifugal", "pump_reciprocating"],
        "centrifugal pump": ["pump_centrifugal"],
        "reciprocating pump": ["pump_reciprocating"],
        "compressor": ["compressor_centrifugal", "compressor_reciprocating"],
        "centrifugal compressor": ["compressor_centrifugal"],
        "reciprocating compressor": ["compressor_reciprocating"], 
        "turbine": ["turbine_gas", "turbine_steam"],
        "gas turbine": ["turbine_gas"],
        "steam turbine": ["turbine_steam"],
        "motor": ["motor_electric"],
        "vessel": ["vessel_pressure", "vessel_storage"],
        "pressure vessel": ["vessel_pressure"],
        "storage tank": ["vessel_storage"],
        "tank": ["vessel_storage"],
        "heat exchanger": ["heat_exchanger"],
        "exchanger": ["heat_exchanger"],
        "pipe": ["pipe"],
        "piping": ["pipe"],
        "valve": ["valve_control", "valve_safety", "valve_manual"],
        "control valve": ["valve_control"],
        "safety valve": ["valve_safety"],
        "manual valve": ["valve_manual"],
        "sensor": ["sensor_pressure", "sensor_temperature", "sensor_flow"],
        "pressure sensor": ["sensor_pressure"],
        "temperature sensor": ["sensor_temperature"],
        "flow sensor": ["sensor_flow"],
        "transmitter": ["sensor_pressure", "sensor_temperature", "sensor_flow"],
        "plc": ["plc"],
        "controller": ["plc"],
        "generator": ["turbine_gas", "turbine_steam"],
        "transformer": ["transformer"],
        "switchgear": ["switchgear"],
        "extruder": ["extruder"],
        "filter": ["pipe"],
        "boiler": ["heat_exchanger"],
        "furnace": ["heat_exchanger"],
        "heater": ["heat_exchanger"],
        "cooler": ["heat_exchanger"],
        "fan": ["motor_electric"],
        "blower": ["compressor_centrifugal"],
    }
    
    for keyword, type_ids in equipment_mapping.items():
        if keyword in equipment_lower:
            return type_ids
    return []


async def auto_translate_failure_mode(fm_id: str, fm_data: dict, created_by: str = None):
    """Background task to auto-translate a failure mode to Dutch and German."""
    try:
        translation_service = TranslationService(db)
        
        # Translate to Dutch and German
        await translation_service.translate_entity(
            entity_type=EntityType.FAILURE_MODE,
            entity_id=fm_id,
            entity_data=fm_data,
            target_languages=["nl", "de"],
            created_by=created_by
        )
        logger.info(f"Auto-translated failure mode {fm_id} to Dutch and German")
    except Exception as e:
        logger.error(f"Failed to auto-translate failure mode {fm_id}: {e}")


@router.post("/failure-modes")
async def create_failure_mode(
    data: FailureModeCreate,
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user)
):
    """Create a new failure mode in MongoDB."""
    # Auto-link equipment types if not provided
    equipment_type_ids = data.equipment_type_ids if data.equipment_type_ids else auto_link_equipment_types(data.equipment)
    
    try:
        # Check for duplicate failure mode name
        existing = await failure_modes_service.find_by_name(data.failure_mode)
        if existing:
            raise HTTPException(
                status_code=400, 
                detail="A failure mode with this name already exists"
            )
        
        new_fm = await failure_modes_service.create(
            data={
                "category": data.category,
                "equipment": data.equipment,
                "failure_mode": data.failure_mode,
                "keywords": data.keywords,
                "severity": data.severity,
                "occurrence": data.occurrence,
                "detectability": data.detectability,
                "recommended_actions": data.recommended_actions,
                "equipment_type_ids": equipment_type_ids,
                "description": data.description,
                "source": data.source,
                "linked_threat_id": data.linked_threat_id,
                "failure_mode_type": data.failure_mode_type or "generic",
                "process": data.process,
                "potential_effects": data.potential_effects,
                "potential_causes": data.potential_causes,
                "iso14224_mechanism": data.iso14224_mechanism,
            },
            created_by=current_user["id"]
        )
        
        # Trigger auto-translation in background
        # NOTE: Use the failure_mode NAME as the entity_id so that the frontend
        # can look up translations via fm.failure_mode (consistent with library data).
        fm_data_for_translation = {
            "name": data.failure_mode,
            "description": data.description or "",
            "effects": data.potential_effects or "",
            "causes": data.potential_causes or "",
            "recommended_actions": ", ".join(data.recommended_actions) if data.recommended_actions else "",
        }
        background_tasks.add_task(
            auto_translate_failure_mode, 
            data.failure_mode,
            fm_data_for_translation,
            current_user["id"]
        )
        
        return new_fm
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating failure mode in MongoDB: {e}")
        # Fallback to in-memory (will be lost on restart)
        max_id = max((fm["id"] for fm in FAILURE_MODES_LIBRARY), default=0)
        new_id = max_id + 1
        
        new_fm = {
            "id": new_id,
            "category": data.category,
            "equipment": data.equipment,
            "failure_mode": data.failure_mode,
            "keywords": data.keywords,
            "severity": data.severity,
            "occurrence": data.occurrence,
            "detectability": data.detectability,
            "rpn": data.severity * data.occurrence * data.detectability,
            "recommended_actions": data.recommended_actions,
            "equipment_type_ids": equipment_type_ids,
            "process": data.process,
            "potential_effects": data.potential_effects,
            "potential_causes": data.potential_causes,
            "iso14224_mechanism": data.iso14224_mechanism,
            "is_custom": True,
            "created_by": current_user["id"]
        }
        
        FAILURE_MODES_LIBRARY.append(new_fm)
        return new_fm


@router.patch("/failure-modes/{mode_id}")
async def update_failure_mode(
    mode_id: str,
    data: FailureModeUpdate,
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user)
):
    """Update a failure mode in MongoDB."""
    try:
        update_data = {}
        if data.category is not None:
            update_data["category"] = data.category
        if data.equipment is not None:
            update_data["equipment"] = data.equipment
            # Auto-link if equipment changed and no explicit types provided
            if data.equipment_type_ids is None:
                auto_types = auto_link_equipment_types(data.equipment)
                if auto_types:
                    update_data["equipment_type_ids"] = auto_types
        if data.failure_mode is not None:
            update_data["failure_mode"] = data.failure_mode
        if data.keywords is not None:
            update_data["keywords"] = data.keywords
        if data.severity is not None:
            update_data["severity"] = data.severity
        if data.occurrence is not None:
            update_data["occurrence"] = data.occurrence
        if data.detectability is not None:
            update_data["detectability"] = data.detectability
        if data.recommended_actions is not None:
            update_data["recommended_actions"] = data.recommended_actions
        if data.equipment_type_ids is not None:
            update_data["equipment_type_ids"] = data.equipment_type_ids
        # New fields
        if data.process is not None:
            update_data["process"] = data.process
        if data.potential_effects is not None:
            update_data["potential_effects"] = data.potential_effects
        if data.potential_causes is not None:
            update_data["potential_causes"] = data.potential_causes
        if data.iso14224_mechanism is not None:
            update_data["iso14224_mechanism"] = data.iso14224_mechanism
        if data.failure_mode_type is not None:
            update_data["failure_mode_type"] = data.failure_mode_type
        if data.ai_improved_at is not None:
            update_data["ai_improved_at"] = data.ai_improved_at
        
        result = await failure_modes_service.update(
            mode_id, 
            update_data,
            updated_by=current_user.get("name", current_user.get("email", "Unknown")),
            change_reason=data.change_reason
        )
        
        if result:
            # If FMEA scores changed, recalculate all linked threat scores
            if result.get("fmea_changed"):
                old_name = result.get("old_failure_mode_name", result["failure_mode"])
                updated_threats = await recalculate_threat_scores_for_failure_mode(
                    old_name,
                    result["severity"],
                    result["occurrence"],
                    result["detectability"]
                )
                logger.info(f"Updated {updated_threats} threat scores after FMEA change")
                result["threats_updated"] = updated_threats
                
                # Also propagate changes to EFMs
                try:
                    efms_updated = await efm_service.propagate_template_change(
                        failure_mode_id=mode_id,
                        new_severity=data.severity,
                        new_occurrence=data.occurrence,
                        new_detectability=data.detectability
                    )
                    result["efms_updated"] = efms_updated
                    logger.info(f"Propagated FMEA change to {efms_updated} EFMs")
                except Exception as efm_err:
                    logger.error(f"Failed to propagate to EFMs: {efm_err}")
            
            # Trigger auto-translation in background when content changes
            # NOTE: Use the failure_mode NAME as the entity_id (matches frontend lookup).
            if any([data.failure_mode, data.potential_effects, 
                    data.potential_causes, data.recommended_actions]):
                fm_data_for_translation = {
                    "name": result.get("failure_mode", ""),
                    "description": result.get("mechanism", ""),  # Use mechanism as description
                    "effects": result.get("potential_effects", ""),
                    "causes": result.get("potential_causes", ""),
                    "recommended_actions": ", ".join(
                        [str(a) if isinstance(a, str) else a.get("description", str(a)) if isinstance(a, dict) else str(a) 
                         for a in (result.get("recommended_actions") or [])]
                    ),
                }
                background_tasks.add_task(
                    auto_translate_failure_mode,
                    result.get("failure_mode", "") or mode_id,
                    fm_data_for_translation,
                    current_user["id"]
                )
            
            return result
        
        raise HTTPException(status_code=404, detail="Failure mode not found")
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating failure mode: {e}")
        # Fallback to in-memory
        try:
            legacy_id = int(mode_id)
            for fm in FAILURE_MODES_LIBRARY:
                if fm["id"] == legacy_id:
                    if data.category is not None:
                        fm["category"] = data.category
                    if data.equipment is not None:
                        fm["equipment"] = data.equipment
                    if data.failure_mode is not None:
                        fm["failure_mode"] = data.failure_mode
                    if data.keywords is not None:
                        fm["keywords"] = data.keywords
                    if data.severity is not None:
                        fm["severity"] = data.severity
                    if data.occurrence is not None:
                        fm["occurrence"] = data.occurrence
                    if data.detectability is not None:
                        fm["detectability"] = data.detectability
                    if data.recommended_actions is not None:
                        fm["recommended_actions"] = data.recommended_actions
                    if data.equipment_type_ids is not None:
                        fm["equipment_type_ids"] = data.equipment_type_ids
                    fm["rpn"] = fm["severity"] * fm["occurrence"] * fm["detectability"]
                    return fm
        except ValueError:
            pass
        raise HTTPException(status_code=404, detail="Failure mode not found")


# ============= FAILURE MODE VERSION HISTORY ENDPOINTS =============

@router.get("/failure-modes/{mode_id}/versions")
async def get_failure_mode_versions(
    mode_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get version history for a failure mode."""
    try:
        versions = await failure_modes_service.get_versions(mode_id)
        return {"versions": versions, "total": len(versions)}
    except Exception as e:
        logger.error(f"Error getting failure mode versions: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class RollbackRequest(BaseModel):
    version_id: str
    reason: Optional[str] = None


@router.post("/failure-modes/{mode_id}/rollback")
async def rollback_failure_mode(
    mode_id: str,
    data: RollbackRequest,
    current_user: dict = Depends(get_current_user)
):
    """Rollback a failure mode to a specific version."""
    try:
        user_name = current_user.get("name", current_user.get("email", "Unknown"))
        result = await failure_modes_service.rollback_to_version(
            mode_id,
            data.version_id,
            rolled_back_by=user_name
        )
        
        if result:
            return {
                **result,
                "message": f"Rolled back to version {result.get('rolled_back_from_version', '?')}"
            }
        
        raise HTTPException(status_code=404, detail="Version or failure mode not found")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error rolling back failure mode: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/failure-modes/{mode_id}/validate")
async def validate_failure_mode(
    mode_id: str,
    data: FailureModeValidation,
    current_user: dict = Depends(get_current_user)
):
    """Validate a failure mode with validator name and position."""
    try:
        # Use provided user_id or fall back to current_user's id
        validated_by_id = data.validated_by_id or current_user.get("id") or current_user.get("user_id")
        
        result = await failure_modes_service.validate(
            mode_id,
            data.validated_by_name,
            data.validated_by_position,
            validated_by_id
        )
        if result:
            return result
        raise HTTPException(status_code=404, detail="Failure mode not found")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error validating failure mode: {e}")
        # Fallback
        try:
            legacy_id = int(mode_id)
            for fm in FAILURE_MODES_LIBRARY:
                if fm["id"] == legacy_id:
                    fm["is_validated"] = True
                    fm["validated_by_name"] = data.validated_by_name
                    fm["validated_by_position"] = data.validated_by_position
                    fm["validated_by_id"] = validated_by_id
                    fm["validated_at"] = datetime.now(timezone.utc).isoformat()
                    return fm
        except ValueError:
            pass
        raise HTTPException(status_code=404, detail="Failure mode not found")


@router.post("/failure-modes/{mode_id}/unvalidate")
async def unvalidate_failure_mode(
    mode_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Remove validation from a failure mode."""
    try:
        result = await failure_modes_service.unvalidate(mode_id)
        if result:
            return result
        raise HTTPException(status_code=404, detail="Failure mode not found")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error unvalidating failure mode: {e}")
        # Fallback
        try:
            legacy_id = int(mode_id)
            for fm in FAILURE_MODES_LIBRARY:
                if fm["id"] == legacy_id:
                    fm["is_validated"] = False
                    fm["validated_by_name"] = None
                    fm["validated_by_position"] = None
                    fm["validated_at"] = None
                    return fm
        except ValueError:
            pass
        raise HTTPException(status_code=404, detail="Failure mode not found")


@router.post("/failure-modes/merge")
async def merge_failure_modes(
    payload: dict = Body(...),
    current_user: dict = Depends(get_current_user),
):
    """Merge N "loser" failure modes into a single "winner" failure mode.

    Used by the AI similarity-review tool to consolidate near-duplicate
    failure modes (e.g. "Bearing Failure" + "Bearing Damage" on a pump).
    Always backs up loser docs to `fm_merge_log` before deleting.

    Body: {"winner_id": "...", "loser_ids": ["...", "..."], "canonical_name": "..."}
    """
    from bson import ObjectId

    winner_id = (payload.get("winner_id") or "").strip()
    loser_ids = [str(x).strip() for x in (payload.get("loser_ids") or []) if str(x).strip()]
    canonical_name = (payload.get("canonical_name") or "").strip()

    if not winner_id or not loser_ids:
        raise HTTPException(status_code=400, detail="winner_id and loser_ids are required")
    if winner_id in loser_ids:
        raise HTTPException(status_code=400, detail="winner_id cannot also be a loser")

    def _id_query(mid: str):
        if ObjectId.is_valid(mid):
            return {"_id": ObjectId(mid)}
        try:
            return {"legacy_id": int(mid)}
        except ValueError:
            return None

    win_q = _id_query(winner_id)
    if not win_q:
        raise HTTPException(status_code=400, detail=f"Invalid winner_id: {winner_id}")

    winner = await db.failure_modes.find_one(win_q)
    if not winner:
        raise HTTPException(status_code=404, detail="Winner failure mode not found")

    loser_docs = []
    for lid in loser_ids:
        q = _id_query(lid)
        if not q:
            continue
        doc = await db.failure_modes.find_one(q)
        if doc and doc["_id"] != winner["_id"]:
            loser_docs.append(doc)

    if not loser_docs:
        raise HTTPException(status_code=404, detail="No valid loser failure modes found")

    # Helper: dedupe-preserving union
    def _dedup(items, key=lambda x: x):
        seen, out = set(), []
        for it in items or []:
            k = key(it)
            if k in seen:
                continue
            seen.add(k)
            out.append(it)
        return out

    def _action_key(a):
        if isinstance(a, str):
            return a.strip().lower()
        if isinstance(a, dict):
            return (a.get("action") or a.get("description") or "").strip().lower()
        return str(a)

    def _str_key(v):
        return (v or "").strip().lower() if isinstance(v, str) else str(v)

    merged_ets = _dedup(
        list(winner.get("equipment_type_ids") or [])
        + [eid for ld in loser_docs for eid in (ld.get("equipment_type_ids") or [])]
    )
    merged_kw = _dedup(
        list(winner.get("keywords") or [])
        + [k for ld in loser_docs for k in (ld.get("keywords") or [])],
        key=_str_key,
    )
    merged_actions = _dedup(
        list(winner.get("recommended_actions") or [])
        + [a for ld in loser_docs for a in (ld.get("recommended_actions") or [])],
        key=_action_key,
    )
    merged_effects = _dedup(
        list(winner.get("potential_effects") or [])
        + [e for ld in loser_docs for e in (ld.get("potential_effects") or [])],
        key=_str_key,
    )
    merged_causes = _dedup(
        list(winner.get("potential_causes") or [])
        + [c for ld in loser_docs for c in (ld.get("potential_causes") or [])],
        key=_str_key,
    )

    update_fields = {
        "equipment_type_ids": merged_ets,
        "keywords": merged_kw,
        "recommended_actions": merged_actions,
        "potential_effects": merged_effects,
        "potential_causes": merged_causes,
        "updated_at": datetime.now(timezone.utc),
        "version": (winner.get("version") or 1) + 1,
    }
    if canonical_name:
        update_fields["failure_mode"] = canonical_name

    # Backup losers first — small audit collection so the merge can be undone manually.
    await db.fm_merge_log.insert_one({
        "merged_at": datetime.now(timezone.utc),
        "merged_by": current_user.get("id") or current_user.get("user_id"),
        "winner_id": str(winner["_id"]),
        "winner_failure_mode": canonical_name or winner.get("failure_mode"),
        "previous_winner_name": winner.get("failure_mode"),
        "losers": [
            {**{k: v for k, v in ld.items() if k != "_id"}, "_mongo_id": str(ld["_id"])}
            for ld in loser_docs
        ],
    })

    await db.failure_modes.update_one({"_id": winner["_id"]}, {"$set": update_fields})
    deleted = 0
    for ld in loser_docs:
        res = await db.failure_modes.delete_one({"_id": ld["_id"]})
        deleted += res.deleted_count

    # Invalidate FM list cache so the next GET refetches.
    try:
        from services.failure_modes_service import _invalidate_cache
        _invalidate_cache()
    except Exception:
        pass

    return {
        "winner_id": str(winner["_id"]),
        "deleted_count": deleted,
        "canonical_name": canonical_name or winner.get("failure_mode"),
    }


@router.delete("/failure-modes/{mode_id}")
async def delete_failure_mode(
    mode_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a custom failure mode from MongoDB."""
    try:
        deleted = await failure_modes_service.delete(mode_id)
        if deleted:
            return {"message": "Failure mode deleted"}
        raise HTTPException(status_code=404, detail="Failure mode not found")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting failure mode: {e}")
        # Fallback
        try:
            legacy_id = int(mode_id)
            for i, fm in enumerate(FAILURE_MODES_LIBRARY):
                if fm["id"] == legacy_id:
                    if not fm.get("is_custom"):
                        raise HTTPException(status_code=400, detail="Cannot delete built-in failure modes")
                    FAILURE_MODES_LIBRARY.pop(i)
                    return {"message": "Failure mode deleted"}
        except ValueError:
            pass
        raise HTTPException(status_code=404, detail="Failure mode not found")


# ============= ISO 14224 EQUIPMENT HIERARCHY ENDPOINTS =============

