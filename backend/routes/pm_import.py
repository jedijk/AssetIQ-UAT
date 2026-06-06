"""
PM Intelligence Import Routes - Upload and process maintenance plans.

Converts preventive maintenance plans (Excel, PDF, images) into structured
failure mode intelligence for the AssetIQ Failure Mode Library.
"""

import asyncio
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, BackgroundTasks
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import Optional, List, Dict, Any
from datetime import datetime, timezone
import uuid
import logging
import io
import json

from database import db
from auth import get_current_user, require_permission
from services.pm_import_service import PMImportService, normalize_pm_import_display_status

_library_write = require_permission("library:write")

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/pm-import", tags=["PM Intelligence Import"])


# ============== MODELS ==============

class TaskUpdateRequest(BaseModel):
    """Request model for updating a task (PM Import refactor shape)."""
    equipment_tag: Optional[str] = None
    equipment_description: Optional[str] = None
    task_description: Optional[str] = None
    task_type: Optional[str] = None  # PM | PDM | CBM | CM
    discipline: Optional[str] = None
    frequency: Optional[str] = None
    frequency_days: Optional[int] = None
    estimated_hours: Optional[float] = None


class ImportRequest(BaseModel):
    """Request model for final import."""
    include_low_confidence: bool = True


class BulkActionRequest(BaseModel):
    """Request model for bulk actions."""
    task_ids: Optional[List[str]] = None
    action: str  # "accept", "reject", "accept_high_confidence"


class SelectMatchRequest(BaseModel):
    """Request model for selecting a failure mode match."""
    match_id: str


class ApproveNewFMRequest(BaseModel):
    """Request model for approving a new failure mode creation."""
    failure_mode: str
    equipment: Optional[str] = None
    category: Optional[str] = None
    severity: Optional[int] = 5
    occurrence: Optional[int] = 5
    detectability: Optional[int] = 5


class ApplySuggestionRequest(BaseModel):
    """Request model for applying an AI suggestion."""
    action: str  # "merge", "new_failure_mode", "new_task", "keep_custom", "reject"
    target_failure_mode_id: Optional[str] = None
    new_failure_mode_data: Optional[Dict[str, Any]] = None
    # 0-based index of an existing recommended_action to replace. When omitted
    # the backend falls back to the AI's stored choice on the suggestion, then to
    # a lexical similarity check.
    replace_action_index: Optional[int] = None


class ApplyToFailureModeRequest(BaseModel):
    """Apply a Custom PM Import task to a failure mode from the import list."""
    target_failure_mode_id: str
    placement_mode: str = "add"  # "add" | "replace"
    replace_action_index: Optional[int] = None


# Background task for processing
async def process_pm_file_background(
    session_id: str,
    file_name: str,
    file_type: str,
    file_content: bytes,
    created_by: str
):
    """Process PM file in background."""
    pm_service = PMImportService(db)
    try:
        await pm_service.process_session(
            session_id=session_id,
            file_name=file_name,
            file_type=file_type,
            file_content=file_content
        )
    except Exception as e:
        logger.error(f"Background processing error for session {session_id}: {e}", exc_info=True)
        # Update session with error
        await db.pm_import_sessions.update_one(
            {"session_id": session_id},
            {"$set": {
                "status": "error",
                "error_message": str(e),
                "progress_message": f"Error: {str(e)[:200]}",
                "updated_at": datetime.now(timezone.utc)
            }}
        )


# ============== ROUTES ==============

@router.get("/template")
async def download_pm_template():
    """
    Download an empty PM Import Excel template.
    
    The template includes:
    - Column headers: Tag, Task, Frequency, Discipline
    - Example rows (in gray italic)
    - Instructions sheet with valid discipline values
    """
    import os
    from fastapi.responses import FileResponse
    
    template_path = "/app/backend/static/templates/pm_import_template.xlsx"
    
    if not os.path.exists(template_path):
        raise HTTPException(status_code=404, detail="Template file not found")
    
    return FileResponse(
        path=template_path,
        filename="pm_import_template.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.post("/upload")
async def upload_pm_plan(
    file: UploadFile = File(...),
    background_tasks: BackgroundTasks = None,
    current_user: dict = Depends(_library_write),
):
    """
    Upload a preventive maintenance plan for processing.
    
    Supported file types:
    - Excel (.xlsx, .xls)
    - PDF
    - Images (.png, .jpg, .jpeg, .webp)
    
    Returns a session ID to track processing progress.
    """
    
    # Validate file type
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")
    
    # Get file extension
    filename_lower = file.filename.lower()
    
    if filename_lower.endswith(('.xlsx', '.xls')):
        file_type = 'xlsx' if filename_lower.endswith('.xlsx') else 'xls'
    elif filename_lower.endswith('.pdf'):
        file_type = 'pdf'
    elif filename_lower.endswith('.png'):
        file_type = 'png'
    elif filename_lower.endswith(('.jpg', '.jpeg')):
        file_type = 'jpg'
    elif filename_lower.endswith('.webp'):
        file_type = 'webp'
    else:
        raise HTTPException(
            status_code=400, 
            detail="Unsupported file type. Supported: Excel (.xlsx, .xls), PDF, Images (.png, .jpg, .jpeg, .webp)"
        )
    
    # Read file content
    content = await file.read()
    
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="Empty file uploaded")
    
    # Max file size: 20MB
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large. Maximum size is 20MB")
    
    # Create PM Import service
    pm_service = PMImportService(db)
    user_id = current_user.get("id", current_user.get("email", "unknown"))
    
    # Create session first (quick - just DB insert)
    try:
        session_id = await pm_service.create_session_placeholder(
            file_name=file.filename,
            file_type=file_type,
            created_by=user_id
        )
        
        # Process file in background to avoid timeout
        asyncio.create_task(process_pm_file_background(
            session_id=session_id,
            file_name=file.filename,
            file_type=file_type,
            file_content=content,
            created_by=user_id
        ))
        
        return {
            "session_id": session_id,
            "status": "processing",
            "message": "File uploaded. Processing started in background. Poll session for status.",
            "stats": None,
            "tasks_count": 0
        }
        
    except Exception as e:
        logger.error(f"Upload error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/session/{session_id}")
async def get_session(
    session_id: str,
    current_user: dict = Depends(_library_write)
):
    """
    Get the status and results of a PM import session.
    
    Returns:
    - Session status (processing, ready_for_review, imported, error)
    - Progress percentage
    - Extracted tasks with AI analysis
    - Statistics
    """
    
    pm_service = PMImportService(db)
    session = await pm_service.ensure_equipment_impacts(session_id, current_user)
    
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Remove internal fields
    session.pop("_id", None)
    
    return session


@router.patch("/session/{session_id}/task/{task_id}")
async def update_task(
    session_id: str,
    task_id: str,
    updates: TaskUpdateRequest,
    current_user: dict = Depends(_library_write)
):
    """
    Update a specific task in the review session.
    
    Allows editing:
    - Component name
    - Task type
    - Suggested failure modes
    - Failure mechanisms
    - Detection methods
    - Existing control description
    - Frequency
    """
    
    pm_service = PMImportService(db)
    
    update_data = updates.model_dump(exclude_none=True)
    if not update_data:
        raise HTTPException(status_code=400, detail="No updates provided")
    
    result = await pm_service.update_task(session_id, task_id, update_data)
    
    if not result:
        raise HTTPException(status_code=404, detail="Session or task not found")
    
    return {
        "success": True,
        "stats": result["stats"]
    }


@router.post("/session/{session_id}/task/{task_id}/accept")
async def accept_task(
    session_id: str,
    task_id: str,
    current_user: dict = Depends(_library_write)
):
    """Mark a task as accepted for import."""
    
    pm_service = PMImportService(db)
    result = await pm_service.accept_task(session_id, task_id)
    
    if not result:
        raise HTTPException(status_code=404, detail="Session or task not found")
    
    return {"success": True, "stats": result["stats"]}


@router.post("/session/{session_id}/task/{task_id}/reject")
async def reject_task(
    session_id: str,
    task_id: str,
    current_user: dict = Depends(_library_write)
):
    """Mark a task as rejected (won't be imported)."""
    
    pm_service = PMImportService(db)
    result = await pm_service.reject_task(session_id, task_id)
    
    if not result:
        raise HTTPException(status_code=404, detail="Session or task not found")
    
    return {"success": True, "stats": result["stats"]}


@router.delete("/session/{session_id}/task/{task_id}")
async def delete_task(
    session_id: str,
    task_id: str,
    current_user: dict = Depends(_library_write)
):
    """Permanently remove a task from a session."""
    
    pm_service = PMImportService(db)
    result = await pm_service.delete_task(session_id, task_id)
    
    if not result:
        raise HTTPException(status_code=404, detail="Session or task not found")
    
    return {"success": True, "stats": result["stats"]}


class TaskMappingRequest(BaseModel):
    """Request model for manually mapping a task to a hierarchy equipment node."""
    equipment_id: Optional[str] = None


@router.patch("/session/{session_id}/task/{task_id}/mapping")
async def update_task_mapping(
    session_id: str,
    task_id: str,
    body: TaskMappingRequest,
    current_user: dict = Depends(_library_write)
):
    """
    Manually map an imported task to an equipment hierarchy node.
    PM Import refactor: no longer touches failure modes or equipment types.
    """
    
    session = await db.pm_import_sessions.find_one({"session_id": session_id})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    tasks = session.get("tasks_extracted", [])
    matched = False
    
    for task in tasks:
        if task.get("task_id") != task_id:
            continue
        matched = True
        
        if body.equipment_id is not None:
            node = await db.equipment_nodes.find_one({"id": body.equipment_id})
            if not node:
                raise HTTPException(status_code=400, detail="Equipment node not found")
            pm_service = PMImportService(db)
            type_map = await pm_service._load_equipment_type_name_map()
            task["equipment_match"] = pm_service._apply_equipment_type_to_match(
                {
                    "equipment_id": node.get("id"),
                    "tag": node.get("tag"),
                    "name": node.get("name"),
                    "match_type": "manual",
                    "confidence": 100,
                },
                node,
                type_map,
            )
            # Also propagate tag/description so the row reads cleanly
            if node.get("tag"):
                task["equipment_tag"] = node.get("tag")
            if node.get("name"):
                task["equipment_description"] = node.get("name")
        
        task["updated_at"] = datetime.now(timezone.utc).isoformat()
        break
    
    if not matched:
        raise HTTPException(status_code=404, detail="Task not found")
    
    await db.pm_import_sessions.update_one(
        {"session_id": session_id},
        {"$set": {
            "tasks_extracted": tasks,
            "updated_at": datetime.now(timezone.utc),
        }}
    )
    
    return {"success": True}


@router.get("/lookup/equipment")
async def lookup_equipment(
    q: Optional[str] = None,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Search equipment hierarchy nodes by tag or name (for manual mapping dropdown)."""
    query = {}
    if q:
        query = {"$or": [
            {"tag": {"$regex": q, "$options": "i"}},
            {"name": {"$regex": q, "$options": "i"}},
        ]}
    cursor = db.equipment_nodes.find(query).limit(limit)
    items = []
    async for n in cursor:
        items.append({
            "id": n.get("id"),
            "tag": n.get("tag"),
            "name": n.get("name"),
            "level": n.get("level"),
        })
    return {"items": items}


@router.post("/session/{session_id}/import-to-pm-tasks")
async def import_session_to_pm_tasks(
    session_id: str,
    current_user: dict = Depends(_library_write)
):
    """
    Promote all `accepted` tasks from a session into the `pm_tasks` collection.
    PM Import refactor: writes to `pm_tasks` only — NOT to maintenance_programs,
    failure_modes, or fmea_library.
    """
    session = await db.pm_import_sessions.find_one({"session_id": session_id})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    accepted = [
        t for t in (session.get("tasks_extracted") or [])
        if t.get("review_status") == "accepted"
    ]
    
    if not accepted:
        return {"success": True, "imported": 0, "message": "No accepted tasks"}
    
    now = datetime.now(timezone.utc).isoformat()
    docs = []
    for t in accepted:
        em = t.get("equipment_match") or {}
        docs.append({
            "task_id": str(uuid.uuid4()),
            "source_task_id": t.get("task_id"),
            "equipment_id": em.get("equipment_id"),
            "equipment_tag": t.get("equipment_tag") or em.get("tag") or "",
            "equipment_description": t.get("equipment_description") or em.get("name") or "",
            "task_description": t.get("task_description") or "",
            "task_type": t.get("task_type") or "PM",
            "discipline": t.get("discipline") or "Mechanical",
            "frequency": t.get("frequency") or "Monthly",
            "frequency_days": t.get("frequency_days"),
            "estimated_hours": t.get("estimated_hours") or 0.5,
            "source_import_session": session_id,
            "created_at": now,
            "created_by": current_user.get("id") or current_user.get("email"),
        })
    
    if docs:
        await db.pm_tasks.insert_many(docs)
    
    # Mark the session as imported (without deleting tasks_extracted, for traceability)
    await db.pm_import_sessions.update_one(
        {"session_id": session_id},
        {"$set": {"status": "imported", "updated_at": datetime.now(timezone.utc)}}
    )
    
    return {"success": True, "imported": len(docs)}


@router.post("/session/{session_id}/task/{task_id}/select-match")
async def select_match(
    session_id: str,
    task_id: str,
    request: SelectMatchRequest,
    current_user: dict = Depends(_library_write)
):
    """
    Select a failure mode match for a task (Scenario B).
    
    When a task has multiple possible matches, user selects one.
    The task will be linked to the selected failure mode during import.
    """
    
    pm_service = PMImportService(db)
    
    result = await pm_service.update_task(
        session_id, 
        task_id, 
        {
            "selected_match_id": request.match_id,
            "approved_new_fm": None,  # Clear prior new-FM approval since user is linking existing
            "review_status": "accepted"
        }
    )
    
    if not result:
        raise HTTPException(status_code=404, detail="Session or task not found")
    
    return {"success": True, "stats": result["stats"]}


@router.post("/session/{session_id}/task/{task_id}/approve-new-fm")
async def approve_new_failure_mode(
    session_id: str,
    task_id: str,
    request: ApproveNewFMRequest,
    current_user: dict = Depends(_library_write)
):
    """
    Approve creation of a new failure mode for a task (Scenario C).
    
    When no reliable match is found, AI proposes a new failure mode.
    User reviews and approves the details before it gets created.
    """
    
    pm_service = PMImportService(db)
    
    result = await pm_service.update_task(
        session_id, 
        task_id, 
        {
            "approved_new_fm": {
                "failure_mode": request.failure_mode,
                "equipment": request.equipment,
                "category": request.category,
                "severity": request.severity,
                "occurrence": request.occurrence,
                "detectability": request.detectability
            },
            "selected_match_id": None,  # Clear prior existing-match selection since user is creating new
            "review_status": "accepted"
        }
    )
    
    if not result:
        raise HTTPException(status_code=404, detail="Session or task not found")
    
    return {"success": True, "stats": result["stats"]}


@router.post("/session/{session_id}/bulk-action")
async def bulk_action(
    session_id: str,
    request: BulkActionRequest,
    current_user: dict = Depends(_library_write)
):
    """
    Perform bulk actions on tasks.
    
    Actions:
    - accept: Accept specified tasks
    - reject: Reject specified tasks
    - accept_high_confidence: Accept all tasks with confidence >= 70%
    """
    
    pm_service = PMImportService(db)
    
    if request.action == "accept_high_confidence":
        result = await pm_service.accept_all_high_confidence(session_id)
        if not result:
            raise HTTPException(status_code=404, detail="Session not found")
        return {
            "success": True,
            "accepted_count": result["accepted_count"],
            "stats": result["stats"]
        }
    
    elif request.action in ["accept", "reject"]:
        if not request.task_ids:
            raise HTTPException(status_code=400, detail="task_ids required for accept/reject")
        
        for task_id in request.task_ids:
            if request.action == "accept":
                await pm_service.accept_task(session_id, task_id)
            else:
                await pm_service.reject_task(session_id, task_id)
        
        session = await pm_service.get_session(session_id)
        return {
            "success": True,
            "processed_count": len(request.task_ids),
            "stats": session.get("stats") if session else {}
        }
    
    else:
        raise HTTPException(status_code=400, detail=f"Unknown action: {request.action}")


@router.post("/session/{session_id}/import")
async def import_to_library(
    session_id: str,
    request: ImportRequest = ImportRequest(),
    current_user: dict = Depends(_library_write)
):
    """
    Import accepted tasks to the Failure Mode Library.
    
    - Tasks with existing matches: Links as preventive control
    - New proposed tasks: Creates new failure mode entries
    - Low confidence items: Imported with warning if include_low_confidence=True
    """
    
    pm_service = PMImportService(db)
    
    try:
        result = await pm_service.import_to_library(
            session_id=session_id,
            created_by=current_user.get("id", current_user.get("email", "unknown"))
        )
        
        return result
        
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.error(f"Import error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/session/{session_id}/export")
async def export_review(
    session_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Export the review results as an Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    pm_service = PMImportService(db)
    session = await pm_service.get_session(session_id)
    
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    tasks = session.get("tasks_extracted", [])
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "PM Import Review"
    
    # Headers
    headers = [
        "Original Task", "Component", "Task Type", "Suggested Failure Modes",
        "Failure Mechanisms", "Detection Methods", "Existing Control", 
        "Frequency", "Confidence", "Library Match", "Review Status"
    ]
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
    
    # Write data
    for row_idx, task in enumerate(tasks, 2):
        failure_modes = task.get("suggested_failure_modes", [])
        fm_str = ", ".join(fm if isinstance(fm, str) else fm.get("name", "") for fm in failure_modes)
        
        mechanisms = task.get("failure_mechanisms", [])
        mech_str = ", ".join(mechanisms) if isinstance(mechanisms, list) else str(mechanisms)
        
        detection = task.get("detection_methods", [])
        det_str = ", ".join(detection) if isinstance(detection, list) else str(detection)
        
        library_match = task.get("library_match", {})
        match_str = library_match.get("status", "unknown")
        if library_match.get("matched_name"):
            match_str += f" ({library_match['matched_name']})"
        
        row_data = [
            task.get("original_task", ""),
            task.get("component", ""),
            task.get("task_type", ""),
            fm_str,
            mech_str,
            det_str,
            task.get("existing_control", ""),
            task.get("frequency", ""),
            task.get("confidence_score", 0),
            match_str,
            task.get("review_status", "pending")
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Color code confidence
            if col == 9:  # Confidence column
                if value >= 90:
                    cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                elif value >= 70:
                    cell.fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    # Adjust column widths
    column_widths = [40, 20, 15, 35, 30, 25, 35, 15, 12, 25, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze header
    ws.freeze_panes = "A2"
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"pm_import_review_{session_id[:8]}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.delete("/session/{session_id}")
async def delete_session(
    session_id: str,
    current_user: dict = Depends(_library_write)
):
    """Delete a PM import session."""
    
    result = await db.pm_import_sessions.delete_one({"session_id": session_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Session not found")
    
    return {"success": True, "message": "Session deleted"}


@router.get("/sessions")
async def list_sessions(
    limit: int = 20,
    skip: int = 0,
    current_user: dict = Depends(get_current_user)
):
    """List PM import sessions for the current user."""
    
    user_id = current_user.get("id", current_user.get("email", "unknown"))
    
    cursor = db.pm_import_sessions.find(
        {"created_by": user_id}
    ).sort("created_at", -1).skip(skip).limit(limit)
    
    sessions = []
    async for session in cursor:
        sessions.append({
            "session_id": session["session_id"],
            "file_name": session["file_name"],
            "status": session["status"],
            "stats": session.get("stats", {}),
            "created_at": session["created_at"].isoformat() if session.get("created_at") else None
        })
    
    total = await db.pm_import_sessions.count_documents({"created_by": user_id})
    
    return {
        "sessions": sessions,
        "total": total
    }


@router.delete("/tasks")
async def delete_all_tasks(
    current_user: dict = Depends(_library_write)
):
    """
    Delete ALL imported tasks across ALL sessions for the current user.
    Sessions are preserved (file metadata kept), but `tasks_extracted` is cleared.
    """
    
    user_id = current_user.get("id", current_user.get("email", "unknown"))
    
    # Count tasks before
    cursor = db.pm_import_sessions.find({"created_by": user_id})
    deleted_count = 0
    session_count = 0
    async for session in cursor:
        session_count += 1
        deleted_count += len(session.get("tasks_extracted", []) or [])
    
    # Clear all tasks across all sessions
    cleared_stats = {
        "total_tasks": 0, "failure_modes_identified": 0,
        "existing_matches": 0, "new_proposed": 0,
        "low_confidence_items": 0, "manual_review_required": 0,
        "accepted": 0, "rejected": 0, "pending": 0,
    }
    result = await db.pm_import_sessions.update_many(
        {"created_by": user_id},
        {"$set": {
            "tasks_extracted": [],
            "stats": cleared_stats,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {
        "success": True,
        "deleted_tasks": deleted_count,
        "sessions_cleared": result.modified_count
    }


@router.get("/tasks")
async def list_all_tasks(
    current_user: dict = Depends(get_current_user)
):
    """
    List all extracted tasks across all PM import sessions for the current user.
    
    Returns a flattened view of tasks with the fields needed for the
    Custom PM Import tab: equipment, task, task type, discipline, frequency.
    """
    
    user_id = current_user.get("id", current_user.get("email", "unknown"))
    pm_service = PMImportService(db)
    
    cursor = db.pm_import_sessions.find(
        {"created_by": user_id}
    ).sort("created_at", -1)
    
    tasks = []
    session_count = 0
    async for session in cursor:
        session_count += 1
        session_id = session.get("session_id")
        file_name = session.get("file_name")
        created_at = session.get("created_at")
        created_at_iso = created_at.isoformat() if created_at else None
        
        for task in session.get("tasks_extracted", []) or []:
            equipment_match = task.get("equipment_match")
            tasks.append({
                "task_id": task.get("task_id"),
                "session_id": session_id,
                "file_name": file_name,
                "imported_at": created_at_iso,
                # ===== PM Import refactor shape =====
                "equipment_tag": task.get("equipment_tag") or task.get("asset") or "",
                "equipment_description": task.get("equipment_description") or task.get("component") or "",
                "equipment_type_name": (equipment_match or {}).get("equipment_type_name") or "",
                "task_description": task.get("task_description") or task.get("original_task") or "",
                "task_type": task.get("task_type") or "PM",
                "discipline": task.get("discipline") or "Mechanical",
                "frequency": task.get("frequency") or "Monthly",
                "frequency_days": task.get("frequency_days"),
                "estimated_hours": task.get("estimated_hours") or 0.5,
                "confidence_score": task.get("confidence_score") or 50,
                "equipment_match": equipment_match,
                "review_status": task.get("review_status") or "pending",
                "import_status": normalize_pm_import_display_status(task),
                "apply_mode": task.get("apply_mode"),
                "target_failure_mode_id": task.get("target_failure_mode_id"),
                "original_task": task.get("original_task") or "",
            })

    await pm_service.enrich_task_rows_equipment_types(tasks)
    
    return {
        "tasks": tasks,
        "total": len(tasks),
        "session_count": session_count
    }


# ============== AI REVIEW ROUTES ==============

@router.post("/session/{session_id}/ai-review")
async def ai_review_session(
    session_id: str,
    current_user: dict = Depends(_library_write)
):
    """
    Run AI review on all accepted tasks in a session.
    
    For each accepted task:
    1. Matches equipment type by tag → equipment_nodes → equipment_type_id
    2. Finds similar failure modes in library
    3. Generates AI recommendation (merge, new_failure_mode, new_task, keep_custom)
    
    Returns suggestions for each accepted task.
    """
    pm_service = PMImportService(db)
    
    try:
        result = await pm_service.ai_review_accepted_tasks(session_id)
        return result
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.error(f"AI review error for session {session_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"AI review failed: {str(e)}")


@router.get("/session/{session_id}/ai-review")
async def get_ai_review_results(
    session_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get the AI review results for a session."""
    session = await db.pm_import_sessions.find_one(
        {"session_id": session_id},
        {"_id": 0, "ai_review_suggestions": 1, "ai_review_status": 1}
    )
    
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    return {
        "suggestions": session.get("ai_review_suggestions", []),
        "status": session.get("ai_review_status", "not_started")
    }


@router.post("/session/{session_id}/task/{task_id}/apply-to-failure-mode")
async def apply_task_to_failure_mode(
    session_id: str,
    task_id: str,
    request: ApplyToFailureModeRequest,
    current_user: dict = Depends(_library_write),
):
    """
    Apply an accepted PM Import task to a failure mode.
    Use placement_mode 'replace' with replace_action_index, or 'add' for a new action.
    Marks the task as implemented in the session.
    """
    pm_service = PMImportService(db)
    user_id = current_user.get("id") or current_user.get("user_id") or current_user.get("email")

    try:
        return await pm_service.apply_task_to_failure_mode(
            session_id=session_id,
            task_id=task_id,
            target_failure_mode_id=request.target_failure_mode_id,
            placement_mode=request.placement_mode,
            replace_action_index=request.replace_action_index,
            user_id=user_id,
        )
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.error(f"Apply to failure mode error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to apply task: {str(e)}")


@router.post("/session/{session_id}/task/{task_id}/apply-suggestion")
async def apply_ai_suggestion(
    session_id: str,
    task_id: str,
    request: ApplySuggestionRequest,
    current_user: dict = Depends(_library_write)
):
    """
    Apply an AI suggestion for a specific task.
    
    Actions:
    - "merge": Merge task into existing failure mode's recommended_actions
    - "new_failure_mode": Create new failure mode with this task
    - "new_task": Add task under existing failure mode
    - "keep_custom": Mark as custom (stays in session)
    - "reject": Reject suggestion (stays as accepted task)
    """
    pm_service = PMImportService(db)
    
    try:
        result = await pm_service.apply_ai_suggestion(
            session_id=session_id,
            task_id=task_id,
            action=request.action,
            target_failure_mode_id=request.target_failure_mode_id,
            new_failure_mode_data=request.new_failure_mode_data,
            replace_action_index=request.replace_action_index,
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.error(f"Apply suggestion error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to apply suggestion: {str(e)}")


@router.post("/session/{session_id}/apply-all-suggestions")
async def apply_all_suggestions(
    session_id: str,
    current_user: dict = Depends(_library_write)
):
    """Apply all AI suggestions that have action != 'keep_custom'."""
    pm_service = PMImportService(db)
    
    session = await db.pm_import_sessions.find_one({"session_id": session_id})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    suggestions = session.get("ai_review_suggestions", [])
    results = {"applied": 0, "failed": 0, "skipped": 0, "replaced": 0, "added": 0, "details": []}
    
    for suggestion in suggestions:
        if suggestion.get("status") == "applied":
            results["skipped"] += 1
            continue
        
        recommendation = suggestion.get("recommendation", {})
        action = recommendation.get("action", "keep_custom")
        
        if action == "keep_custom":
            results["skipped"] += 1
            continue
        
        try:
            result = await pm_service.apply_ai_suggestion(
                session_id=session_id,
                task_id=suggestion.get("task_id"),
                action=action,
                target_failure_mode_id=recommendation.get("target_failure_mode_id"),
                replace_action_index=recommendation.get("replace_action_index"),
            )
            
            if result.get("success"):
                results["applied"] += 1
                if result.get("mode") == "replaced":
                    results["replaced"] += 1
                elif result.get("mode") == "added":
                    results["added"] += 1
            else:
                results["failed"] += 1
            
            results["details"].append(result)
        except Exception as e:
            results["failed"] += 1
            results["details"].append({
                "task_id": suggestion.get("task_id"),
                "success": False,
                "error": str(e)
            })
    
    return results
