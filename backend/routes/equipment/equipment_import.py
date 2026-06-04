"""
Equipment Hierarchy Import operations.
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from pydantic import BaseModel
from datetime import datetime, timezone
import uuid
import io
import logging
import pandas as pd
from database import db
from auth import get_current_user
from iso14224_models import (
    ISOLevel, ISO_LEVEL_LABELS, LEGACY_LEVEL_MAP, detect_equipment_type,
    UnstructuredItemCreate, ParseEquipmentListRequest, AssignToHierarchyRequest,
    is_valid_parent_child, get_valid_child_levels,
)
from services.criticality_score import compute_criticality_score

logger = logging.getLogger(__name__)
router = APIRouter()


# =============================================================================
# UNSTRUCTURED ITEMS
# =============================================================================

@router.get("/equipment-hierarchy/unstructured")
async def get_unstructured_items(
    current_user: dict = Depends(get_current_user)
):
    """Get all unstructured (unassigned) equipment items."""
    items = await db.unstructured_items.find(
        {"created_by": current_user["id"]},
        {"_id": 0}
    ).to_list(1000)
    return {"items": items}


@router.post("/equipment-hierarchy/unstructured")
async def create_unstructured_item(
    item_data: UnstructuredItemCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a single unstructured equipment item."""
    detected = detect_equipment_type(item_data.name)
    
    item_id = str(uuid.uuid4())
    item_doc = {
        "id": item_id,
        "name": item_data.name,
        "detected_type_id": item_data.detected_type_id or (detected["id"] if detected else None),
        "detected_type_name": detected["name"] if detected else None,
        "detected_discipline": item_data.detected_discipline or (detected["discipline"] if detected else None),
        "detected_icon": detected["icon"] if detected else None,
        "source": item_data.source or "manual",
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.unstructured_items.insert_one(item_doc)
    item_doc.pop("_id", None)
    return item_doc


@router.post("/equipment-hierarchy/parse-list")
async def parse_equipment_list(
    request: ParseEquipmentListRequest,
    current_user: dict = Depends(get_current_user)
):
    """Parse a text list and create unstructured items with auto-detection."""
    import re
    
    content = request.content.strip()
    items = re.split(r'[\n\r,;\t]+', content)
    
    seen = set()
    unique_items = []
    for item in items:
        cleaned = item.strip()
        cleaned = re.sub(r'^[\d]+[.\)]\s*', '', cleaned)
        cleaned = re.sub(r'^[-•*]\s*', '', cleaned)
        cleaned = cleaned.strip()
        
        if cleaned and len(cleaned) > 1 and cleaned.lower() not in seen:
            seen.add(cleaned.lower())
            unique_items.append(cleaned)
    
    created_items = []
    for name in unique_items:
        detected = detect_equipment_type(name)
        
        item_id = str(uuid.uuid4())
        item_doc = {
            "id": item_id,
            "name": name,
            "detected_type_id": detected["id"] if detected else None,
            "detected_type_name": detected["name"] if detected else None,
            "detected_discipline": detected["discipline"] if detected else None,
            "detected_icon": detected["icon"] if detected else None,
            "source": request.source,
            "created_by": current_user["id"],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.unstructured_items.insert_one(item_doc)
        item_doc.pop("_id", None)
        created_items.append(item_doc)
    
    return {"parsed_count": len(created_items), "items": created_items}


@router.post("/equipment-hierarchy/parse-file")
async def parse_equipment_file(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Parse an uploaded file (Excel, PDF, CSV, TXT) and extract equipment items."""
    import re
    
    filename = file.filename.lower()
    content = await file.read()
    
    extracted_items = []
    
    try:
        if filename.endswith('.csv') or filename.endswith('.txt'):
            text_content = content.decode('utf-8', errors='ignore')
            items = text_content.strip().split('\n')
            extracted_items = [item.strip() for item in items if item.strip()]
            
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            import pandas as pd
            df = pd.read_excel(io.BytesIO(content), header=None)
            for col in df.columns:
                for val in df[col].dropna():
                    if isinstance(val, str) and val.strip():
                        extracted_items.append(val.strip())
                    elif not pd.isna(val):
                        extracted_items.append(str(val).strip())
                        
        elif filename.endswith('.pdf'):
            try:
                import pdfplumber
                with pdfplumber.open(io.BytesIO(content)) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text()
                        if text:
                            lines = text.split('\n')
                            extracted_items.extend([ln.strip() for ln in lines if ln.strip()])
            except ImportError:
                import PyPDF2
                reader = PyPDF2.PdfReader(io.BytesIO(content))
                for page in reader.pages:
                    text = page.extract_text()
                    if text:
                        lines = text.split('\n')
                        extracted_items.extend([ln.strip() for ln in lines if ln.strip()])
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type")
    
    except Exception as e:
        logger.error(f"File parsing error: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")
    
    seen = set()
    unique_items = []
    for item in extracted_items:
        cleaned = item.strip()
        cleaned = re.sub(r'^[\d]+[.\)]\s*', '', cleaned)
        cleaned = re.sub(r'^[-•*]\s*', '', cleaned)
        cleaned = cleaned.strip()
        
        if cleaned and len(cleaned) > 1 and cleaned.lower() not in seen:
            seen.add(cleaned.lower())
            unique_items.append(cleaned)
    
    created_items = []
    for name in unique_items[:100]:
        detected = detect_equipment_type(name)
        
        item_id = str(uuid.uuid4())
        item_doc = {
            "id": item_id,
            "name": name,
            "detected_type_id": detected["id"] if detected else None,
            "detected_type_name": detected["name"] if detected else None,
            "detected_discipline": detected["discipline"] if detected else None,
            "detected_icon": detected["icon"] if detected else None,
            "source": "file",
            "created_by": current_user["id"],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.unstructured_items.insert_one(item_doc)
        item_doc.pop("_id", None)
        created_items.append(item_doc)
    
    return {"filename": file.filename, "parsed_count": len(created_items), "items": created_items}


@router.post("/equipment-hierarchy/unstructured/{item_id}/assign")
async def assign_unstructured_to_hierarchy(
    item_id: str,
    assignment: AssignToHierarchyRequest,
    current_user: dict = Depends(get_current_user)
):
    """Move an unstructured item into the ISO hierarchy."""
    item = await db.unstructured_items.find_one(
        {"id": item_id, "created_by": current_user["id"]}
    )
    if not item:
        raise HTTPException(status_code=404, detail="Unstructured item not found")
    
    existing = await db.equipment_nodes.find_one({
        "name": item["name"],
        "parent_id": assignment.parent_id,
        "created_by": current_user["id"]
    })
    if existing:
        raise HTTPException(
            status_code=400, 
            detail=f"A node with name '{item['name']}' already exists under this parent"
        )
    
    parent = await db.equipment_nodes.find_one(
        {"id": assignment.parent_id},
        {"_id": 0}
    )
    if not parent:
        raise HTTPException(status_code=400, detail="Parent node not found")
    
    try:
        target_level = ISOLevel(assignment.level)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid ISO level: {assignment.level}")
    
    parent_level = ISOLevel(parent["level"])
    if not is_valid_parent_child(parent_level, target_level):
        valid_children = get_valid_child_levels(parent_level)
        raise HTTPException(
            status_code=400,
            detail=f"Cannot add {target_level.value} under {parent_level.value}. Valid: {[c.value for c in valid_children]}"
        )
    
    node_id = str(uuid.uuid4())
    node_doc = {
        "id": node_id,
        "name": item["name"],
        "level": target_level.value,
        "parent_id": assignment.parent_id,
        "equipment_type_id": item.get("detected_type_id"),
        "description": f"Imported from unstructured list (source: {item.get('source', 'unknown')})",
        "criticality": None,
        "discipline": item.get("detected_discipline"),
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.equipment_nodes.insert_one(node_doc)
    node_doc.pop("_id", None)
    
    await db.unstructured_items.delete_one({"id": item_id})
    
    return node_doc


@router.delete("/equipment-hierarchy/unstructured/{item_id}")
async def delete_unstructured_item(
    item_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete an unstructured item."""
    result = await db.unstructured_items.delete_one(
        {"id": item_id, "created_by": current_user["id"]}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Item deleted"}


@router.delete("/equipment-hierarchy/unstructured")
async def clear_unstructured_items(
    current_user: dict = Depends(get_current_user)
):
    """Delete all unstructured items for the current user."""
    result = await db.unstructured_items.delete_many(
        {"created_by": current_user["id"]}
    )
    return {"message": f"Deleted {result.deleted_count} items"}


# =============================================================================
# EXCEL HIERARCHY IMPORT
# =============================================================================

EXCEL_LEVEL_MAPPING = {
    "Plant/Unit": "plant_unit",
    "Section/System": "section_system",
    "Equipment Unit": "equipment_unit",
    "Subunit": "subunit",
    "Maintainable Item": "maintainable_item",
    "Installation": "installation",
}


def calculate_criticality_from_excel(safety: int, production: int, environmental: int, reputation: int):
    """Calculate criticality data including level, color, and risk score."""
    max_impact = max(safety, production, environmental, reputation)
    
    if safety >= 4 or max_impact == 5:
        level = "safety_critical"
        color = "#EF4444"
    elif production >= 4 or max_impact >= 4:
        level = "production_critical"
        color = "#F97316"
    elif max_impact >= 3:
        level = "medium"
        color = "#EAB308"
    else:
        level = "low"
        color = "#22C55E"
    
    return {
        "safety_impact": safety,
        "production_impact": production,
        "environmental_impact": environmental,
        "reputation_impact": reputation,
        "level": level,
        "color": color,
        "max_impact": max_impact,
        "risk_score": compute_criticality_score(
            safety, production, environmental, reputation
        ),
    }


class ExcelHierarchyImportRequest(BaseModel):
    installation_id: str
    excel_url: str
    replace_existing: bool = True


# ISO 14224 Level mapping for file imports (display labels)
ISO_LEVEL_MAPPING = dict(EXCEL_LEVEL_MAPPING)

# Level hierarchy order (installation is the selected target, not imported as a child row)
ISO_LEVEL_ORDER = ["plant_unit", "section_system", "equipment_unit", "subunit", "maintainable_item"]

# Accept many spreadsheet variants for the Level column
_LEVEL_INPUT_MAP: dict[str, str] = {}


def _build_level_input_map() -> dict[str, str]:
    if _LEVEL_INPUT_MAP:
        return _LEVEL_INPUT_MAP
    mapping: dict[str, str] = {}
    for label, iso in ISO_LEVEL_MAPPING.items():
        mapping[label.lower()] = iso
        mapping[label.replace("/", " ").lower()] = iso
        mapping[label.replace("/", "_").lower()] = iso
    for enum_member in ISOLevel:
        val = enum_member.value
        if val == "installation":
            continue
        mapping[val.lower()] = val
        label = ISO_LEVEL_LABELS.get(enum_member)
        if label:
            mapping[label.lower()] = val
            mapping[label.replace("/", " ").lower()] = val
            mapping[label.replace("/", "_").lower()] = val
    for legacy, iso in LEGACY_LEVEL_MAP.items():
        mapping[legacy.lower()] = iso
    # Common spreadsheet aliases
    mapping.update({
        "plant": "plant_unit",
        "unit": "plant_unit",
        "section": "section_system",
        "system": "section_system",
        "equipment": "equipment_unit",
        "maintainable item": "maintainable_item",
        "maintainable": "maintainable_item",
        "equipment unit": "equipment_unit",
    })
    _LEVEL_INPUT_MAP.update(mapping)
    return _LEVEL_INPUT_MAP


def resolve_iso_level(level_raw) -> str | None:
    """Map a spreadsheet Level cell to canonical ISO level key."""
    if level_raw is None or (isinstance(level_raw, float) and pd.isna(level_raw)):
        return None
    level_str = str(level_raw).strip()
    if not level_str:
        return None
    if level_str in ISO_LEVEL_MAPPING:
        return ISO_LEVEL_MAPPING[level_str]
    key = level_str.lower()
    resolved = _build_level_input_map().get(key)
    if resolved:
        return resolved
    compact = key.replace("/", "_").replace("-", "_").replace(" ", "_")
    return _build_level_input_map().get(compact)


def _find_df_column(df: pd.DataFrame, *candidates: str) -> str | None:
    """Case-insensitive column lookup."""
    lower_map = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        if cand in df.columns:
            return cand
        hit = lower_map.get(cand.lower())
        if hit is not None:
            return hit
    return None


def _is_uuid_string(value: str) -> bool:
    try:
        uuid.UUID(str(value).strip())
        return True
    except (ValueError, TypeError, AttributeError):
        return False


def _cell_str(row: dict, col: str | None) -> str | None:
    if not col:
        return None
    val = row.get(col)
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    text = str(val).strip()
    return text or None


def _tag_lookup_key(tag: str) -> str:
    return str(tag).strip().upper()


def _index_hierarchy_nodes(existing_nodes: list) -> tuple[dict, dict, dict, dict]:
    """Build tag/name indexes for parent resolution (case-insensitive tags)."""
    nodes_by_tag: dict = {}
    nodes_by_name: dict = {}
    nodes_by_id = {n["id"]: n for n in existing_nodes}
    nodes_by_name_parent = {
        (n.get("name", "").lower(), n.get("parent_id")): n for n in existing_nodes
    }
    for n in existing_nodes:
        name_key = (n.get("name") or "").lower()
        if name_key and name_key not in nodes_by_name:
            nodes_by_name[name_key] = n
        tag = n.get("tag")
        if tag:
            nodes_by_tag[tag] = n
            nodes_by_tag[_tag_lookup_key(tag)] = n
    return nodes_by_tag, nodes_by_name, nodes_by_id, nodes_by_name_parent


async def _load_installation_hierarchy_nodes(installation_id: str, user_id: str) -> list:
    """Load all nodes under an installation (matches Equipment Manager tree)."""
    from services.installation_filter_service import installation_filter

    equipment_ids = await installation_filter.get_all_equipment_ids_for_installations(
        [installation_id], user_id
    )
    if not equipment_ids:
        return []
    return await db.equipment_nodes.find(
        {"id": {"$in": list(equipment_ids)}},
        {"_id": 0, "id": 1, "name": 1, "tag": 1, "parent_id": 1, "level": 1, "sort_order": 1},
    ).to_list(10000)


async def _ensure_subunit_parent_for_tag(
    parent_ref: str,
    installation_id: str,
    current_user_id: str,
    existing_nodes: list,
    nodes_by_tag: dict,
    nodes_by_name: dict,
    nodes_by_name_parent: dict,
) -> dict | None:
    """
    When importing maintainable items under a tag like 1F-3001, ensure a subunit exists.
    Creates it under the Strainer equipment unit when missing.
    """
    if not parent_ref:
        return None

    hit = nodes_by_tag.get(parent_ref) or nodes_by_tag.get(_tag_lookup_key(parent_ref))
    if hit:
        if hit.get("level") == "subunit":
            return hit
        if hit.get("level") != "equipment_unit":
            return hit

    strainer = None
    if hit and hit.get("level") == "equipment_unit":
        strainer = hit
    if not strainer:
        for n in existing_nodes:
            if n.get("level") == "equipment_unit" and (n.get("name") or "").strip().lower() == "strainer":
                strainer = n
                break
    if not strainer:
        return None

    now = datetime.now(timezone.utc).isoformat()
    sort_orders = [n.get("sort_order") or 0 for n in existing_nodes if n.get("parent_id") == strainer["id"]]
    tag = parent_ref.strip()
    name = tag if " " in tag else f"Strainer {tag}"
    node_doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "level": "subunit",
        "parent_id": strainer["id"],
        "installation_id": installation_id,
        "tag": tag,
        "created_by": current_user_id,
        "created_at": now,
        "updated_at": now,
        "sort_order": (max(sort_orders) if sort_orders else 0) + 1,
    }
    await db.equipment_nodes.insert_one(node_doc)
    existing_nodes.append(node_doc)
    nodes_by_tag[tag] = node_doc
    nodes_by_tag[_tag_lookup_key(tag)] = node_doc
    nodes_by_name[name.lower()] = node_doc
    nodes_by_name_parent[(name.lower(), strainer["id"])] = node_doc
    logger.info("Auto-created subunit %s (%s) under Strainer for Excel import", tag, name)
    return node_doc


@router.post("/equipment-hierarchy/import-excel")
async def import_excel_file(
    file: UploadFile = File(...),
    installation_id: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Import equipment hierarchy from an uploaded Excel file.
    
    The Excel file should have columns:
    - ID: Tag/identifier for the equipment (required)
    - Name: Display name (required)
    - Level: ISO 14224 level (Plant/Unit, Section/System, Equipment Unit, Subunit, Maintainable Item) (required)
    - Parent: (optional) - If not provided, hierarchy is inferred from sequential order
    - Equipment Type: Type of equipment (optional)
    - Description: Description text (optional)
    - Safety, Production, Environmental, Reputation: Criticality scores 0-5 (optional)
    """
    
    if current_user.get("role") not in ["admin", "owner"]:
        raise HTTPException(status_code=403, detail="Admin or Owner access required")
    
    # Get the installation
    if not installation_id:
        raise HTTPException(status_code=400, detail="installation_id is required")
    
    installation = await db.equipment_nodes.find_one({
        "id": installation_id,
        "level": "installation"
    })
    if not installation:
        raise HTTPException(status_code=404, detail="Installation not found")
    
    # Read the Excel file
    try:
        content = await file.read()
        df = pd.read_excel(io.BytesIO(content))
        df.columns = [str(c).strip() for c in df.columns]
    except Exception as e:
        logger.error(f"Failed to read Excel file: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")
    
    if df.empty:
        raise HTTPException(status_code=400, detail="Excel file has no data rows")
    
    # All nodes under this installation (by tree walk — not installation_id field)
    existing_nodes = await _load_installation_hierarchy_nodes(
        installation_id, current_user["id"]
    )
    nodes_by_tag, nodes_by_name, nodes_by_id, nodes_by_name_parent = _index_hierarchy_nodes(
        existing_nodes
    )
    
    # Track parent at each level for sequential hierarchy inference
    current_parents = {
        -1: {"id": installation_id, "name": installation.get("name")}
    }
    
    # Determine column names (case-insensitive)
    tag_col = _find_df_column(df, "Tag", "Equipment Tag")
    id_col = _find_df_column(df, "ID", "Identifier")
    name_col = _find_df_column(df, "Name", "Line Item Name", "Equipment Name")
    level_col = _find_df_column(df, "Level", "ISO Level", "Hierarchy Level")
    parent_col = _find_df_column(df, "Parent", "Parent Name")
    path_col = _find_df_column(df, "Full Path", "Path", "Hierarchy Path")
    description_col = _find_df_column(df, "Description")
    equipment_type_col = _find_df_column(df, "Equipment Type", "Type")
    safety_col = _find_df_column(df, "Safety")
    production_col = _find_df_column(df, "Production")
    environmental_col = _find_df_column(df, "Environmental")
    reputation_col = _find_df_column(df, "Reputation")
    
    if not name_col:
        raise HTTPException(status_code=400, detail="Excel file must have a 'Name' or 'Line Item Name' column")
    if not level_col:
        raise HTTPException(status_code=400, detail="Excel file must have a 'Level' column")
    
    # Stats
    created_count = 0
    updated_count = 0
    skipped_count = 0
    invalid_level_count = 0
    empty_row_count = 0
    errors = []
    invalid_level_samples: list[str] = []
    
    def resolve_parent_id(parent_name: str | None, iso_level: str) -> str | None:
        """Resolve parent from Parent column or Full Path (match by name or tag)."""
        if parent_name:
            parent_key = parent_name.lower()
            if parent_key in nodes_by_name:
                return nodes_by_name[parent_key]["id"]
            tag_match = (
                nodes_by_tag.get(parent_name)
                or nodes_by_tag.get(_tag_lookup_key(parent_name))
            )
            if tag_match:
                return tag_match["id"]
            install_name = (installation.get("name") or "").lower()
            if parent_key == install_name:
                return installation_id
        return None
    
    # Process rows (sort by level depth so sequential parent stack is more reliable)
    rows = df.to_dict("records")

    def row_sort_key(row):
        iso = resolve_iso_level(row.get(level_col)) if level_col else None
        if iso and iso in ISO_LEVEL_ORDER:
            return ISO_LEVEL_ORDER.index(iso)
        return 99

    rows = sorted(rows, key=row_sort_key)
    batch_inserts = []
    for idx, row in enumerate(rows):
        try:
            name = _cell_str(row, name_col)
            level_raw = row.get(level_col) if level_col else None
            iso_level = resolve_iso_level(level_raw)

            if not iso_level:
                if name or (level_raw is not None and str(level_raw).strip()):
                    invalid_level_count += 1
                    if len(invalid_level_samples) < 5 and level_raw is not None:
                        invalid_level_samples.append(str(level_raw).strip())
                else:
                    empty_row_count += 1
                continue

            if iso_level == "installation":
                empty_row_count += 1
                continue

            tag = _cell_str(row, tag_col)
            if not tag and id_col:
                id_val = _cell_str(row, id_col)
                if id_val and not _is_uuid_string(id_val):
                    tag = id_val

            if not name and tag:
                name = tag
            if not name:
                empty_row_count += 1
                continue

            description = _cell_str(row, description_col)
            equipment_type = _cell_str(row, equipment_type_col)

            # Calculate criticality if scores provided
            criticality = None
            safety = int(row.get(safety_col) or 0) if safety_col and pd.notna(row.get(safety_col)) else 0
            production = int(row.get(production_col) or 0) if production_col and pd.notna(row.get(production_col)) else 0
            environmental = int(row.get(environmental_col) or 0) if environmental_col and pd.notna(row.get(environmental_col)) else 0
            reputation = int(row.get(reputation_col) or 0) if reputation_col and pd.notna(row.get(reputation_col)) else 0

            if safety > 0 or production > 0 or environmental > 0 or reputation > 0:
                criticality = calculate_criticality_from_excel(safety, production, environmental, reputation)

            level_idx = ISO_LEVEL_ORDER.index(iso_level)
            parent_level_idx = level_idx - 1
            parent_id = None

            # Full Path: "Installation > Plant > Section" — parent is second-to-last segment
            full_path = _cell_str(row, path_col)
            if full_path:
                parts = [p.strip() for p in full_path.replace(">", "/").split("/") if p.strip()]
                if len(parts) >= 2:
                    parent_id = resolve_parent_id(parts[-2], iso_level)

            if not parent_id:
                parent_name = _cell_str(row, parent_col)
                parent_id = resolve_parent_id(parent_name, iso_level)

            parent_ref = _cell_str(row, parent_col)
            if iso_level == "maintainable_item" and parent_ref:
                parent_node = nodes_by_id.get(parent_id) if parent_id else None
                needs_subunit = (
                    not parent_id
                    or (parent_node and parent_node.get("level") == "equipment_unit")
                )
                if needs_subunit:
                    ensured = await _ensure_subunit_parent_for_tag(
                        parent_ref,
                        installation_id,
                        current_user["id"],
                        existing_nodes,
                        nodes_by_tag,
                        nodes_by_name,
                        nodes_by_name_parent,
                    )
                    if ensured:
                        parent_id = ensured["id"]

            if not parent_id:
                if parent_level_idx not in current_parents:
                    skipped_count += 1
                    errors.append(
                        f"Row {idx + 2}: missing parent for '{name}' ({iso_level}); "
                        "add parent rows above or use Parent / Full Path columns"
                    )
                    continue
                parent_id = current_parents[parent_level_idx]["id"]
            
            # Check if exists by tag
            tag_hit = tag and (
                nodes_by_tag.get(tag) or nodes_by_tag.get(_tag_lookup_key(tag))
            )
            if tag_hit:
                existing = tag_hit
                current_parents[level_idx] = existing
                nodes_by_name[name.lower()] = existing
                skipped_count += 1
                if len(errors) < 10:
                    errors.append(f"Row {idx + 2}: tag '{tag}' already exists ({existing.get('name')})")
                continue
            
            # Check if exists by name + parent
            lookup_key = (name.lower(), parent_id)
            if lookup_key in nodes_by_name_parent:
                existing = nodes_by_name_parent[lookup_key]
                # Update with tag if missing
                if tag and not existing.get("tag"):
                    update_data = {"tag": tag, "updated_at": datetime.now(timezone.utc).isoformat()}
                    if description:
                        update_data["description"] = description
                    if criticality:
                        update_data["criticality"] = criticality
                    await db.equipment_nodes.update_one({"id": existing["id"]}, {"$set": update_data})
                    updated_count += 1
                current_parents[level_idx] = existing
                nodes_by_name[name.lower()] = existing
                if tag:
                    nodes_by_tag[tag] = existing
                continue
            
            # Create new node
            node_id = str(uuid.uuid4())
            node_doc = {
                "id": node_id,
                "name": name,
                "level": iso_level,
                "parent_id": parent_id,
                "installation_id": installation_id,
                "tag": tag,
                "equipment_type": equipment_type,
                "description": description,
                "criticality": criticality,
                "created_by": current_user["id"],
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            
            batch_inserts.append(node_doc)
            
            current_parents[level_idx] = node_doc
            nodes_by_name[name.lower()] = node_doc
            if tag:
                nodes_by_tag[tag] = node_doc
            nodes_by_name_parent[(name.lower(), parent_id)] = node_doc
            created_count += 1
            
        except Exception as e:
            errors.append(f"Row {idx + 2}: {str(e)}")
            logger.error(f"Error processing row {idx + 2}: {e}")
    
    # Batch insert all new nodes
    if batch_inserts:
        await db.equipment_nodes.insert_many(batch_inserts)
        # Remove _id from docs
        for doc in batch_inserts:
            doc.pop("_id", None)
    
    if created_count == 0 and updated_count == 0 and invalid_level_count > 0:
        sample = ", ".join(repr(s) for s in invalid_level_samples)
        raise HTTPException(
            status_code=400,
            detail=(
                f"No rows imported: {invalid_level_count} row(s) had unrecognized Level values. "
                f"Use labels like Plant/Unit, Section/System, Equipment Unit, Subunit, Maintainable Item "
                f"(or plant_unit, section_system, etc.). Examples from file: {sample}"
            ),
        )

    return {
        "success": True,
        "installation_id": installation_id,
        "installation_name": installation.get("name"),
        "created_count": created_count,
        "updated_count": updated_count,
        "skipped_count": skipped_count,
        "invalid_level_count": invalid_level_count,
        "empty_row_count": empty_row_count,
        "invalid_level_samples": invalid_level_samples,
        "errors": errors[:10] if errors else [],
        "message": (
            f"Import complete: {created_count} created, {updated_count} updated, "
            f"{skipped_count} skipped"
            + (f", {invalid_level_count} invalid level(s)" if invalid_level_count else "")
        ),
    }


@router.post("/equipment/import-hierarchy-excel")
async def import_hierarchy_from_excel(
    request: ExcelHierarchyImportRequest,
    current_user: dict = Depends(get_current_user)
):
    """Import equipment hierarchy from an Excel file URL."""
    import requests as req_lib
    from openpyxl import load_workbook
    
    if current_user.get("role") not in ["admin", "owner"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    installation = await db.equipment_nodes.find_one({
        "id": request.installation_id,
        "level": "installation"
    })
    if not installation:
        raise HTTPException(status_code=404, detail="Installation not found")
    
    installation_id = request.installation_id
    
    try:
        response = req_lib.get(request.excel_url, timeout=30)
        response.raise_for_status()
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
    except Exception as e:
        logger.error(f"Failed to load Excel file: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to load Excel file: {str(e)}")
    
    headers = [cell.value for cell in ws[1]]
    level_order = {lvl: i for i, lvl in enumerate(ISO_LEVEL_ORDER)}
    
    current_path = []
    all_items = []
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = dict(zip(headers, row))
        name = data.get('Name')
        level_raw = data.get('Level')
        
        if not name or not level_raw:
            continue
        
        level = resolve_iso_level(level_raw)
        if not level or level == "installation":
            logger.warning(f"Unknown level '{level_raw}' for '{name}', skipping...")
            continue
        
        name = name.strip() if isinstance(name, str) else name
        level_num = level_order[level]
        
        safety = int(data.get('Safety') or 0)
        production = int(data.get('Production') or 0)
        environmental = int(data.get('Environmental') or 0)
        reputation = int(data.get('Reputation') or 0)
        
        criticality = None
        if safety > 0 or production > 0 or environmental > 0 or reputation > 0:
            criticality = calculate_criticality_from_excel(safety, production, environmental, reputation)
        
        while current_path and level_order.get(current_path[-1][0], 99) >= level_num:
            current_path.pop()
        
        parent_name = current_path[-1][1] if current_path else None
        current_path.append((level, name))
        full_path = ' > '.join([p[1] for p in current_path])
        
        all_items.append({
            'name': name,
            'level': level,
            'parent_name': parent_name,
            'full_path': full_path,
            'equipment_type': data.get('Equipment Type'),
            'discipline': data.get('Discipline'),
            'description': data.get('Description'),
            'criticality': criticality
        })
    
    unique_items = {}
    for item in all_items:
        key = item['full_path']
        if key not in unique_items:
            unique_items[key] = item
        elif item['criticality'] and not unique_items[key].get('criticality'):
            unique_items[key]['criticality'] = item['criticality']
    
    items_list = list(unique_items.values())
    logger.info(f"Parsed {len(items_list)} unique equipment items from Excel")
    
    deleted_count = 0
    if request.replace_existing:
        result = await db.equipment_nodes.delete_many({"installation_id": installation_id})
        deleted_count = result.deleted_count
    
    sorted_items = sorted(items_list, key=lambda x: level_order.get(x['level'], 5))
    
    path_to_id = {}
    equipment_list = []
    sort_order = 0
    
    for item in sorted_items:
        sort_order += 1
        eq_id = str(uuid.uuid4())
        
        path_to_id[item['full_path']] = eq_id
        
        parent_id = installation_id
        if item.get('parent_name'):
            parent_path_parts = item['full_path'].rsplit(' > ', 1)
            if len(parent_path_parts) > 1:
                parent_path = parent_path_parts[0]
                parent_id = path_to_id.get(parent_path, installation_id)
        
        eq = {
            "id": eq_id,
            "name": item['name'],
            "parent_id": parent_id,
            "installation_id": installation_id,
            "level": item['level'],
            "equipment_type": item.get('equipment_type'),
            "discipline": item.get('discipline'),
            "description": item.get('description'),
            "criticality": item.get('criticality'),
            "sort_order": sort_order,
            "created_by": current_user["id"],
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        equipment_list.append(eq)
    
    inserted_count = 0
    items_with_criticality = 0
    if equipment_list:
        result = await db.equipment_nodes.insert_many(equipment_list)
        inserted_count = len(result.inserted_ids)
        items_with_criticality = sum(1 for eq in equipment_list if eq.get('criticality'))
    
    from collections import Counter
    level_counts = dict(Counter([eq['level'] for eq in equipment_list]))
    
    logger.info(f"Excel hierarchy import: deleted={deleted_count}, inserted={inserted_count}")
    
    return {
        "success": True,
        "installation_id": installation_id,
        "installation_name": installation.get("name"),
        "deleted_count": deleted_count,
        "inserted_count": inserted_count,
        "items_with_criticality": items_with_criticality,
        "by_level": level_counts,
        "message": f"Successfully imported {inserted_count} equipment items"
    }


class HierarchyImportRequest(BaseModel):
    installation_id: str
    hierarchy: dict
    replace_existing: bool = True


@router.post("/equipment/import-hierarchy")
async def import_equipment_hierarchy(
    request: HierarchyImportRequest,
    current_user: dict = Depends(get_current_user)
):
    """Import a complete equipment hierarchy for an installation."""
    if current_user.get("role") not in ["admin", "owner"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    installation = await db.installations.find_one({"id": request.installation_id})
    if not installation:
        raise HTTPException(status_code=404, detail="Installation not found")
    
    installation_id = request.installation_id
    
    deleted_count = 0
    if request.replace_existing:
        result = await db.equipment.delete_many({"installation_id": installation_id})
        deleted_count = result.deleted_count
    
    equipment_list = []
    sort_order = 0
    
    def create_equipment(name, parent_id, eq_type, level):
        nonlocal sort_order
        sort_order += 1
        return {
            "id": str(uuid.uuid4()),
            "name": name,
            "parent_id": parent_id,
            "installation_id": installation_id,
            "type": eq_type,
            "level": level,
            "sort_order": sort_order,
            "created_by": current_user["id"],
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
    
    def process_hierarchy(data, parent_id=None, level=0):
        items = []
        for name, info in data.items():
            if isinstance(info, dict):
                eq_type = info.get("type", "equipment")
                eq = create_equipment(name, parent_id, eq_type, level)
                items.append(eq)
                
                children = info.get("children", {})
                if isinstance(children, dict):
                    items.extend(process_hierarchy(children, eq["id"], level + 1))
                elif isinstance(children, list):
                    for child_name in children:
                        child_eq = create_equipment(child_name, eq["id"], "maintainable_item", level + 1)
                        items.append(child_eq)
        return items
    
    equipment_list = process_hierarchy(request.hierarchy)
    
    inserted_count = 0
    if equipment_list:
        result = await db.equipment.insert_many(equipment_list)
        inserted_count = len(result.inserted_ids)
    
    logger.info(f"Hierarchy import: deleted={deleted_count}, inserted={inserted_count}")
    
    return {
        "success": True,
        "installation_id": installation_id,
        "deleted_count": deleted_count,
        "inserted_count": inserted_count,
        "message": f"Successfully imported {inserted_count} equipment items"
    }
