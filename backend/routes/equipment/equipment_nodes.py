"""
Equipment Node CRUD operations.
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone
import uuid
import io
import logging
from database import db, efm_service, installation_filter
from auth import get_current_user
from iso14224_models import (
    ISOLevel, ISO_LEVEL_ORDER, EQUIPMENT_TYPES, ISO_LEVEL_LABELS,
    get_valid_child_levels, is_valid_parent_child, EquipmentNodeCreate, EquipmentNodeUpdate
)

logger = logging.getLogger(__name__)
router = APIRouter()


@router.get("/equipment-hierarchy/nodes")
async def get_equipment_nodes(
    current_user: dict = Depends(get_current_user)
):
    """Get equipment hierarchy nodes filtered by user's assigned installations."""
    from services.query_cache import query_cache
    import time
    
    # Check cache first (keyed by user ID since results are user-specific)
    cache_key = f"equipment_nodes:{current_user['id']}"
    cached = query_cache.get(cache_key)
    if cached is not None:
        return cached
    
    # Get user's installation filter data
    installation_ids = await installation_filter.get_user_installation_ids(current_user)
    
    # If no installations assigned, return empty list
    if not installation_ids:
        result = {"nodes": []}
        query_cache.set(cache_key, result, ttl=60)
        return result
    
    # Get all equipment IDs under assigned installations (shared equipment - no created_by filter)
    equipment_ids = await installation_filter.get_all_equipment_ids_for_installations(
        installation_ids, current_user["id"]
    )
    
    if not equipment_ids:
        result = {"nodes": []}
        query_cache.set(cache_key, result, ttl=60)
        return result
    
    # Get all nodes that belong to assigned installations (no created_by filter)
    nodes = await db.equipment_nodes.find(
        {"id": {"$in": list(equipment_ids)}},
        {"_id": 0}
    ).sort("sort_order", 1).to_list(5000)
    
    result = {"nodes": nodes}
    # Cache for 2 minutes
    query_cache.set(cache_key, result, ttl=120)
    return result


@router.get("/equipment-hierarchy/installations")
async def get_all_installations(
    current_user: dict = Depends(get_current_user)
):
    """Get all installation-level nodes across all users (for admin assignment)."""
    # Get nodes that are ONLY installations (top-level per ISO 14224)
    nodes = await db.equipment_nodes.find(
        {"level": "installation"},
        {"_id": 0, "id": 1, "name": 1, "level": 1, "created_by": 1}
    ).sort("name", 1).to_list(500)
    
    # Deduplicate by name (in case same installation exists for multiple users)
    seen_names = set()
    unique_nodes = []
    for node in nodes:
        if node.get("name") not in seen_names:
            seen_names.add(node.get("name"))
            unique_nodes.append(node)
    
    return {"installations": unique_nodes}


@router.get("/equipment-hierarchy/export")
async def export_equipment_hierarchy_excel(
    current_user: dict = Depends(get_current_user)
):
    """Export equipment hierarchy to an Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Use the same installation-based filtering as get_equipment_nodes
    installation_ids = await installation_filter.get_user_installation_ids(current_user)
    
    if not installation_ids:
        # Return empty Excel if no installations assigned
        wb = Workbook()
        ws = wb.active
        ws.title = "Equipment Hierarchy"
        ws['A1'] = "No equipment found"
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=equipment_hierarchy_empty.xlsx"}
        )
    
    # Get all equipment IDs under assigned installations
    equipment_ids = await installation_filter.get_all_equipment_ids_for_installations(
        installation_ids, current_user["id"]
    )
    
    # Fetch nodes using installation filter (same as hierarchy view)
    nodes = await db.equipment_nodes.find(
        {"id": {"$in": list(equipment_ids)}} if equipment_ids else {},
        {"_id": 0}
    ).sort("sort_order", 1).to_list(5000)
    
    # Fetch equipment types for lookup (both user's custom and defaults)
    equipment_types = await db.custom_equipment_types.find(
        {},
        {"_id": 0}
    ).to_list(100)
    
    # Merge with defaults
    all_types = {t["id"]: t["name"] for t in EQUIPMENT_TYPES}
    for t in equipment_types:
        all_types[t["id"]] = t["name"]
    
    # Build parent lookup for path generation
    node_lookup = {n["id"]: n for n in nodes}
    
    def get_path(node):
        """Generate full path from root to node."""
        if not isinstance(node, dict):
            return ""
        path_parts = [node.get("name", "")]
        current = node
        while isinstance(current, dict) and current.get("parent_id") and current["parent_id"] in node_lookup:
            current = node_lookup[current["parent_id"]]
            if isinstance(current, dict):
                path_parts.insert(0, current.get("name", ""))
            else:
                break
        return " > ".join(path_parts)
    
    def get_criticality_score(node):
        """Calculate total criticality score."""
        if not isinstance(node, dict):
            return 0
        crit = node.get("criticality")
        if not isinstance(crit, dict):
            return 0
        return sum([
            crit.get("safety_impact", crit.get("safety", 0)) or 0,
            crit.get("production_impact", crit.get("production", 0)) or 0,
            crit.get("environmental_impact", crit.get("environmental", 0)) or 0,
            crit.get("reputation_impact", crit.get("reputation", 0)) or 0
        ])
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment Hierarchy"
    
    # Define headers
    headers = [
        "ID", "Name", "Tag", "Level", "Parent", "Full Path", "Equipment Type",
        "Discipline", "Process Step", "Description",
        "Safety", "Production", "Environmental", "Reputation", "Total Criticality",
        "Created At"
    ]
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Level colors for visual hierarchy
    level_colors = {
        "installation": "E8F5E9",
        "plant_unit": "C8E6C9",
        "section_system": "A5D6A7",
        "equipment_unit": "81C784",
        "equipment": "66BB6A",
        "subunit": "4CAF50",
        "maintainable_item": "43A047"
    }
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    for row_idx, node in enumerate(nodes, 2):
        if not isinstance(node, dict):
            logger.warning(f"Skipping corrupted node at row {row_idx}")
            continue
            
        parent_name = ""
        if node.get("parent_id") and node["parent_id"] in node_lookup:
            parent_node = node_lookup[node["parent_id"]]
            if isinstance(parent_node, dict):
                parent_name = parent_node.get("name", "")
        
        equipment_type_name = ""
        if node.get("equipment_type_id"):
            equipment_type_name = all_types.get(node["equipment_type_id"], "")
        
        criticality = node.get("criticality")
        if not isinstance(criticality, dict):
            criticality = {}
        
        level_label = ""
        node_level = node.get("level")
        if node_level:
            try:
                level_label = ISO_LEVEL_LABELS.get(ISOLevel(node_level), str(node_level))
            except (ValueError, TypeError):
                level_label = str(node_level)
        
        row_data = [
            node.get("id", ""),
            node.get("name", ""),
            node.get("tag", ""),
            level_label,
            parent_name,
            get_path(node),
            equipment_type_name or node.get("equipment_type", ""),
            (node.get("discipline") or "").replace("_", " ").title(),
            node.get("process_step", ""),
            node.get("description", ""),
            criticality.get("safety_impact", criticality.get("safety", 0)) or 0,
            criticality.get("production_impact", criticality.get("production", 0)) or 0,
            criticality.get("environmental_impact", criticality.get("environmental", 0)) or 0,
            criticality.get("reputation_impact", criticality.get("reputation", 0)) or 0,
            get_criticality_score(node),
            node.get("created_at", "")[:10] if node.get("created_at") else ""
        ]
        
        level_color = level_colors.get(node.get("level"), "FFFFFF")
        row_fill = PatternFill(start_color=level_color, end_color=level_color, fill_type="solid")
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col <= 3:
                cell.fill = row_fill
    
    # Adjust column widths (added Tag column)
    column_widths = [38, 25, 15, 18, 20, 50, 20, 15, 20, 30, 8, 10, 12, 10, 14, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    ws.freeze_panes = "A2"
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"equipment_hierarchy_{timestamp}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/equipment-hierarchy/nodes/{node_id}")
async def get_equipment_node(
    node_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get a specific equipment node."""
    node = await db.equipment_nodes.find_one(
        {"id": node_id},
        {"_id": 0}
    )
    if not node:
        raise HTTPException(status_code=404, detail="Equipment node not found")
    return node


@router.post("/equipment-hierarchy/nodes")
async def create_equipment_node(
    node_data: EquipmentNodeCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a new equipment hierarchy node with ISO 14224 validation."""
    # Check if user is trying to create an installation (root node)
    if not node_data.parent_id and node_data.level == ISOLevel.INSTALLATION:
        if current_user.get("role") != "owner":
            raise HTTPException(
                status_code=403,
                detail="Only owners can create new installations"
            )
    
    # Check for duplicate name under the same parent
    existing = await db.equipment_nodes.find_one({
        "name": node_data.name,
        "parent_id": node_data.parent_id
    })
    if existing:
        parent_info = f"under parent '{node_data.parent_id}'" if node_data.parent_id else "at root level"
        raise HTTPException(
            status_code=400, 
            detail=f"A node with name '{node_data.name}' already exists {parent_info}"
        )
    
    # Validate parent-child relationship if parent specified
    if node_data.parent_id:
        parent = await db.equipment_nodes.find_one(
            {"id": node_data.parent_id},
            {"_id": 0}
        )
        if not parent:
            raise HTTPException(status_code=400, detail="Parent node not found")
        
        parent_level = ISOLevel(parent["level"])
        if not is_valid_parent_child(parent_level, node_data.level):
            valid_children = get_valid_child_levels(parent_level)
            raise HTTPException(
                status_code=400, 
                detail=f"Invalid parent-child relationship. {parent_level.value} can only have {[c.value for c in valid_children]} as children"
            )
    else:
        if node_data.level != ISOLevel.INSTALLATION:
            raise HTTPException(
                status_code=400, 
                detail="Root nodes must be of level 'installation'"
            )
    
    node_id = str(uuid.uuid4())
    
    max_sort = await db.equipment_nodes.find_one(
        {"parent_id": node_data.parent_id},
        sort=[("sort_order", -1)],
        projection={"sort_order": 1}
    )
    next_sort_order = (max_sort.get("sort_order", 0) if max_sort else 0) + 1
    
    # Inherit process_step from parent if not provided
    inherited_process_step = node_data.process_step
    if not inherited_process_step and node_data.parent_id and node_data.level in [ISOLevel.SUBUNIT, ISOLevel.MAINTAINABLE_ITEM]:
        parent = await db.equipment_nodes.find_one(
            {"id": node_data.parent_id},
            {"process_step": 1}
        )
        if parent and parent.get("process_step"):
            inherited_process_step = parent["process_step"]
    
    node_doc = {
        "id": node_id,
        "name": node_data.name,
        "level": node_data.level.value,
        "parent_id": node_data.parent_id,
        "equipment_type_id": node_data.equipment_type_id,
        "description": node_data.description,
        "tag": node_data.tag,
        "process_step": inherited_process_step,
        "criticality": None,
        "discipline": None,
        "sort_order": next_sort_order,
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.equipment_nodes.insert_one(node_doc)
    
    # Auto-generate EFMs if equipment_type_id is specified
    if node_data.equipment_type_id:
        try:
            efms = await efm_service.generate_efms_for_equipment(
                equipment_id=node_id,
                equipment_name=node_data.name,
                equipment_type_id=node_data.equipment_type_id
            )
            node_doc["efms_generated"] = len(efms)
            logger.info(f"Auto-generated {len(efms)} EFMs for equipment {node_id}")
        except Exception as e:
            logger.error(f"Failed to generate EFMs for {node_id}: {e}")
    
    node_doc.pop("_id", None)
    return node_doc


@router.patch("/equipment-hierarchy/nodes/{node_id}")
async def update_equipment_node(
    node_id: str,
    update: EquipmentNodeUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update an equipment hierarchy node."""
    node = await db.equipment_nodes.find_one({"id": node_id})
    if not node:
        raise HTTPException(status_code=404, detail="Equipment node not found")
    
    # Validate new parent if changing parent
    if update.parent_id is not None and update.parent_id != node.get("parent_id"):
        if update.parent_id:
            new_parent = await db.equipment_nodes.find_one(
                {"id": update.parent_id},
                {"_id": 0}
            )
            if not new_parent:
                raise HTTPException(status_code=400, detail="New parent node not found")
            
            parent_level = ISOLevel(new_parent["level"])
            child_level = ISOLevel(node["level"])
            if not is_valid_parent_child(parent_level, child_level):
                raise HTTPException(
                    status_code=400, 
                    detail="Invalid parent-child relationship per ISO 14224"
                )
            
            # Check for circular references
            current_parent = update.parent_id
            while current_parent:
                if current_parent == node_id:
                    raise HTTPException(status_code=400, detail="Circular reference detected")
                parent_node = await db.equipment_nodes.find_one({"id": current_parent})
                current_parent = parent_node.get("parent_id") if parent_node else None
        else:
            if node["level"] != ISOLevel.INSTALLATION.value:
                raise HTTPException(
                    status_code=400, 
                    detail="Only installations can be root nodes"
                )
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.equipment_nodes.update_one(
            {"id": node_id},
            {"$set": update_data}
        )
        
        # Sync EFMs if equipment_type_id changed
        if update.equipment_type_id is not None:
            old_type = node.get("equipment_type_id")
            new_type = update.equipment_type_id
            if old_type != new_type:
                try:
                    sync_result = await efm_service.sync_efms_for_equipment(
                        equipment_id=node_id,
                        equipment_name=node.get("name"),
                        new_equipment_type_id=new_type,
                        old_equipment_type_id=old_type
                    )
                    logger.info(f"Synced EFMs for {node_id}: {sync_result}")
                except Exception as e:
                    logger.error(f"Failed to sync EFMs for {node_id}: {e}")
    
    updated = await db.equipment_nodes.find_one({"id": node_id}, {"_id": 0})
    return updated


@router.get("/equipment-hierarchy/nodes/{node_id}/deletion-impact")
async def get_deletion_impact(
    node_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get impact analysis for deleting an equipment node."""
    node = await db.equipment_nodes.find_one({"id": node_id})
    if not node:
        raise HTTPException(status_code=404, detail="Equipment node not found")
    
    async def get_children_ids(parent_id):
        children = await db.equipment_nodes.find(
            {"parent_id": parent_id},
            {"_id": 0, "id": 1, "name": 1}
        ).to_list(1000)
        all_items = [{"id": c["id"], "name": c["name"]} for c in children]
        for child in children:
            all_items.extend(await get_children_ids(child["id"]))
        return all_items
    
    children = await get_children_ids(node_id)
    all_equipment_ids = [node_id] + [c["id"] for c in children]
    
    impacted_tasks = await db.task_instances.find(
        {"equipment_id": {"$in": all_equipment_ids}, "status": {"$ne": "completed"}},
        {"_id": 0, "id": 1, "title": 1, "status": 1, "equipment_name": 1}
    ).to_list(100)
    
    impacted_actions = await db.central_actions.find(
        {"equipment_id": {"$in": all_equipment_ids}, "status": {"$nin": ["completed", "closed"]}},
        {"_id": 0, "id": 1, "title": 1, "status": 1, "equipment_name": 1}
    ).to_list(100)
    
    impacted_investigations = await db.threats.find(
        {"asset_id": {"$in": all_equipment_ids}, "status": {"$nin": ["closed", "mitigated"]}},
        {"_id": 0, "id": 1, "title": 1, "status": 1, "asset": 1}
    ).to_list(100)
    
    impacted_plans = await db.task_plans.find(
        {"equipment_id": {"$in": all_equipment_ids}, "is_active": True},
        {"_id": 0, "id": 1, "task_template_name": 1, "equipment_name": 1}
    ).to_list(100)
    
    return {
        "node": {"id": node_id, "name": node.get("name"), "level": node.get("level")},
        "children_count": len(children),
        "children": children[:10],
        "impact": {
            "tasks": {"count": len(impacted_tasks), "items": impacted_tasks[:5], "will_be": "orphaned"},
            "actions": {"count": len(impacted_actions), "items": impacted_actions[:5], "will_be": "orphaned"},
            "investigations": {"count": len(impacted_investigations), "items": impacted_investigations[:5], "will_be": "orphaned"},
            "task_plans": {"count": len(impacted_plans), "items": impacted_plans[:5], "will_be": "deactivated"}
        },
        "total_impacted": len(impacted_tasks) + len(impacted_actions) + len(impacted_investigations) + len(impacted_plans)
    }


@router.delete("/equipment-hierarchy/nodes/{node_id}")
async def delete_equipment_node(
    node_id: str,
    cascade: bool = False,
    current_user: dict = Depends(get_current_user)
):
    """Delete an equipment node and optionally cascade to related items."""
    node = await db.equipment_nodes.find_one({"id": node_id})
    if not node:
        raise HTTPException(status_code=404, detail="Equipment node not found")
    
    # Only Owner can delete installations
    if node.get("level") == "installation":
        if current_user.get("role") != "owner":
            raise HTTPException(
                status_code=403, 
                detail="Only Owner can delete installations."
            )
    
    async def get_children_ids(parent_id):
        children = await db.equipment_nodes.find(
            {"parent_id": parent_id},
            {"_id": 0, "id": 1}
        ).to_list(1000)
        all_ids = [c["id"] for c in children]
        for child in children:
            all_ids.extend(await get_children_ids(child["id"]))
        return all_ids
    
    children_ids = await get_children_ids(node_id)
    all_ids = [node_id] + children_ids
    
    # Orphan related items
    orphaned_tasks = await db.task_instances.update_many(
        {"equipment_id": {"$in": all_ids}},
        {"$set": {"equipment_id": None, "equipment_name": f"[Deleted: {node.get('name')}]"}}
    )
    
    orphaned_actions = await db.central_actions.update_many(
        {"equipment_id": {"$in": all_ids}},
        {"$set": {"equipment_id": None, "equipment_name": f"[Deleted: {node.get('name')}]"}}
    )
    
    orphaned_investigations = await db.threats.update_many(
        {"asset_id": {"$in": all_ids}},
        {"$set": {"asset_id": None, "asset": f"[Deleted: {node.get('name')}]"}}
    )
    
    deactivated_plans = await db.task_plans.update_many(
        {"equipment_id": {"$in": all_ids}},
        {"$set": {"is_active": False, "deactivation_reason": f"Equipment deleted: {node.get('name')}"}}
    )
    
    result = await db.equipment_nodes.delete_many({"id": {"$in": all_ids}})
    
    logger.info(f"Deleted equipment node {node_id} and {len(children_ids)} children")
    
    return {
        "message": f"Deleted {result.deleted_count} nodes", 
        "deleted_ids": all_ids,
        "orphaned": {
            "tasks": orphaned_tasks.modified_count,
            "actions": orphaned_actions.modified_count,
            "investigations": orphaned_investigations.modified_count,
            "plans_deactivated": deactivated_plans.modified_count
        }
    }
