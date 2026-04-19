"""
Production Log Ingestion & History Builder.

Endpoints for uploading, parsing, previewing, and ingesting production log files.
Supports CSV, TXT, LOG, ZIP formats with template-based and AI-assisted parsing.
Owner-only access.
"""
import os
import io
import re
import csv
import json
import uuid
import zipfile
import logging
import asyncio
from datetime import datetime, timezone
from typing import List, Optional, Dict, Any
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, BackgroundTasks
from pydantic import BaseModel
from auth import get_current_user
from database import db
from openai import OpenAI
from services.storage_service import put_object_async

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/production-logs", tags=["Production Logs"])

ALLOWED_EXTENSIONS = {"csv", "txt", "log", "zip", "xlsx", "xls"}
MAX_FILE_SIZE = 100 * 1024 * 1024  # 100MB


# ======================== Models ========================

class ColumnMapping(BaseModel):
    timestamp: Optional[str] = None
    asset_id: Optional[str] = None
    status: Optional[str] = None
    event_type: Optional[str] = None
    metric_columns: List[str] = []

class ParseTemplate(BaseModel):
    delimiter: str = ","
    has_header: bool = True
    skip_rows: int = 0
    timestamp_format: Optional[str] = None
    column_mapping: ColumnMapping = ColumnMapping()

class IngestRequest(BaseModel):
    job_id: str
    confirm: bool = True

class BatchIngestRequest(BaseModel):
    job_ids: List[str]
    template: dict


class ColumnAlias(BaseModel):
    """Defines aliases for flexible column matching."""
    canonical_name: str  # The standard name used internally
    aliases: List[str] = []  # Alternative names that map to this column


class SavedTemplate(BaseModel):
    name: str
    description: Optional[str] = None
    delimiter: str = ","
    has_header: bool = True
    skip_rows: int = 0
    timestamp_format: Optional[str] = None
    column_mapping: ColumnMapping = ColumnMapping()
    # Column aliases for fuzzy matching
    column_aliases: Dict[str, List[str]] = {}  # {canonical_name: [alias1, alias2, ...]}


class SaveTemplateRequest(BaseModel):
    name: str
    description: Optional[str] = None
    template: dict
    column_aliases: Dict[str, List[str]] = {}


class BatchIngestWithTemplateRequest(BaseModel):
    job_ids: List[str]
    template_id: str


# ======================== Helpers ========================

def _owner_only(user: dict):
    if user.get("role") != "owner":
        raise HTTPException(status_code=403, detail="Owner access required")


def _normalize_column_name(name: str) -> str:
    """Normalize column name for fuzzy matching."""
    # Lowercase, remove special chars, collapse whitespace
    normalized = name.lower().strip()
    normalized = re.sub(r'[^a-z0-9]', '', normalized)
    return normalized


def _fuzzy_match_columns(
    file_columns: List[str], 
    template_mapping: ColumnMapping,
    column_aliases: Dict[str, List[str]] = None
) -> Dict[str, str]:
    """
    Match file columns to template columns using fuzzy matching.
    Returns a mapping of {template_column: matched_file_column}.
    """
    column_aliases = column_aliases or {}
    
    # Build a lookup of normalized names -> original file column names
    file_col_lookup = {}
    for col in file_columns:
        norm = _normalize_column_name(col)
        file_col_lookup[norm] = col
        # Also try without common prefixes/suffixes
        for prefix in ['col_', 'field_', 'column_']:
            if norm.startswith(prefix):
                file_col_lookup[norm[len(prefix):]] = col
    
    def find_match(template_col: str) -> Optional[str]:
        """Find matching file column for a template column."""
        if not template_col:
            return None
            
        # 1. Exact match
        if template_col in file_columns:
            return template_col
            
        # 2. Normalized exact match
        norm_template = _normalize_column_name(template_col)
        if norm_template in file_col_lookup:
            return file_col_lookup[norm_template]
        
        # 3. Check aliases
        aliases = column_aliases.get(template_col, [])
        for alias in aliases:
            if alias in file_columns:
                return alias
            norm_alias = _normalize_column_name(alias)
            if norm_alias in file_col_lookup:
                return file_col_lookup[norm_alias]
        
        # 4. Fuzzy substring match (template col contains or is contained in file col)
        for norm_file, orig_file in file_col_lookup.items():
            if norm_template in norm_file or norm_file in norm_template:
                return orig_file
                
        return None
    
    # Match all template columns
    matched = {}
    
    if template_mapping.timestamp:
        match = find_match(template_mapping.timestamp)
        matched['timestamp'] = match or template_mapping.timestamp
        
    if template_mapping.asset_id:
        match = find_match(template_mapping.asset_id)
        matched['asset_id'] = match or template_mapping.asset_id
        
    if template_mapping.status:
        match = find_match(template_mapping.status)
        matched['status'] = match or template_mapping.status
        
    if template_mapping.event_type:
        match = find_match(template_mapping.event_type)
        matched['event_type'] = match or template_mapping.event_type
    
    # Match metric columns
    matched_metrics = []
    for metric_col in template_mapping.metric_columns:
        match = find_match(metric_col)
        if match:
            matched_metrics.append(match)
        else:
            matched_metrics.append(metric_col)
    matched['metric_columns'] = matched_metrics
    
    return matched


def _get_ext(filename: str) -> str:
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def _detect_delimiter(sample: str) -> str:
    """Auto-detect delimiter from first few lines."""
    for delim in [",", ";", "\t", "|"]:
        if delim in sample:
            return delim
    return ","


def _parse_timestamp(val: str, fmt: Optional[str] = None, base_date: Optional[str] = None) -> Optional[str]:
    """Try to parse a timestamp string into ISO format.
    
    Args:
        val: The timestamp string to parse
        fmt: Optional format string
        base_date: Optional base date (YYYY-MM-DD) to combine with time-only values
    """
    if not val or not val.strip():
        return None
    val = str(val).strip()
    
    # Handle time-only values (HH:MM:SS or HH:MM)
    time_only_formats = ["%H:%M:%S", "%H:%M"]
    for tf in time_only_formats:
        try:
            time_obj = datetime.strptime(val, tf)
            # If we have a base date, combine them
            if base_date:
                try:
                    date_obj = datetime.strptime(base_date, "%Y-%m-%d")
                    combined = datetime.combine(date_obj.date(), time_obj.time())
                    return combined.isoformat()
                except:
                    pass
            # Otherwise, use today's date as fallback
            today = datetime.now().date()
            combined = datetime.combine(today, time_obj.time())
            return combined.isoformat()
        except ValueError:
            continue
    
    formats = []
    if fmt:
        formats.append(fmt)
    formats.extend([
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%m/%d/%Y %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d",
        "%d-%m-%Y",
    ])
    for f in formats:
        try:
            dt = datetime.strptime(val, f)
            return dt.isoformat()
        except ValueError:
            continue
    # Try ISO parse as last resort
    try:
        dt = datetime.fromisoformat(val.replace("Z", "+00:00"))
        return dt.isoformat()
    except Exception:
        return None


def _classify_event(status: Optional[str], metrics: dict) -> str:
    """Basic event classification rules."""
    if not status:
        return "normal"
    s = str(status).lower().strip()
    if any(w in s for w in ["stop", "down", "off", "shutdown", "error", "fault", "trip"]):
        return "downtime"
    if any(w in s for w in ["waste", "reject", "scrap", "discard"]):
        return "waste"
    if any(w in s for w in ["alarm", "alert", "warning", "abnormal", "high", "low", "critical"]):
        return "alarm"
    return "normal"


def _parse_csv_content(content: str, template: ParseTemplate) -> List[dict]:
    """Parse CSV/TXT content using a template configuration."""
    lines = content.splitlines()
    if template.skip_rows > 0:
        lines = lines[template.skip_rows:]
    if not lines:
        return []

    delimiter = template.delimiter
    if delimiter == "\\t":
        delimiter = "\t"

    reader = csv.reader(lines, delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return []

    headers = None
    data_start = 0
    if template.has_header:
        headers = [h.strip() for h in rows[0]]
        data_start = 1
    else:
        headers = [f"col_{i}" for i in range(len(rows[0]))]

    mapping = template.column_mapping
    records = []
    errors = []

    for row_idx, row in enumerate(rows[data_start:], start=data_start + 1):
        if not row or all(not cell.strip() for cell in row):
            continue  # Skip empty rows

        record = {"_row": row_idx}

        # Build a dict from row
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = val.strip()

        # Map timestamp
        ts_val = row_dict.get(mapping.timestamp) if mapping.timestamp else None
        if ts_val:
            parsed_ts = _parse_timestamp(ts_val, template.timestamp_format)
            record["timestamp"] = parsed_ts
            if not parsed_ts:
                record["_errors"] = record.get("_errors", []) + [f"Invalid timestamp: {ts_val}"]
        else:
            record["timestamp"] = None
            if mapping.timestamp:
                record["_errors"] = record.get("_errors", []) + ["Missing timestamp"]

        # Map asset_id
        record["asset_id"] = row_dict.get(mapping.asset_id, "").strip() if mapping.asset_id else None

        # Map status
        record["status"] = row_dict.get(mapping.status, "").strip() if mapping.status else None

        # Map event_type (auto-classify if not mapped)
        if mapping.event_type and row_dict.get(mapping.event_type):
            record["event_type"] = row_dict[mapping.event_type].strip().lower()
        else:
            record["event_type"] = _classify_event(record.get("status"), {})

        # Map metrics (all metric_columns)
        metrics = {}
        for col in mapping.metric_columns:
            val = row_dict.get(col, "").strip()
            if val:
                try:
                    metrics[col] = float(val.replace(",", "."))
                except ValueError:
                    metrics[col] = val
        record["metrics"] = metrics

        # Also classify based on metrics
        if record["event_type"] == "normal":
            record["event_type"] = _classify_event(record.get("status"), metrics)

        records.append(record)

    return records


def _parse_excel_content(file_bytes: bytes, ext: str, template: ParseTemplate) -> List[dict]:
    """Parse XLSX/XLS content using openpyxl or xlrd."""
    import openpyxl

    all_rows = []  # Keep all rows including header section for date extraction
    
    if ext == "xls":
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        ws = wb.sheet_by_index(0)
        for r in range(ws.nrows):
            all_rows.append([str(ws.cell_value(r, c)) for c in range(ws.ncols)])
    else:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            all_rows.append([str(c) if c is not None else "" for c in row])
        wb.close()

    if not all_rows:
        return []

    # Try to extract base date from header section (before skip_rows)
    base_date = None
    if template.skip_rows > 0:
        header_section = all_rows[:template.skip_rows]
        for row in header_section:
            for cell in row:
                if cell:
                    # Look for date patterns in header cells
                    cell_str = str(cell).strip()
                    # Try common date formats
                    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"]:
                        try:
                            dt = datetime.strptime(cell_str.split()[0] if ' ' in cell_str else cell_str, fmt.split()[0])
                            base_date = dt.strftime("%Y-%m-%d")
                            logger.debug(f"[Excel Parse] Found base date in header: {base_date}")
                            break
                        except ValueError:
                            continue
                    if base_date:
                        break
            if base_date:
                break

    # Skip rows to get to data
    rows = all_rows[template.skip_rows:] if template.skip_rows > 0 else all_rows
    if not rows:
        return []

    mapping = template.column_mapping
    headers = None
    data_start = 0
    if template.has_header:
        headers = [h.strip() for h in rows[0]]
        data_start = 1
    else:
        headers = [f"col_{i}" for i in range(len(rows[0]))]

    records = []
    for row_idx, row in enumerate(rows[data_start:], start=data_start + 1):
        if all(not cell.strip() for cell in row):
            continue

        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = val.strip()

        record = {"_row": row_idx}

        ts_val = row_dict.get(mapping.timestamp) if mapping.timestamp else None
        if ts_val:
            # Pass base_date for time-only values
            parsed_ts = _parse_timestamp(ts_val, template.timestamp_format, base_date)
            record["timestamp"] = parsed_ts
            if not parsed_ts:
                record["_errors"] = record.get("_errors", []) + [f"Invalid timestamp: {ts_val}"]
        else:
            record["timestamp"] = None
            if mapping.timestamp:
                record["_errors"] = record.get("_errors", []) + ["Missing timestamp"]

        record["asset_id"] = row_dict.get(mapping.asset_id, "").strip() if mapping.asset_id else None
        record["status"] = row_dict.get(mapping.status, "").strip() if mapping.status else None

        if mapping.event_type and row_dict.get(mapping.event_type):
            record["event_type"] = row_dict[mapping.event_type].strip().lower()
        else:
            record["event_type"] = _classify_event(record.get("status"), {})

        metrics = {}
        for col in mapping.metric_columns:
            val = row_dict.get(col, "").strip()
            if val:
                try:
                    metrics[col] = float(val.replace(",", "."))
                except ValueError:
                    metrics[col] = val
        record["metrics"] = metrics

        if record["event_type"] == "normal":
            record["event_type"] = _classify_event(record.get("status"), metrics)

        records.append(record)

    return records


# ======================== Endpoints ========================

@router.post("/upload")
async def upload_log_files(
    files: List[UploadFile] = File(...),
    job_id: Optional[str] = Form(None),
    current_user: dict = Depends(get_current_user),
):
    """Upload one or more log files. Creates or appends to an ingestion job."""
    _owner_only(current_user)

    is_new = job_id is None
    if is_new:
        job_id = str(uuid.uuid4())

    # Read all files into memory first (fast)
    pending_uploads = []
    for file in files:
        ext = _get_ext(file.filename or "")
        if ext not in ALLOWED_EXTENSIONS:
            continue  # Skip invalid silently in chunked mode
        content = await file.read()
        if len(content) > MAX_FILE_SIZE:
            continue

        if ext == "zip":
            try:
                with zipfile.ZipFile(io.BytesIO(content)) as zf:
                    for name in zf.namelist():
                        inner_ext = _get_ext(name)
                        if inner_ext not in ("csv", "txt", "log", "xlsx", "xls"):
                            continue
                        inner_content = zf.read(name)
                        file_id = str(uuid.uuid4())
                        storage_path = f"production-logs/{job_id}/{file_id}.{inner_ext}"
                        mime = {"csv": "text/csv", "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xls": "application/vnd.ms-excel"}.get(inner_ext, "text/plain")
                        pending_uploads.append((file_id, name, storage_path, inner_content, mime, inner_ext))
            except zipfile.BadZipFile:
                continue
        else:
            file_id = str(uuid.uuid4())
            storage_path = f"production-logs/{job_id}/{file_id}.{ext}"
            mime = {"csv": "text/csv", "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xls": "application/vnd.ms-excel"}.get(ext, "text/plain")
            pending_uploads.append((file_id, file.filename, storage_path, content, mime, ext))

    if not pending_uploads:
        if is_new:
            raise HTTPException(status_code=400, detail="No valid files found")
        return {"job_id": job_id, "files_uploaded": 0, "files": []}

    # Upload to R2 in parallel batches of 10
    uploaded_files = []
    BATCH_SIZE = 10
    for i in range(0, len(pending_uploads), BATCH_SIZE):
        batch = pending_uploads[i:i + BATCH_SIZE]
        tasks = [put_object_async(p[2], p[3], p[4]) for p in batch]
        await asyncio.gather(*tasks)
        for file_id, filename, storage_path, content, mime, ext in batch:
            uploaded_files.append({
                "file_id": file_id,
                "filename": filename,
                "storage_path": storage_path,
                "size": len(content),
                "extension": ext,
            })

    if is_new:
        # Create new job
        job = {
            "id": job_id,
            "status": "uploaded",
            "files": uploaded_files,
            "total_files": len(uploaded_files),
            "records_parsed": 0,
            "records_ingested": 0,
            "records_failed": 0,
            "parse_template": None,
            "created_by": current_user.get("id"),
            "created_by_name": current_user.get("name", "Unknown"),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.log_ingestion_jobs.insert_one(job)
    else:
        # Append files to existing job
        await db.log_ingestion_jobs.update_one(
            {"id": job_id},
            {
                "$push": {"files": {"$each": uploaded_files}},
                "$inc": {"total_files": len(uploaded_files)},
                "$set": {"updated_at": datetime.now(timezone.utc).isoformat()},
            }
        )

    if not uploaded_files:
        raise HTTPException(status_code=400, detail="No valid files found")

    # Create ingestion job
    job = {
        "id": job_id,
        "status": "uploaded",
        "files": uploaded_files,
        "total_files": len(uploaded_files),
        "records_parsed": 0,
        "records_ingested": 0,
        "records_failed": 0,
        "parse_template": None,
        "created_by": current_user.get("id"),
        "created_by_name": current_user.get("name", "Unknown"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.log_ingestion_jobs.insert_one(job)

    return {
        "job_id": job_id,
        "files_uploaded": len(uploaded_files),
        "files": [{"filename": f["filename"], "size": f["size"]} for f in uploaded_files],
    }


@router.post("/detect-columns")
async def detect_columns(
    job_id: str = Form(...),
    file_id: Optional[str] = Form(None),
    delimiter: str = Form(","),
    has_header: bool = Form(True),
    skip_rows: int = Form(0),
    current_user: dict = Depends(get_current_user),
):
    """Read the first file in a job and return detected columns/sample rows."""
    _owner_only(current_user)

    job = await db.log_ingestion_jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    # Pick target file
    target = None
    if file_id:
        target = next((f for f in job["files"] if f["file_id"] == file_id), None)
    if not target:
        target = job["files"][0]

    # Read file content
    from services.storage_service import get_object_async
    try:
        data, _ = await get_object_async(target["storage_path"])
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read file: {e}")

    file_ext = target.get("extension", "").lower()
    is_excel = file_ext in ("xlsx", "xls")

    if is_excel:
        # Parse Excel to rows
        import openpyxl
        if file_ext == "xls":
            import xlrd
            wb = xlrd.open_workbook(file_contents=data)
            ws = wb.sheet_by_index(0)
            all_rows = []
            for r in range(ws.nrows):
                all_rows.append([str(ws.cell_value(r, c)) for c in range(ws.ncols)])
        else:
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            all_rows = []
            for row in ws.iter_rows(values_only=True):
                all_rows.append([str(c) if c is not None else "" for c in row])
            wb.close()

        if skip_rows > 0:
            all_rows = all_rows[skip_rows:]

        rows = all_rows[:51]
        total_data_lines = len(all_rows) - (1 if has_header else 0)
    else:
        content = data.decode("utf-8", errors="replace")
        lines = content.splitlines()
        if skip_rows > 0:
            lines = lines[skip_rows:]
        if not lines:
            return {"columns": [], "sample_rows": [], "total_lines": 0}

        delim = delimiter
        if delim == "\\t":
            delim = "\t"

        reader = csv.reader(lines[:51], delimiter=delim)
        rows = list(reader)
        total_data_lines = len(lines) - (1 if has_header else 0)

    columns = []
    sample_rows = []
    if has_header and rows:
        columns = [h.strip() for h in rows[0]]
        sample_rows = rows[1:11]  # Next 10 rows
    elif rows:
        columns = [f"col_{i}" for i in range(len(rows[0]))]
        sample_rows = rows[:10]

    # Auto-detect timestamp and asset columns
    suggestions = {"timestamp": None, "asset_id": None, "status": None, "metrics": []}
    for col in columns:
        cl = col.lower()
        if any(w in cl for w in ["time", "date", "timestamp", "datetime"]):
            if not suggestions["timestamp"]:
                suggestions["timestamp"] = col
        elif any(w in cl for w in ["asset", "equipment", "tag", "unit", "machine", "device"]):
            if not suggestions["asset_id"]:
                suggestions["asset_id"] = col
        elif any(w in cl for w in ["status", "state", "condition"]):
            if not suggestions["status"]:
                suggestions["status"] = col
        else:
            # Check if column has numeric data
            numeric_count = 0
            for row in sample_rows[:5]:
                idx = columns.index(col) if col in columns else -1
                if idx >= 0 and idx < len(row):
                    try:
                        float(row[idx].replace(",", "."))
                        numeric_count += 1
                    except ValueError:
                        pass
            if numeric_count >= 2:
                suggestions["metrics"].append(col)

    return {
        "columns": columns,
        "sample_rows": [dict(zip(columns, row)) for row in sample_rows],
        "total_lines": total_data_lines,
        "suggestions": suggestions,
        "detected_delimiter": _detect_delimiter(rows[0][0] if rows else "") if not is_excel else ",",
    }


@router.post("/parse-preview")
async def parse_preview(
    job_id: str = Form(...),
    template_json: str = Form(...),
    current_user: dict = Depends(get_current_user),
):
    """Parse files using a template and return a preview of the first 100 records."""
    _owner_only(current_user)

    job = await db.log_ingestion_jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    try:
        template = ParseTemplate(**json.loads(template_json))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid template: {e}")

    from services.storage_service import get_object_async

    all_records = []
    file_stats = []

    for f in job["files"]:
        try:
            data, _ = await get_object_async(f["storage_path"])
            file_ext = f.get("extension", "").lower()
            if file_ext in ("xlsx", "xls"):
                records = _parse_excel_content(data, file_ext, template)
            else:
                content = data.decode("utf-8", errors="replace")
                records = _parse_csv_content(content, template)
            file_stats.append({
                "filename": f["filename"],
                "total_rows": len(records),
                "errors": sum(1 for r in records if r.get("_errors")),
            })
            all_records.extend(records)
        except Exception as e:
            file_stats.append({"filename": f["filename"], "total_rows": 0, "errors": 0, "error": str(e)})

    # Calculate stats
    total = len(all_records)
    with_errors = sum(1 for r in all_records if r.get("_errors"))
    with_timestamp = sum(1 for r in all_records if r.get("timestamp"))
    with_asset = sum(1 for r in all_records if r.get("asset_id"))
    success_rate = round(((total - with_errors) / total * 100) if total > 0 else 0, 1)

    # Save template to job
    await db.log_ingestion_jobs.update_one(
        {"id": job_id},
        {"$set": {
            "parse_template": json.loads(template_json),
            "records_parsed": total,
            "status": "previewed",
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }}
    )

    # Return preview (first 100)
    preview = all_records[:100]
    # Strip internal fields for response
    clean_preview = []
    for r in preview:
        entry = {k: v for k, v in r.items() if not k.startswith("_")}
        entry["_errors"] = r.get("_errors", [])
        entry["_row"] = r.get("_row")
        clean_preview.append(entry)

    return {
        "job_id": job_id,
        "total_records": total,
        "records_with_errors": with_errors,
        "records_with_timestamp": with_timestamp,
        "records_with_asset_id": with_asset,
        "success_rate": success_rate,
        "file_stats": file_stats,
        "preview": clean_preview,
        "event_summary": {
            "normal": sum(1 for r in all_records if r.get("event_type") == "normal"),
            "downtime": sum(1 for r in all_records if r.get("event_type") == "downtime"),
            "waste": sum(1 for r in all_records if r.get("event_type") == "waste"),
            "alarm": sum(1 for r in all_records if r.get("event_type") == "alarm"),
        }
    }


@router.post("/ingest")
async def ingest_logs(
    request: IngestRequest,
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user),
):
    """Confirm and start async ingestion of parsed logs into production_logs collection."""
    _owner_only(current_user)

    job = await db.log_ingestion_jobs.find_one({"id": request.job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if job.get("status") == "completed":
        raise HTTPException(status_code=400, detail="Job already ingested")
    if not job.get("parse_template"):
        raise HTTPException(status_code=400, detail="No parse template configured. Run preview first.")

    # Update status to processing
    await db.log_ingestion_jobs.update_one(
        {"id": request.job_id},
        {"$set": {"status": "processing", "updated_at": datetime.now(timezone.utc).isoformat()}}
    )

    # Run ingestion in background
    background_tasks.add_task(_run_ingestion, request.job_id, job)

    return {"job_id": request.job_id, "status": "processing", "message": "Ingestion started"}


@router.post("/batch-ingest")
async def batch_ingest_logs(
    request: BatchIngestRequest,
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user),
):
    """Parse and ingest multiple jobs at once using the same template."""
    _owner_only(current_user)

    if not request.job_ids:
        raise HTTPException(status_code=400, detail="No jobs selected")

    try:
        template = ParseTemplate(**request.template)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid template: {e}")

    template_dict = request.template
    started = []

    for job_id in request.job_ids:
        job = await db.log_ingestion_jobs.find_one({"id": job_id}, {"_id": 0})
        if not job:
            continue
        if job.get("status") == "completed":
            continue

        # Set template and mark as processing
        await db.log_ingestion_jobs.update_one(
            {"id": job_id},
            {"$set": {
                "parse_template": template_dict,
                "status": "processing",
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }}
        )
        job["parse_template"] = template_dict
        background_tasks.add_task(_run_ingestion, job_id, job)
        started.append(job_id)

    return {
        "started": len(started),
        "job_ids": started,
        "message": f"Batch ingestion started for {len(started)} job(s)",
    }


async def _run_ingestion(job_id: str, job: dict):
    """Background task: parse all files and insert into production_logs."""
    from services.storage_service import get_object_async

    template = ParseTemplate(**job["parse_template"])
    total_ingested = 0
    total_failed = 0

    try:
        for f in job["files"]:
            try:
                data, _ = await get_object_async(f["storage_path"])
                file_ext = f.get("extension", "").lower()
                if file_ext in ("xlsx", "xls"):
                    records = _parse_excel_content(data, file_ext, template)
                else:
                    content = data.decode("utf-8", errors="replace")
                    records = _parse_csv_content(content, template)

                docs = []
                for r in records:
                    if not r.get("timestamp"):
                        total_failed += 1
                        continue
                    doc = {
                        "id": str(uuid.uuid4()),
                        "timestamp": r["timestamp"],
                        "asset_id": r.get("asset_id") or "unknown",
                        "metrics": r.get("metrics", {}),
                        "status": r.get("status"),
                        "event_type": r.get("event_type", "normal"),
                        "source": {
                            "job_id": job_id,
                            "file_id": f["file_id"],
                            "filename": f["filename"],
                            "row": r.get("_row"),
                        },
                        "ingested_at": datetime.now(timezone.utc).isoformat(),
                    }
                    docs.append(doc)

                if docs:
                    await db.production_logs.insert_many(docs)
                    total_ingested += len(docs)

            except Exception as e:
                logger.error(f"[LogIngest] Failed to process {f['filename']}: {e}")
                total_failed += len(_parse_csv_content("", template))  # approximate

        # Update job as completed
        await db.log_ingestion_jobs.update_one(
            {"id": job_id},
            {"$set": {
                "status": "completed",
                "records_ingested": total_ingested,
                "records_failed": total_failed,
                "completed_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }}
        )
        logger.info(f"[LogIngest] Job {job_id} completed: {total_ingested} ingested, {total_failed} failed")

    except Exception as e:
        logger.error(f"[LogIngest] Job {job_id} failed: {e}")
        await db.log_ingestion_jobs.update_one(
            {"id": job_id},
            {"$set": {
                "status": "failed",
                "error": str(e),
                "records_ingested": total_ingested,
                "records_failed": total_failed,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }}
        )


@router.get("/jobs")
async def list_jobs(
    current_user: dict = Depends(get_current_user),
):
    """List all ingestion jobs."""
    _owner_only(current_user)

    jobs = await db.log_ingestion_jobs.find(
        {}, {"_id": 0}
    ).sort("created_at", -1).to_list(50)

    return {"jobs": jobs}


@router.get("/jobs/{job_id}")
async def get_job(
    job_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Get details of a specific ingestion job."""
    _owner_only(current_user)

    job = await db.log_ingestion_jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    return job


@router.delete("/jobs/{job_id}")
async def delete_job(
    job_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Delete a job and its ingested data."""
    _owner_only(current_user)

    job = await db.log_ingestion_jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    # Delete ingested records
    result = await db.production_logs.delete_many({"source.job_id": job_id})
    logger.info(f"[LogIngest] Deleted {result.deleted_count} records for job {job_id}")

    # Delete job
    await db.log_ingestion_jobs.delete_one({"id": job_id})

    return {"deleted_records": result.deleted_count, "message": "Job and data deleted"}


# ======================== Template Management ========================

@router.post("/templates")
async def save_template(
    request: SaveTemplateRequest,
    current_user: dict = Depends(get_current_user),
):
    """Save a parse template for reuse."""
    _owner_only(current_user)
    
    # Check for duplicate name
    existing = await db.log_parse_templates.find_one({"name": request.name})
    if existing:
        raise HTTPException(status_code=400, detail="Template with this name already exists")
    
    template_doc = {
        "id": str(uuid.uuid4()),
        "name": request.name,
        "description": request.description,
        "template": request.template,
        "column_aliases": request.column_aliases,
        "created_by": current_user.get("id"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "usage_count": 0,
    }
    
    await db.log_parse_templates.insert_one(template_doc)
    del template_doc["_id"]
    
    logger.info(f"[Templates] Created template '{request.name}' by user {current_user.get('id')}")
    return template_doc


@router.get("/templates")
async def list_templates(
    current_user: dict = Depends(get_current_user),
):
    """List all saved parse templates."""
    _owner_only(current_user)
    
    templates = await db.log_parse_templates.find(
        {}, {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    return {"templates": templates}


@router.get("/templates/{template_id}")
async def get_template(
    template_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Get a specific template."""
    _owner_only(current_user)
    
    template = await db.log_parse_templates.find_one({"id": template_id}, {"_id": 0})
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    return template


@router.put("/templates/{template_id}")
async def update_template(
    template_id: str,
    request: SaveTemplateRequest,
    current_user: dict = Depends(get_current_user),
):
    """Update an existing template."""
    _owner_only(current_user)
    
    template = await db.log_parse_templates.find_one({"id": template_id})
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    # Check for duplicate name (if changed)
    if request.name != template.get("name"):
        existing = await db.log_parse_templates.find_one({"name": request.name, "id": {"$ne": template_id}})
        if existing:
            raise HTTPException(status_code=400, detail="Template with this name already exists")
    
    await db.log_parse_templates.update_one(
        {"id": template_id},
        {"$set": {
            "name": request.name,
            "description": request.description,
            "template": request.template,
            "column_aliases": request.column_aliases,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }}
    )
    
    logger.info(f"[Templates] Updated template '{request.name}'")
    return {"message": "Template updated"}


@router.delete("/templates/{template_id}")
async def delete_template(
    template_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Delete a template."""
    _owner_only(current_user)
    
    result = await db.log_parse_templates.delete_one({"id": template_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    
    logger.info(f"[Templates] Deleted template {template_id}")
    return {"message": "Template deleted"}


@router.post("/batch-ingest-with-template")
async def batch_ingest_with_saved_template(
    request: BatchIngestWithTemplateRequest,
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user),
):
    """
    Batch ingest multiple jobs using a saved template with fuzzy column matching.
    This is the "train once, apply to many" workflow.
    """
    _owner_only(current_user)
    
    if not request.job_ids:
        raise HTTPException(status_code=400, detail="No jobs selected")
    
    # Get the saved template
    saved_template = await db.log_parse_templates.find_one({"id": request.template_id}, {"_id": 0})
    if not saved_template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    try:
        base_template = ParseTemplate(**saved_template["template"])
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid template configuration: {e}")
    
    column_aliases = saved_template.get("column_aliases", {})
    started = []
    match_reports = []
    
    for job_id in request.job_ids:
        job = await db.log_ingestion_jobs.find_one({"id": job_id}, {"_id": 0})
        if not job:
            continue
        if job.get("status") == "completed":
            continue
        
        # Get file columns to perform fuzzy matching
        from services.storage_service import get_object_async
        file_columns = []
        try:
            first_file = job["files"][0]
            data, _ = await get_object_async(first_file["storage_path"])
            file_ext = first_file.get("extension", "").lower()
            
            if file_ext in ("xlsx", "xls"):
                import openpyxl
                if file_ext == "xls":
                    import xlrd
                    wb = xlrd.open_workbook(file_contents=data)
                    ws = wb.sheet_by_index(0)
                    if ws.nrows > 0:
                        file_columns = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
                else:
                    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
                    ws = wb.active
                    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                    if first_row:
                        file_columns = [str(c).strip() if c else "" for c in first_row]
                    wb.close()
            else:
                content = data.decode("utf-8", errors="replace")
                lines = content.splitlines()
                if lines and base_template.has_header:
                    delimiter = base_template.delimiter
                    if delimiter == "\\t":
                        delimiter = "\t"
                    file_columns = [c.strip() for c in lines[0].split(delimiter)]
        except Exception as e:
            logger.warning(f"[BatchIngest] Could not read columns for job {job_id}: {e}")
            # Use original template columns if we can't read the file
            file_columns = []
        
        # Perform fuzzy column matching
        matched_columns = _fuzzy_match_columns(
            file_columns, 
            base_template.column_mapping,
            column_aliases
        )
        
        # Create adjusted template with matched columns
        adjusted_template = {
            "delimiter": base_template.delimiter,
            "has_header": base_template.has_header,
            "skip_rows": base_template.skip_rows,
            "timestamp_format": base_template.timestamp_format,
            "column_mapping": {
                "timestamp": matched_columns.get("timestamp"),
                "asset_id": matched_columns.get("asset_id"),
                "status": matched_columns.get("status"),
                "event_type": matched_columns.get("event_type"),
                "metric_columns": matched_columns.get("metric_columns", []),
            }
        }
        
        # Log the matching for debugging
        match_reports.append({
            "job_id": job_id,
            "filename": job["files"][0]["filename"] if job["files"] else "unknown",
            "file_columns": file_columns[:10],  # First 10 columns
            "matched": matched_columns,
        })
        
        # Set template and mark as processing
        await db.log_ingestion_jobs.update_one(
            {"id": job_id},
            {"$set": {
                "parse_template": adjusted_template,
                "template_used": saved_template["name"],
                "status": "processing",
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }}
        )
        job["parse_template"] = adjusted_template
        background_tasks.add_task(_run_ingestion, job_id, job)
        started.append(job_id)
    
    # Update template usage count
    await db.log_parse_templates.update_one(
        {"id": request.template_id},
        {"$inc": {"usage_count": len(started)}}
    )
    
    logger.info(f"[BatchIngest] Started {len(started)} jobs with template '{saved_template['name']}'")
    
    return {
        "started": len(started),
        "job_ids": started,
        "template_name": saved_template["name"],
        "match_reports": match_reports,
        "message": f"Batch ingestion started for {len(started)} job(s) using template '{saved_template['name']}'",
    }


@router.post("/preview-template-match")
async def preview_template_match(
    job_id: str = Form(...),
    template_id: str = Form(...),
    current_user: dict = Depends(get_current_user),
):
    """
    Preview how a saved template would match columns in a specific job.
    Useful for verifying the fuzzy matching before batch processing.
    """
    _owner_only(current_user)
    
    # Get job
    job = await db.log_ingestion_jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Get template
    saved_template = await db.log_parse_templates.find_one({"id": template_id}, {"_id": 0})
    if not saved_template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    try:
        base_template = ParseTemplate(**saved_template["template"])
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid template: {e}")
    
    column_aliases = saved_template.get("column_aliases", {})
    
    # Get file columns
    from services.storage_service import get_object_async
    file_columns = []
    sample_rows = []
    
    try:
        first_file = job["files"][0]
        data, _ = await get_object_async(first_file["storage_path"])
        file_ext = first_file.get("extension", "").lower()
        
        if file_ext in ("xlsx", "xls"):
            import openpyxl
            if file_ext == "xls":
                import xlrd
                wb = xlrd.open_workbook(file_contents=data)
                ws = wb.sheet_by_index(0)
                if ws.nrows > 0:
                    file_columns = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
                for r in range(1, min(ws.nrows, 6)):
                    row_data = {file_columns[c]: str(ws.cell_value(r, c)) for c in range(min(len(file_columns), ws.ncols))}
                    sample_rows.append(row_data)
            else:
                wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
                ws = wb.active
                rows_iter = ws.iter_rows(values_only=True)
                first_row = next(rows_iter, None)
                if first_row:
                    file_columns = [str(c).strip() if c else "" for c in first_row]
                for i, row in enumerate(rows_iter):
                    if i >= 5:
                        break
                    row_data = {file_columns[j]: str(row[j]) if row[j] is not None else "" for j in range(min(len(file_columns), len(row)))}
                    sample_rows.append(row_data)
                wb.close()
        else:
            content = data.decode("utf-8", errors="replace")
            lines = content.splitlines()
            delimiter = base_template.delimiter
            if delimiter == "\\t":
                delimiter = "\t"
            if lines and base_template.has_header:
                file_columns = [c.strip() for c in lines[0].split(delimiter)]
            for line in lines[1:6]:
                vals = line.split(delimiter)
                row_data = {file_columns[j]: vals[j].strip() if j < len(vals) else "" for j in range(len(file_columns))}
                sample_rows.append(row_data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read file: {e}")
    
    # Perform fuzzy matching
    matched_columns = _fuzzy_match_columns(
        file_columns,
        base_template.column_mapping,
        column_aliases
    )
    
    # Build match details showing what was matched
    match_details = []
    template_mapping = base_template.column_mapping
    
    if template_mapping.timestamp:
        match_details.append({
            "field": "timestamp",
            "template_column": template_mapping.timestamp,
            "matched_to": matched_columns.get("timestamp"),
            "match_type": "exact" if matched_columns.get("timestamp") == template_mapping.timestamp else "fuzzy",
            "success": matched_columns.get("timestamp") in file_columns,
        })
    
    if template_mapping.asset_id:
        match_details.append({
            "field": "asset_id",
            "template_column": template_mapping.asset_id,
            "matched_to": matched_columns.get("asset_id"),
            "match_type": "exact" if matched_columns.get("asset_id") == template_mapping.asset_id else "fuzzy",
            "success": matched_columns.get("asset_id") in file_columns,
        })
    
    if template_mapping.status:
        match_details.append({
            "field": "status",
            "template_column": template_mapping.status,
            "matched_to": matched_columns.get("status"),
            "match_type": "exact" if matched_columns.get("status") == template_mapping.status else "fuzzy",
            "success": matched_columns.get("status") in file_columns,
        })
    
    for metric in template_mapping.metric_columns:
        matched_metric = next((m for m in matched_columns.get("metric_columns", []) if _normalize_column_name(m) == _normalize_column_name(metric) or m == metric), None)
        match_details.append({
            "field": "metric",
            "template_column": metric,
            "matched_to": matched_metric,
            "match_type": "exact" if matched_metric == metric else "fuzzy",
            "success": matched_metric in file_columns if matched_metric else False,
        })
    
    return {
        "template_name": saved_template["name"],
        "file_columns": file_columns,
        "matched_columns": matched_columns,
        "match_details": match_details,
        "sample_rows": sample_rows,
        "all_matched": all(d["success"] for d in match_details if d["template_column"]),
    }


@router.get("/entries")
async def query_entries(
    asset_id: Optional[str] = None,
    event_type: Optional[str] = None,
    start: Optional[str] = None,
    end: Optional[str] = None,
    limit: int = 100,
    skip: int = 0,
    current_user: dict = Depends(get_current_user),
):
    """Query production log entries with filters."""
    _owner_only(current_user)

    query = {}
    if asset_id:
        query["asset_id"] = asset_id
    if event_type:
        query["event_type"] = event_type
    if start or end:
        ts_filter = {}
        if start:
            ts_filter["$gte"] = start
        if end:
            ts_filter["$lte"] = end
        query["timestamp"] = ts_filter

    entries = await db.production_logs.find(
        query, {"_id": 0}
    ).sort("timestamp", -1).skip(skip).limit(limit).to_list(limit)

    total = await db.production_logs.count_documents(query)

    return {"entries": entries, "total": total, "limit": limit, "skip": skip}


@router.get("/stats")
async def get_log_stats(
    current_user: dict = Depends(get_current_user),
):
    """Get overall production log statistics."""
    _owner_only(current_user)

    total = await db.production_logs.count_documents({})

    # Event type counts
    event_pipeline = [
        {"$group": {"_id": "$event_type", "count": {"$sum": 1}}}
    ]
    events = await db.production_logs.aggregate(event_pipeline).to_list(10)

    # Unique assets
    asset_pipeline = [
        {"$group": {"_id": "$asset_id"}},
        {"$count": "total"}
    ]
    assets = await db.production_logs.aggregate(asset_pipeline).to_list(1)

    # Jobs summary
    jobs_total = await db.log_ingestion_jobs.count_documents({})
    jobs_completed = await db.log_ingestion_jobs.count_documents({"status": "completed"})

    return {
        "total_entries": total,
        "unique_assets": assets[0]["total"] if assets else 0,
        "events": {e["_id"]: e["count"] for e in events if e["_id"]},
        "jobs_total": jobs_total,
        "jobs_completed": jobs_completed,
    }


# ======================== Aggregation Layer ========================

@router.post("/aggregate")
async def run_aggregation(
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user),
):
    """Build/rebuild asset_history aggregations from production_logs."""
    _owner_only(current_user)

    total = await db.production_logs.count_documents({})
    if total == 0:
        raise HTTPException(status_code=400, detail="No production logs to aggregate")

    background_tasks.add_task(_run_aggregation)
    return {"message": "Aggregation started", "total_source_records": total}


async def _run_aggregation():
    """Background: aggregate production_logs into asset_history (hourly buckets)."""
    try:
        pipeline = [
            {"$addFields": {
                "ts_parsed": {
                    "$cond": {
                        "if": {"$eq": [{"$type": "$timestamp"}, "string"]},
                        "then": {"$dateFromString": {"dateString": "$timestamp", "onError": None}},
                        "else": "$timestamp"
                    }
                }
            }},
            {"$match": {"ts_parsed": {"$ne": None}}},
            {"$group": {
                "_id": {
                    "asset_id": "$asset_id",
                    "hour": {"$dateToString": {"format": "%Y-%m-%dT%H:00:00", "date": "$ts_parsed"}},
                },
                "records": {"$sum": 1},
                "events": {"$push": "$event_type"},
                "statuses": {"$push": "$status"},
                "all_metrics": {"$push": "$metrics"},
            }},
            {"$sort": {"_id.hour": 1}},
        ]

        results = await db.production_logs.aggregate(pipeline).to_list(100000)

        if not results:
            logger.info("[Aggregation] No valid records to aggregate")
            return

        # Clear old aggregations
        await db.asset_history.delete_many({})

        docs = []
        for r in results:
            asset_id = r["_id"]["asset_id"]
            hour = r["_id"]["hour"]

            # Aggregate metrics
            metric_agg = {}
            for m in r["all_metrics"]:
                if not m:
                    continue
                for k, v in m.items():
                    if isinstance(v, (int, float)):
                        if k not in metric_agg:
                            metric_agg[k] = {"values": []}
                        metric_agg[k]["values"].append(v)

            metrics_summary = {}
            for k, data in metric_agg.items():
                vals = data["values"]
                if vals:
                    metrics_summary[k] = {
                        "avg": round(sum(vals) / len(vals), 2),
                        "min": round(min(vals), 2),
                        "max": round(max(vals), 2),
                        "count": len(vals),
                    }

            # Count events
            event_counts = {}
            for e in r["events"]:
                event_counts[e] = event_counts.get(e, 0) + 1

            docs.append({
                "id": str(uuid.uuid4()),
                "asset_id": asset_id,
                "hour": hour,
                "records": r["records"],
                "metrics": metrics_summary,
                "events": event_counts,
                "downtime_count": event_counts.get("downtime", 0),
                "alarm_count": event_counts.get("alarm", 0),
                "waste_count": event_counts.get("waste", 0),
                "aggregated_at": datetime.now(timezone.utc).isoformat(),
            })

        if docs:
            await db.asset_history.insert_many(docs)
            logger.info(f"[Aggregation] Created {len(docs)} hourly buckets")

    except Exception as e:
        logger.error(f"[Aggregation] Failed: {e}")


@router.get("/history")
async def get_asset_history(
    asset_id: Optional[str] = None,
    start: Optional[str] = None,
    end: Optional[str] = None,
    limit: int = 500,
    current_user: dict = Depends(get_current_user),
):
    """Query aggregated asset history."""
    _owner_only(current_user)

    query = {}
    if asset_id:
        query["asset_id"] = asset_id
    if start or end:
        hour_filter = {}
        if start:
            hour_filter["$gte"] = start
        if end:
            hour_filter["$lte"] = end
        query["hour"] = hour_filter

    docs = await db.asset_history.find(query, {"_id": 0}).sort("hour", 1).limit(limit).to_list(limit)
    return {"history": docs, "total": len(docs)}


@router.get("/assets")
async def list_log_assets(
    current_user: dict = Depends(get_current_user),
):
    """List all unique asset_ids in production logs."""
    _owner_only(current_user)

    pipeline = [
        {"$group": {"_id": "$asset_id", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    results = await db.production_logs.aggregate(pipeline).to_list(1000)
    return {"assets": [{"asset_id": r["_id"], "count": r["count"]} for r in results if r["_id"]]}


@router.get("/timeseries")
async def get_timeseries(
    asset_id: str,
    metric: Optional[str] = None,
    start: Optional[str] = None,
    end: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """Get time series data for charts — from asset_history aggregation."""
    _owner_only(current_user)

    query = {"asset_id": asset_id}
    if start or end:
        hour_filter = {}
        if start:
            hour_filter["$gte"] = start
        if end:
            hour_filter["$lte"] = end
        query["hour"] = hour_filter

    docs = await db.asset_history.find(query, {"_id": 0}).sort("hour", 1).limit(2000).to_list(2000)

    # Build time series
    timestamps = []
    metrics_data = {}
    events_data = {"downtime": [], "alarm": [], "waste": [], "normal": []}

    for d in docs:
        timestamps.append(d["hour"])
        for et in ["downtime", "alarm", "waste", "normal"]:
            events_data[et].append(d.get("events", {}).get(et, 0))
        for mk, mv in d.get("metrics", {}).items():
            if metric and mk != metric:
                continue
            if mk not in metrics_data:
                metrics_data[mk] = {"avg": [], "min": [], "max": []}
            metrics_data[mk]["avg"].append(mv.get("avg"))
            metrics_data[mk]["min"].append(mv.get("min"))
            metrics_data[mk]["max"].append(mv.get("max"))

    return {
        "asset_id": asset_id,
        "timestamps": timestamps,
        "metrics": metrics_data,
        "events": events_data,
        "total_points": len(timestamps),
    }


# ======================== AI-Assisted Parsing ========================

@router.post("/ai-parse")
async def ai_parse_file(
    job_id: str = Form(...),
    file_id: Optional[str] = Form(None),
    current_user: dict = Depends(get_current_user),
):
    """Use AI to analyze an unstructured log file and suggest column mappings."""
    _owner_only(current_user)

    vision_key = os.environ.get("OPENAI_VISION_KEY") or os.environ.get("OPENAI_API_KEY")
    if not vision_key:
        raise HTTPException(status_code=400, detail="OPENAI_API_KEY not configured")

    job = await db.log_ingestion_jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    target = None
    if file_id:
        target = next((f for f in job["files"] if f["file_id"] == file_id), None)
    if not target:
        target = job["files"][0]

    from services.storage_service import get_object_async
    try:
        data, _ = await get_object_async(target["storage_path"])
        file_ext = target.get("extension", "").lower()
        if file_ext in ("xlsx", "xls"):
            # Convert Excel to text for AI
            import openpyxl
            if file_ext == "xls":
                import xlrd
                wb = xlrd.open_workbook(file_contents=data)
                ws = wb.sheet_by_index(0)
                lines = []
                for r in range(min(ws.nrows, 30)):
                    lines.append(",".join(str(ws.cell_value(r, c)) for c in range(ws.ncols)))
                sample_text = "\n".join(lines)
            else:
                wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
                ws = wb.active
                lines = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 30:
                        break
                    lines.append(",".join(str(c) if c is not None else "" for c in row))
                sample_text = "\n".join(lines)
                wb.close()
        else:
            content = data.decode("utf-8", errors="replace")
            lines = content.splitlines()[:30]
            sample_text = "\n".join(lines)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read file: {e}")

    # Call OpenAI to analyze the log structure
    try:
        client = OpenAI(api_key=vision_key)
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": """You are a production log analyst. Analyze the sample data and return a JSON object with:
{
  "delimiter": "detected delimiter character (comma, semicolon, tab, pipe, or space)",
  "has_header": true/false,
  "skip_rows": number of rows to skip before data starts,
  "columns": ["list of column names"],
  "column_mapping": {
    "timestamp": "name of timestamp column or null",
    "asset_id": "name of asset/equipment ID column or null",
    "status": "name of status column or null",
    "metric_columns": ["names of numeric metric columns"]
  },
  "timestamp_format": "detected format like %Y-%m-%d %H:%M:%S or null",
  "notes": "brief description of the data structure"
}
Return ONLY valid JSON, no markdown."""},
                {"role": "user", "content": f"Analyze this production log file sample:\n\n{sample_text}"}
            ],
            max_completion_tokens=1000,
            temperature=0.1,
        )

        raw = response.choices[0].message.content.strip()
        # Strip markdown code fences if present
        if raw.startswith("```"):
            raw = raw.split("\n", 1)[1] if "\n" in raw else raw[3:]
            if raw.endswith("```"):
                raw = raw[:-3]
            raw = raw.strip()

        result = json.loads(raw)
        return {"success": True, "analysis": result, "sample_lines": len(sample_text.splitlines())}

    except json.JSONDecodeError:
        return {"success": False, "error": "AI returned invalid format", "raw": raw[:500]}
    except Exception as e:
        logger.error(f"[AI Parse] Failed: {e}")
        raise HTTPException(status_code=500, detail=f"AI analysis failed: {str(e)}")
