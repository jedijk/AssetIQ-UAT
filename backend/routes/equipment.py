"""
Equipment Hierarchy routes.

This module handles all equipment hierarchy operations including:
- Equipment Types (CRUD) - lines ~31-140
- Search & Utils (search, disciplines, criticality profiles) - lines ~143-265
- Node CRUD (get, create, update, delete) - lines ~265-865
- Node Operations (change level, reorder, move) - lines ~866-1200
- Criticality & Discipline Assignment - lines ~1198-1340
- Stats & Unstructured Items - lines ~1340-end

TODO: Consider splitting into modules:
- equipment_types.py - Equipment type CRUD
- equipment_nodes.py - Node CRUD operations  
- equipment_operations.py - Move, reorder, change level
- equipment_criticality.py - Criticality and discipline assignment
- equipment_utils.py - Search, stats, export
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Body
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, timezone
import uuid
import json
import logging
import base64
import io
from database import db, efm_service, installation_filter
from auth import get_current_user
from services.threat_score_service import recalculate_threat_scores_for_asset
from iso14224_models import (
    ISOLevel, ISO_LEVEL_ORDER, EQUIPMENT_TYPES, CRITICALITY_PROFILES, Discipline,
    get_valid_parent_level, get_valid_child_levels, is_valid_parent_child, normalize_level,
    EquipmentNodeCreate, EquipmentNodeUpdate, CriticalityAssignment, MoveNodeRequest,
    UnstructuredItemCreate, ParseEquipmentListRequest, AssignToHierarchyRequest,
    detect_equipment_type, EquipmentTypeCreate, EquipmentTypeUpdate, ISO_LEVEL_LABELS
)
from openai import OpenAI
from bson import ObjectId

logger = logging.getLogger(__name__)

router = APIRouter(tags=["Equipment Hierarchy"])

# =============================================================================
# SECTION 1: EQUIPMENT TYPES CRUD
# =============================================================================

@router.get("/equipment-hierarchy/types")
async def get_iso_equipment_types(
    current_user: dict = Depends(get_current_user)
):
    """Get all equipment types - merged from defaults and user-custom types."""
    # Get user's custom equipment types
    custom_types = await db.custom_equipment_types.find(
        {"created_by": current_user["id"]},
        {"_id": 0}
    ).to_list(100)
    
    # Merge: custom types override defaults by ID
    custom_ids = {t["id"] for t in custom_types}
    merged_types = [t for t in EQUIPMENT_TYPES if t["id"] not in custom_ids] + custom_types
    
    return {"equipment_types": merged_types}

@router.post("/equipment-hierarchy/types")
async def create_equipment_type(
    type_data: EquipmentTypeCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a custom equipment type."""
    # Check if ID already exists for this user
    existing = await db.custom_equipment_types.find_one(
        {"id": type_data.id, "created_by": current_user["id"]}
    )
    if existing:
        raise HTTPException(status_code=400, detail="Equipment type ID already exists")
    
    type_doc = {
        "id": type_data.id,
        "name": type_data.name,
        "iso_class": type_data.iso_class,
        "discipline": type_data.discipline,
        "icon": type_data.icon,
        "is_custom": True,
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.custom_equipment_types.insert_one(type_doc)
    type_doc.pop("_id", None)
    return type_doc

@router.patch("/equipment-hierarchy/types/{type_id}")
async def update_equipment_type(
    type_id: str,
    update: EquipmentTypeUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update a custom equipment type."""
    # Check if it's a custom type
    existing = await db.custom_equipment_types.find_one(
        {"id": type_id, "created_by": current_user["id"]}
    )
    
    if not existing:
        # It might be a default type - create a custom override
        default_type = next((t for t in EQUIPMENT_TYPES if t["id"] == type_id), None)
        if not default_type:
            raise HTTPException(status_code=404, detail="Equipment type not found")
        
        # Create custom override
        type_doc = {
            **default_type,
            "is_custom": True,
            "created_by": current_user["id"],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        update_data = {k: v for k, v in update.model_dump().items() if v is not None}
        type_doc.update(update_data)
        
        await db.custom_equipment_types.insert_one(type_doc)
        type_doc.pop("_id", None)
        return type_doc
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if update_data:
        await db.custom_equipment_types.update_one(
            {"id": type_id, "created_by": current_user["id"]},
            {"$set": update_data}
        )
    
    updated = await db.custom_equipment_types.find_one(
        {"id": type_id, "created_by": current_user["id"]},
        {"_id": 0}
    )
    return updated

@router.delete("/equipment-hierarchy/types/{type_id}")
async def delete_equipment_type(
    type_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a custom equipment type."""
    result = await db.custom_equipment_types.delete_one(
        {"id": type_id, "created_by": current_user["id"]}
    )
    if result.deleted_count == 0:
        # Check if it's a default type
        default_type = next((t for t in EQUIPMENT_TYPES if t["id"] == type_id), None)
        if default_type:
            raise HTTPException(status_code=400, detail="Cannot delete default equipment types")
        raise HTTPException(status_code=404, detail="Equipment type not found")
    return {"message": "Equipment type deleted"}


# =============================================================================
# SECTION 2: SEARCH & UTILITIES
# =============================================================================

@router.get("/equipment-hierarchy/disciplines")
async def get_disciplines():
    """Get all disciplines."""
    return {"disciplines": [d.value for d in Discipline]}

@router.get("/equipment-hierarchy/search")
async def search_equipment(
    q: str,
    limit: int = 10,
    current_user: dict = Depends(get_current_user)
):
    """Search equipment hierarchy by name."""
    if not q or len(q) < 2:
        return {"results": []}
    
    # Get user's role and assigned installations
    user_role = current_user.get("role", "viewer")
    assigned = current_user.get("assigned_installations", [])
    
    # Owners and admins can see ALL equipment - no filtering needed
    is_admin_or_owner = user_role in ["owner", "admin"]
    
    # First, do the name search
    search_filter = {
        "name": {"$regex": q, "$options": "i"}
    }
    
    # Search equipment nodes
    nodes = await db.equipment_nodes.find(
        search_filter,
        {"_id": 0, "id": 1, "name": 1, "level": 1, "parent_id": 1, "full_path": 1, "installation_id": 1}
    ).limit(limit if is_admin_or_owner else limit * 3).to_list(limit if is_admin_or_owner else limit * 3)
    
    # If user is NOT admin/owner and has assigned installations, filter results by hierarchy
    if not is_admin_or_owner and assigned:
        # Get IDs of assigned installations
        installations = await db.equipment_nodes.find(
            {"level": "installation", "name": {"$in": assigned}},
            {"_id": 0, "id": 1, "name": 1}
        ).to_list(100)
        installation_ids = {i["id"] for i in installations}
        installation_names = {i["name"] for i in installations}
        
        # Filter nodes that belong to assigned installations
        filtered_nodes = []
        for node in nodes:
            # Check if node is an installation itself
            if node.get("level") == "installation" and node.get("id") in installation_ids:
                filtered_nodes.append(node)
                continue
            
            # Check if node has installation_id
            if node.get("installation_id") in installation_ids:
                filtered_nodes.append(node)
                continue
            
            # Trace parent chain to find if it belongs to an assigned installation
            current = node
            depth = 0
            belongs_to_assigned = False
            while current.get("parent_id") and depth < 15:
                parent = await db.equipment_nodes.find_one(
                    {"id": current["parent_id"]},
                    {"_id": 0, "id": 1, "name": 1, "parent_id": 1, "level": 1}
                )
                if parent:
                    if parent.get("level") == "installation" and parent.get("id") in installation_ids:
                        belongs_to_assigned = True
                        break
                    current = parent
                else:
                    break
                depth += 1
            
            if belongs_to_assigned:
                filtered_nodes.append(node)
        
        nodes = filtered_nodes[:limit]
    else:
        nodes = nodes[:limit]
    
    # Build path for nodes without full_path
    for node in nodes:
        if not node.get("full_path") and not node.get("path"):
            # Build path from parent chain
            path_parts = [node["name"]]
            current = node
            depth = 0
            while current.get("parent_id") and depth < 10:
                parent = await db.equipment_nodes.find_one(
                    {"id": current["parent_id"]},
                    {"_id": 0, "name": 1, "parent_id": 1}
                )
                if parent:
                    path_parts.insert(0, parent["name"])
                    current = parent
                else:
                    break
                depth += 1
            node["path"] = " > ".join(path_parts)
        elif node.get("full_path"):
            node["path"] = node["full_path"]
    
    return {"results": nodes}

@router.get("/equipment-hierarchy/criticality-profiles")
async def get_criticality_profiles():
    """Get all criticality profiles."""
    return {"profiles": CRITICALITY_PROFILES}

@router.get("/equipment-hierarchy/iso-levels")
async def get_iso_levels():
    """Get ISO 14224 hierarchy levels with labels."""
    from iso14224_models import ISO_LEVEL_LABELS
    return {
        "levels": [level.value for level in ISO_LEVEL_ORDER],
        "labels": {level.value: ISO_LEVEL_LABELS.get(level, level.value) for level in ISO_LEVEL_ORDER},
        "hierarchy": {
            level.value: {
                "label": ISO_LEVEL_LABELS.get(level, level.value),
                "parent": get_valid_parent_level(level).value if get_valid_parent_level(level) else None,
                "children": [c.value for c in get_valid_child_levels(level)]
            }
            for level in ISO_LEVEL_ORDER
        }
    }


# =============================================================================
# SECTION 3: NODE CRUD OPERATIONS
# =============================================================================

@router.get("/equipment-hierarchy/nodes")
async def get_equipment_nodes(
    current_user: dict = Depends(get_current_user)
):
    """Get equipment hierarchy nodes filtered by user's assigned installations."""
    # Get user's installation filter data
    installation_ids = await installation_filter.get_user_installation_ids(current_user)
    
    # If no installations assigned, return empty list
    if not installation_ids:
        return {"nodes": []}
    
    # Get all equipment IDs under assigned installations (shared equipment - no created_by filter)
    equipment_ids = await installation_filter.get_all_equipment_ids_for_installations(
        installation_ids, current_user["id"]
    )
    
    if not equipment_ids:
        return {"nodes": []}
    
    # Get all nodes that belong to assigned installations (no created_by filter)
    nodes = await db.equipment_nodes.find(
        {"id": {"$in": list(equipment_ids)}},
        {"_id": 0}
    ).sort("sort_order", 1).to_list(5000)
    return {"nodes": nodes}


@router.get("/equipment-hierarchy/installations")
async def get_all_installations(
    current_user: dict = Depends(get_current_user)
):
    """Get all installation-level nodes across all users (for admin assignment)."""
    # Get nodes that are ONLY installations (top-level per ISO 14224)
    # Do NOT include plant_unit as those are children of installations
    nodes = await db.equipment_nodes.find(
        {"level": "installation"},
        {"_id": 0, "id": 1, "name": 1, "level": 1, "created_by": 1}
    ).sort("name", 1).to_list(500)
    
    # Deduplicate by name (in case same installation exists for multiple users)
    seen_names = set()
    unique_nodes = []
    for node in nodes:
        if node.get("name") not in seen_names:
            seen_names.add(node.get("name"))
            unique_nodes.append(node)
    
    return {"installations": unique_nodes}


@router.get("/equipment-hierarchy/export")
async def export_equipment_hierarchy_excel(
    current_user: dict = Depends(get_current_user)
):
    """Export equipment hierarchy to an Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Fetch all nodes
    nodes = await db.equipment_nodes.find(
        {"created_by": current_user["id"]},
        {"_id": 0}
    ).sort("sort_order", 1).to_list(5000)
    
    # Fetch equipment types for lookup
    equipment_types = await db.custom_equipment_types.find(
        {"created_by": current_user["id"]},
        {"_id": 0}
    ).to_list(100)
    
    # Merge with defaults
    all_types = {t["id"]: t["name"] for t in EQUIPMENT_TYPES}
    for t in equipment_types:
        all_types[t["id"]] = t["name"]
    
    # Build parent lookup for path generation
    node_lookup = {n["id"]: n for n in nodes}
    
    def get_path(node):
        """Generate full path from root to node."""
        if not isinstance(node, dict):
            return ""
        path_parts = [node.get("name", "")]
        current = node
        while isinstance(current, dict) and current.get("parent_id") and current["parent_id"] in node_lookup:
            current = node_lookup[current["parent_id"]]
            if isinstance(current, dict):
                path_parts.insert(0, current.get("name", ""))
            else:
                break
        return " > ".join(path_parts)
    
    def get_criticality_score(node):
        """Calculate total criticality score."""
        if not isinstance(node, dict):
            return 0
        crit = node.get("criticality")
        if not isinstance(crit, dict):
            return 0
        return sum([
            crit.get("safety", 0) or 0,
            crit.get("production", 0) or 0,
            crit.get("environmental", 0) or 0,
            crit.get("reputation", 0) or 0
        ])
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment Hierarchy"
    
    # Define headers
    headers = [
        "ID", "Name", "Level", "Parent", "Full Path", "Equipment Type",
        "Discipline", "Process Step", "Description",
        "Safety", "Production", "Environmental", "Reputation", "Total Criticality",
        "Created At"
    ]
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")  # Green
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Level colors for visual hierarchy
    level_colors = {
        "installation": "E8F5E9",
        "plant_unit": "C8E6C9",
        "section_system": "A5D6A7",
        "equipment_unit": "81C784",
        "equipment": "66BB6A",
        "subunit": "4CAF50",
        "maintainable_item": "43A047"
    }
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    for row_idx, node in enumerate(nodes, 2):
        # Skip corrupted nodes that are not dictionaries
        if not isinstance(node, dict):
            logger.warning(f"Skipping corrupted node at row {row_idx}: expected dict, got {type(node).__name__}")
            continue
            
        parent_name = ""
        if node.get("parent_id") and node["parent_id"] in node_lookup:
            parent_node = node_lookup[node["parent_id"]]
            if isinstance(parent_node, dict):
                parent_name = parent_node.get("name", "")
        
        equipment_type_name = ""
        if node.get("equipment_type_id"):
            equipment_type_name = all_types.get(node["equipment_type_id"], "")
        
        # Handle criticality - ensure it's a dict
        criticality = node.get("criticality")
        if not isinstance(criticality, dict):
            criticality = {}
        
        # Safely get level label
        level_label = ""
        node_level = node.get("level")
        if node_level:
            try:
                level_label = ISO_LEVEL_LABELS.get(ISOLevel(node_level), str(node_level))
            except (ValueError, TypeError):
                level_label = str(node_level)
        
        row_data = [
            node.get("id", ""),
            node.get("name", ""),
            level_label,
            parent_name,
            get_path(node),
            equipment_type_name,
            (node.get("discipline") or "").replace("_", " ").title(),
            node.get("process_step", ""),
            node.get("description", ""),
            criticality.get("safety", 0),
            criticality.get("production", 0),
            criticality.get("environmental", 0),
            criticality.get("reputation", 0),
            get_criticality_score(node),
            node.get("created_at", "")[:10] if node.get("created_at") else ""
        ]
        
        # Apply level-based coloring
        level_color = level_colors.get(node.get("level"), "FFFFFF")
        row_fill = PatternFill(start_color=level_color, end_color=level_color, fill_type="solid")
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            # Apply level coloring to first 3 columns
            if col <= 3:
                cell.fill = row_fill
    
    # Adjust column widths
    column_widths = [38, 25, 18, 20, 50, 20, 15, 20, 30, 8, 10, 12, 10, 14, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"equipment_hierarchy_{timestamp}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/equipment-hierarchy/nodes/{node_id}")
async def get_equipment_node(
    node_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get a specific equipment node."""
    # Equipment is shared - no created_by filter
    node = await db.equipment_nodes.find_one(
        {"id": node_id},
        {"_id": 0}
    )
    if not node:
        raise HTTPException(status_code=404, detail="Equipment node not found")
    return node

@router.post("/equipment-hierarchy/nodes")
async def create_equipment_node(
    node_data: EquipmentNodeCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a new equipment hierarchy node with ISO 14224 validation."""
    # Check if user is trying to create an installation (root node)
    # Only owners can create new installations
    if not node_data.parent_id and node_data.level == ISOLevel.INSTALLATION:
        if current_user.get("role") != "owner":
            raise HTTPException(
                status_code=403,
                detail="Only owners can create new installations"
            )
    
    # Check for duplicate name under the same parent (global check, not per-user)
    existing = await db.equipment_nodes.find_one({
        "name": node_data.name,
        "parent_id": node_data.parent_id
    })
    if existing:
        parent_info = f"under parent '{node_data.parent_id}'" if node_data.parent_id else "at root level"
        raise HTTPException(
            status_code=400, 
            detail=f"A node with name '{node_data.name}' already exists {parent_info}"
        )
    
    # Validate parent-child relationship if parent specified
    if node_data.parent_id:
        # Parent lookup should not filter by created_by since equipment is shared
        parent = await db.equipment_nodes.find_one(
            {"id": node_data.parent_id},
            {"_id": 0}
        )
        if not parent:
            raise HTTPException(status_code=400, detail="Parent node not found")
        
        parent_level = ISOLevel(parent["level"])
        if not is_valid_parent_child(parent_level, node_data.level):
            valid_children = get_valid_child_levels(parent_level)
            raise HTTPException(
                status_code=400, 
                detail=f"Invalid parent-child relationship. {parent_level.value} can only have {[c.value for c in valid_children]} as children"
            )
    else:
        # Root nodes must be installations
        if node_data.level != ISOLevel.INSTALLATION:
            raise HTTPException(
                status_code=400, 
                detail="Root nodes must be of level 'installation'"
            )
    
    node_id = str(uuid.uuid4())
    
    # Calculate sort_order - get max sort_order for siblings and add 1 (global, not per-user)
    max_sort = await db.equipment_nodes.find_one(
        {"parent_id": node_data.parent_id},
        sort=[("sort_order", -1)],
        projection={"sort_order": 1}
    )
    next_sort_order = (max_sort.get("sort_order", 0) if max_sort else 0) + 1
    
    # Inherit process_step from parent if not provided and at subunit/maintainable_item level
    inherited_process_step = node_data.process_step
    if not inherited_process_step and node_data.parent_id and node_data.level in [ISOLevel.SUBUNIT, ISOLevel.MAINTAINABLE_ITEM]:
        parent = await db.equipment_nodes.find_one(
            {"id": node_data.parent_id},
            {"process_step": 1}
        )
        if parent and parent.get("process_step"):
            inherited_process_step = parent["process_step"]
    
    node_doc = {
        "id": node_id,
        "name": node_data.name,
        "level": node_data.level.value,
        "parent_id": node_data.parent_id,
        "equipment_type_id": node_data.equipment_type_id,
        "description": node_data.description,
        "process_step": inherited_process_step,
        "criticality": None,
        "discipline": None,
        "sort_order": next_sort_order,
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.equipment_nodes.insert_one(node_doc)
    
    # Auto-generate EFMs if equipment_type_id is specified
    if node_data.equipment_type_id:
        try:
            efms = await efm_service.generate_efms_for_equipment(
                equipment_id=node_id,
                equipment_name=node_data.name,
                equipment_type_id=node_data.equipment_type_id
            )
            node_doc["efms_generated"] = len(efms)
            logger.info(f"Auto-generated {len(efms)} EFMs for equipment {node_id}")
        except Exception as e:
            logger.error(f"Failed to generate EFMs for {node_id}: {e}")
    
    # Remove MongoDB's _id before returning
    node_doc.pop("_id", None)
    return node_doc

@router.patch("/equipment-hierarchy/nodes/{node_id}")
async def update_equipment_node(
    node_id: str,
    update: EquipmentNodeUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update an equipment hierarchy node."""
    # Equipment is shared - no created_by filter for lookup
    node = await db.equipment_nodes.find_one(
        {"id": node_id}
    )
    if not node:
        raise HTTPException(status_code=404, detail="Equipment node not found")
    
    # Validate new parent if changing parent
    if update.parent_id is not None and update.parent_id != node.get("parent_id"):
        if update.parent_id:
            new_parent = await db.equipment_nodes.find_one(
                {"id": update.parent_id},
                {"_id": 0}
            )
            if not new_parent:
                raise HTTPException(status_code=400, detail="New parent node not found")
            
            parent_level = ISOLevel(new_parent["level"])
            child_level = ISOLevel(node["level"])
            if not is_valid_parent_child(parent_level, child_level):
                raise HTTPException(
                    status_code=400, 
                    detail="Invalid parent-child relationship per ISO 14224"
                )
            
            # Check for circular references
            current_parent = update.parent_id
            while current_parent:
                if current_parent == node_id:
                    raise HTTPException(status_code=400, detail="Circular reference detected")
                parent_node = await db.equipment_nodes.find_one({"id": current_parent})
                current_parent = parent_node.get("parent_id") if parent_node else None
        else:
            # Removing parent - node must be installation level
            if node["level"] != ISOLevel.INSTALLATION.value:
                raise HTTPException(
                    status_code=400, 
                    detail="Only installations can be root nodes"
                )
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.equipment_nodes.update_one(
            {"id": node_id},
            {"$set": update_data}
        )
        
        # Sync EFMs if equipment_type_id changed
        if update.equipment_type_id is not None:
            old_type = node.get("equipment_type_id")
            new_type = update.equipment_type_id
            if old_type != new_type:
                try:
                    sync_result = await efm_service.sync_efms_for_equipment(
                        equipment_id=node_id,
                        equipment_name=node.get("name"),
                        new_equipment_type_id=new_type,
                        old_equipment_type_id=old_type
                    )
                    logger.info(f"Synced EFMs for {node_id}: {sync_result}")
                except Exception as e:
                    logger.error(f"Failed to sync EFMs for {node_id}: {e}")
    
    updated = await db.equipment_nodes.find_one({"id": node_id}, {"_id": 0})
    return updated


@router.get("/equipment-hierarchy/nodes/{node_id}/deletion-impact")
async def get_deletion_impact(
    node_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get impact analysis for deleting an equipment node - shows related tasks, actions, and investigations."""
    # Find the node
    node = await db.equipment_nodes.find_one({"id": node_id})
    if not node:
        raise HTTPException(status_code=404, detail="Equipment node not found")
    
    # Get all children recursively
    async def get_children_ids(parent_id):
        children = await db.equipment_nodes.find(
            {"parent_id": parent_id},
            {"_id": 0, "id": 1, "name": 1}
        ).to_list(1000)
        all_items = [{"id": c["id"], "name": c["name"]} for c in children]
        for child in children:
            all_items.extend(await get_children_ids(child["id"]))
        return all_items
    
    children = await get_children_ids(node_id)
    all_equipment_ids = [node_id] + [c["id"] for c in children]
    all_equipment_names = [node.get("name")] + [c["name"] for c in children]
    
    # Find impacted tasks
    impacted_tasks = await db.task_instances.find(
        {"equipment_id": {"$in": all_equipment_ids}, "status": {"$ne": "completed"}},
        {"_id": 0, "id": 1, "title": 1, "status": 1, "equipment_name": 1}
    ).to_list(100)
    
    # Find impacted actions
    impacted_actions = await db.central_actions.find(
        {"equipment_id": {"$in": all_equipment_ids}, "status": {"$nin": ["completed", "closed"]}},
        {"_id": 0, "id": 1, "title": 1, "status": 1, "equipment_name": 1}
    ).to_list(100)
    
    # Find impacted investigations/threats
    impacted_investigations = await db.threats.find(
        {"asset_id": {"$in": all_equipment_ids}, "status": {"$nin": ["closed", "mitigated"]}},
        {"_id": 0, "id": 1, "title": 1, "status": 1, "asset": 1}
    ).to_list(100)
    
    # Find task plans linked to this equipment
    impacted_plans = await db.task_plans.find(
        {"equipment_id": {"$in": all_equipment_ids}, "is_active": True},
        {"_id": 0, "id": 1, "task_template_name": 1, "equipment_name": 1}
    ).to_list(100)
    
    return {
        "node": {
            "id": node_id,
            "name": node.get("name"),
            "level": node.get("level")
        },
        "children_count": len(children),
        "children": children[:10],  # First 10 for display
        "impact": {
            "tasks": {
                "count": len(impacted_tasks),
                "items": impacted_tasks[:5],
                "will_be": "orphaned (equipment reference cleared)"
            },
            "actions": {
                "count": len(impacted_actions),
                "items": impacted_actions[:5],
                "will_be": "orphaned (equipment reference cleared)"
            },
            "investigations": {
                "count": len(impacted_investigations),
                "items": impacted_investigations[:5],
                "will_be": "orphaned (asset reference cleared)"
            },
            "task_plans": {
                "count": len(impacted_plans),
                "items": impacted_plans[:5],
                "will_be": "deactivated"
            }
        },
        "total_impacted": len(impacted_tasks) + len(impacted_actions) + len(impacted_investigations) + len(impacted_plans)
    }


@router.delete("/equipment-hierarchy/nodes/{node_id}")
async def delete_equipment_node(
    node_id: str,
    cascade: bool = False,
    current_user: dict = Depends(get_current_user)
):
    """Delete an equipment node and optionally cascade to related items. Installations can only be deleted by Owner."""
    # Equipment is shared - no created_by filter
    node = await db.equipment_nodes.find_one(
        {"id": node_id}
    )
    if not node:
        raise HTTPException(status_code=404, detail="Equipment node not found")
    
    # Only Owner can delete installations
    if node.get("level") == "installation":
        if current_user.get("role") != "owner":
            raise HTTPException(
                status_code=403, 
                detail="Only Owner can delete installations. Contact your system administrator."
            )
    
    # Get all children recursively (global, not per-user)
    async def get_children_ids(parent_id):
        children = await db.equipment_nodes.find(
            {"parent_id": parent_id},
            {"_id": 0, "id": 1}
        ).to_list(1000)
        all_ids = [c["id"] for c in children]
        for child in children:
            all_ids.extend(await get_children_ids(child["id"]))
        return all_ids
    
    children_ids = await get_children_ids(node_id)
    all_ids = [node_id] + children_ids
    
    # Clear equipment references in related items (orphan them rather than delete)
    orphaned_tasks = await db.task_instances.update_many(
        {"equipment_id": {"$in": all_ids}},
        {"$set": {"equipment_id": None, "equipment_name": f"[Deleted: {node.get('name')}]"}}
    )
    
    orphaned_actions = await db.central_actions.update_many(
        {"equipment_id": {"$in": all_ids}},
        {"$set": {"equipment_id": None, "equipment_name": f"[Deleted: {node.get('name')}]"}}
    )
    
    orphaned_investigations = await db.threats.update_many(
        {"asset_id": {"$in": all_ids}},
        {"$set": {"asset_id": None, "asset": f"[Deleted: {node.get('name')}]"}}
    )
    
    # Deactivate task plans linked to this equipment
    deactivated_plans = await db.task_plans.update_many(
        {"equipment_id": {"$in": all_ids}},
        {"$set": {"is_active": False, "deactivation_reason": f"Equipment deleted: {node.get('name')}"}}
    )
    
    # Delete the equipment nodes
    result = await db.equipment_nodes.delete_many(
        {"id": {"$in": all_ids}}
    )
    
    logger.info(f"Deleted equipment node {node_id} ({node.get('name')}) and {len(children_ids)} children. " 
                f"Orphaned: {orphaned_tasks.modified_count} tasks, {orphaned_actions.modified_count} actions, "
                f"{orphaned_investigations.modified_count} investigations. Deactivated {deactivated_plans.modified_count} plans.")
    
    return {
        "message": f"Deleted {result.deleted_count} nodes", 
        "deleted_ids": all_ids,
        "orphaned": {
            "tasks": orphaned_tasks.modified_count,
            "actions": orphaned_actions.modified_count,
            "investigations": orphaned_investigations.modified_count,
            "plans_deactivated": deactivated_plans.modified_count
        }
    }


# =============================================================================
# SECTION 4: NODE OPERATIONS (Change Level, Reorder, Move)
# =============================================================================

class ChangeLevelRequest(BaseModel):
    new_level: ISOLevel
    new_parent_id: Optional[str] = None  # Required when demoting, optional when promoting


@router.post("/equipment-hierarchy/nodes/{node_id}/change-level")
async def change_node_level(
    node_id: str,
    request: ChangeLevelRequest,
    current_user: dict = Depends(get_current_user)
):
    """Change the hierarchy level of a node (promote or demote)."""
    # Equipment is shared - no created_by filter
    node = await db.equipment_nodes.find_one(
        {"id": node_id}
    )
    if not node:
        raise HTTPException(status_code=404, detail="Equipment node not found")
    
    current_level = normalize_level(ISOLevel(node["level"]))
    new_level = normalize_level(request.new_level)
    
    current_idx = ISO_LEVEL_ORDER.index(current_level)
    new_idx = ISO_LEVEL_ORDER.index(new_level)
    
    # Validate level change
    if new_idx == current_idx:
        raise HTTPException(status_code=400, detail="Node is already at this level")
    
    is_promoting = new_idx < current_idx  # Moving up in hierarchy
    is_demoting = new_idx > current_idx   # Moving down in hierarchy
    
    # Get current parent (no created_by filter - equipment is shared)
    current_parent = None
    if node.get("parent_id"):
        current_parent = await db.equipment_nodes.find_one(
            {"id": node["parent_id"]}
        )
    
    if is_promoting:
        # When promoting, the node becomes a sibling of its current parent
        # The new parent is the grandparent (parent of current parent)
        if not current_parent:
            raise HTTPException(status_code=400, detail="Cannot promote a root node")
        
        new_parent_id = current_parent.get("parent_id")  # Grandparent (can be None for installation)
        
        # Validate the new level is correct for the new parent
        if new_parent_id:
            grandparent = await db.equipment_nodes.find_one(
                {"id": new_parent_id}
            )
            if grandparent:
                grandparent_level = normalize_level(ISOLevel(grandparent["level"]))
                if not is_valid_parent_child(grandparent_level, new_level):
                    raise HTTPException(
                        status_code=400, 
                        detail=f"Cannot promote to {new_level.value}. Invalid parent-child relationship."
                    )
        else:
            # Promoting to root level - must be installation
            if new_level != ISOLevel.INSTALLATION:
                raise HTTPException(status_code=400, detail="Only installations can be root nodes")
        
    else:  # is_demoting
        # When demoting, user must specify a new parent
        if not request.new_parent_id:
            raise HTTPException(status_code=400, detail="Must specify new_parent_id when demoting")
        
        new_parent = await db.equipment_nodes.find_one(
            {"id": request.new_parent_id}
        )
        if not new_parent:
            raise HTTPException(status_code=400, detail="New parent node not found")
        
        parent_level = normalize_level(ISOLevel(new_parent["level"]))
        if not is_valid_parent_child(parent_level, new_level):
            raise HTTPException(
                status_code=400, 
                detail=f"Cannot demote to {new_level.value} under {parent_level.value}"
            )
        
        new_parent_id = request.new_parent_id
    
    # Check for duplicate name at new location (global check)
    existing = await db.equipment_nodes.find_one({
        "name": node["name"],
        "parent_id": new_parent_id,
        "id": {"$ne": node_id}
    })
    if existing:
        raise HTTPException(
            status_code=400, 
            detail=f"A node with name '{node['name']}' already exists at the target location"
        )
    
    # Update the node
    await db.equipment_nodes.update_one(
        {"id": node_id},
        {"$set": {
            "level": new_level.value,
            "parent_id": new_parent_id,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # If promoting, also need to update children's parent to maintain hierarchy
    # Children of this node stay as children
    
    updated = await db.equipment_nodes.find_one({"id": node_id}, {"_id": 0})
    action = "promoted" if is_promoting else "demoted"
    return {
        "message": f"Node {action} to {new_level.value}",
        "node": updated
    }


class ReorderRequest(BaseModel):
    direction: str  # "up" or "down"


@router.post("/equipment-hierarchy/nodes/{node_id}/reorder")
async def reorder_equipment_node(
    node_id: str,
    request: ReorderRequest,
    current_user: dict = Depends(get_current_user)
):
    """Reorder a node among its siblings (move up or down in the list)."""
    if request.direction not in ["up", "down"]:
        raise HTTPException(status_code=400, detail="Direction must be 'up' or 'down'")
    
    # Equipment is shared - no created_by filter
    node = await db.equipment_nodes.find_one(
        {"id": node_id}
    )
    if not node:
        raise HTTPException(status_code=404, detail="Equipment node not found")
    
    current_sort = node.get("sort_order", 0)
    
    # Get all siblings (same parent, global)
    siblings = await db.equipment_nodes.find(
        {"parent_id": node["parent_id"]},
        {"_id": 0}
    ).sort("sort_order", 1).to_list(1000)
    
    if len(siblings) <= 1:
        raise HTTPException(status_code=400, detail="No siblings to reorder with")
    
    # Find current index
    current_idx = next((i for i, s in enumerate(siblings) if s["id"] == node_id), -1)
    if current_idx == -1:
        raise HTTPException(status_code=400, detail="Node not found in siblings")
    
    # Calculate target index
    if request.direction == "up":
        if current_idx == 0:
            raise HTTPException(status_code=400, detail="Already at the top")
        target_idx = current_idx - 1
    else:  # down
        if current_idx == len(siblings) - 1:
            raise HTTPException(status_code=400, detail="Already at the bottom")
        target_idx = current_idx + 1
    
    # Swap sort_order values
    target_node = siblings[target_idx]
    target_sort = target_node.get("sort_order", target_idx)
    
    # Update both nodes
    await db.equipment_nodes.update_one(
        {"id": node_id},
        {"$set": {"sort_order": target_sort, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    await db.equipment_nodes.update_one(
        {"id": target_node["id"]},
        {"$set": {"sort_order": current_sort, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": f"Node moved {request.direction}", "new_sort_order": target_sort}


class ReorderToPositionRequest(BaseModel):
    target_node_id: str  # The node to position relative to
    position: str  # "before" or "after"
    new_parent_id: Optional[str] = None  # If moving to a different parent


@router.post("/equipment-hierarchy/nodes/{node_id}/reorder-to")
async def reorder_node_to_position(
    node_id: str,
    request: ReorderToPositionRequest,
    current_user: dict = Depends(get_current_user)
):
    """Reorder a node to a specific position relative to another node via drag-and-drop."""
    if request.position not in ["before", "after"]:
        raise HTTPException(status_code=400, detail="Position must be 'before' or 'after'")
    
    # Equipment is shared - no created_by filter
    node = await db.equipment_nodes.find_one(
        {"id": node_id}
    )
    if not node:
        raise HTTPException(status_code=404, detail="Node not found")
    
    target = await db.equipment_nodes.find_one(
        {"id": request.target_node_id}
    )
    if not target:
        raise HTTPException(status_code=404, detail="Target node not found")
    
    # Determine the new parent - use target's parent if not specified
    new_parent_id = request.new_parent_id if request.new_parent_id is not None else target.get("parent_id")
    
    # If moving to a different parent, validate the level relationship
    if new_parent_id != node.get("parent_id"):
        if new_parent_id:
            new_parent = await db.equipment_nodes.find_one(
                {"id": new_parent_id}
            )
            if not new_parent:
                raise HTTPException(status_code=400, detail="New parent not found")
            
            parent_level = ISOLevel(new_parent["level"])
            child_level = ISOLevel(node["level"])
            if not is_valid_parent_child(parent_level, child_level):
                raise HTTPException(
                    status_code=400,
                    detail=f"Cannot place {child_level.value} under {parent_level.value}"
                )
        else:
            # Moving to root level - must be installation
            if node["level"] != "installation":
                raise HTTPException(status_code=400, detail="Only installations can be at root level")
    
    # Get all siblings at the target location (same parent as target)
    # Equipment is shared - no created_by filter
    siblings = await db.equipment_nodes.find(
        {"parent_id": new_parent_id},
        {"_id": 0}
    ).sort("sort_order", 1).to_list(1000)
    
    # Remove the dragged node from siblings if it's in the same parent
    siblings = [s for s in siblings if s["id"] != node_id]
    
    # Find target's position
    target_idx = next((i for i, s in enumerate(siblings) if s["id"] == request.target_node_id), -1)
    
    if target_idx == -1:
        # Target not in siblings (was the dragged node itself), just append
        insert_idx = len(siblings)
    elif request.position == "before":
        insert_idx = target_idx
    else:  # after
        insert_idx = target_idx + 1
    
    # Reassign sort_order for all siblings
    for i, sibling in enumerate(siblings):
        new_sort = i if i < insert_idx else i + 1
        if sibling.get("sort_order", 0) != new_sort:
            await db.equipment_nodes.update_one(
                {"id": sibling["id"]},
                {"$set": {"sort_order": new_sort, "updated_at": datetime.now(timezone.utc).isoformat()}}
            )
    
    # Update the moved node with new sort_order and possibly new parent
    await db.equipment_nodes.update_one(
        {"id": node_id},
        {"$set": {
            "sort_order": insert_idx,
            "parent_id": new_parent_id,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    updated = await db.equipment_nodes.find_one({"id": node_id}, {"_id": 0})
    return {
        "message": f"Node moved {request.position} target",
        "node": updated
    }


@router.post("/equipment-hierarchy/nodes/{node_id}/move")
async def move_equipment_node(
    node_id: str,
    move_request: MoveNodeRequest,
    current_user: dict = Depends(get_current_user)
):
    """Move a node to a new parent with ISO 14224 validation."""
    # Equipment is shared - no created_by filter
    node = await db.equipment_nodes.find_one(
        {"id": node_id}
    )
    if not node:
        raise HTTPException(status_code=404, detail="Equipment node not found")
    
    # Equipment is shared - no created_by filter
    new_parent = await db.equipment_nodes.find_one(
        {"id": move_request.new_parent_id}
    )
    if not new_parent:
        raise HTTPException(status_code=400, detail="New parent node not found")
    
    # Check for duplicate name under the new parent (no created_by filter - equipment is shared)
    existing = await db.equipment_nodes.find_one({
        "name": node["name"],
        "parent_id": move_request.new_parent_id,
        "id": {"$ne": node_id}  # Exclude the node itself
    })
    if existing:
        raise HTTPException(
            status_code=400, 
            detail=f"A node with name '{node['name']}' already exists under the target parent"
        )
    
    # Validate the move per ISO 14224
    parent_level = ISOLevel(new_parent["level"])
    child_level = ISOLevel(node["level"])
    
    if not is_valid_parent_child(parent_level, child_level):
        valid_children = get_valid_child_levels(parent_level)
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot move {child_level.value} under {parent_level.value}. Valid children: {[c.value for c in valid_children]}"
        )
    
    # Update the node
    await db.equipment_nodes.update_one(
        {"id": node_id},
        {"$set": {
            "parent_id": move_request.new_parent_id,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    updated = await db.equipment_nodes.find_one({"id": node_id}, {"_id": 0})
    return updated

@router.post("/equipment-hierarchy/nodes/{node_id}/criticality")
async def assign_criticality(
    node_id: str,
    assignment: CriticalityAssignment,
    current_user: dict = Depends(get_current_user)
):
    """Assign criticality to an equipment node using 4-dimension model (Safety, Production, Environmental, Reputation)."""
    # Equipment is shared - no created_by filter
    node = await db.equipment_nodes.find_one(
        {"id": node_id}
    )
    if not node:
        raise HTTPException(status_code=404, detail="Equipment node not found")
    
    # Check if all 4 dimensions are None/0 - if so, clear criticality
    has_any_dimension = (
        (assignment.safety_impact and assignment.safety_impact > 0) or
        (assignment.production_impact and assignment.production_impact > 0) or
        (assignment.environmental_impact and assignment.environmental_impact > 0) or
        (assignment.reputation_impact and assignment.reputation_impact > 0)
    )
    
    if not has_any_dimension:
        await db.equipment_nodes.update_one(
            {"id": node_id},
            {"$set": {
                "criticality": None,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        updated = await db.equipment_nodes.find_one({"id": node_id}, {"_id": 0})
        return updated
    
    # Build 4-dimension criticality data
    safety = assignment.safety_impact or 0
    production = assignment.production_impact or 0
    environmental = assignment.environmental_impact or 0
    reputation = assignment.reputation_impact or 0
    
    # Calculate overall criticality level based on max dimension
    max_impact = max(safety, production, environmental, reputation)
    
    # Determine legacy level for backwards compatibility
    if safety >= 4 or max_impact == 5:
        level = "safety_critical"
        color = "#EF4444"  # Red
    elif production >= 4 or max_impact >= 4:
        level = "production_critical"
        color = "#F97316"  # Orange
    elif max_impact >= 3:
        level = "medium"
        color = "#EAB308"  # Yellow
    else:
        level = "low"
        color = "#22C55E"  # Green
    
    criticality_data = {
        # 4-dimension model
        "safety_impact": safety,
        "production_impact": production,
        "environmental_impact": environmental,
        "reputation_impact": reputation,
        # Derived values for backwards compatibility
        "level": level,
        "color": color,
        "max_impact": max_impact,
        # Legacy fields preserved
        "profile_id": assignment.profile_id,
        "fatality_risk": assignment.fatality_risk or 0,
        "production_loss_per_day": assignment.production_loss_per_day or 0,
        "failure_probability": assignment.failure_probability or 0,
        "downtime_days": assignment.downtime_days or 0,
    }
    
    # Calculate risk score weighted by dimensions
    risk_score = (
        (safety * 25) +  # Safety has highest weight
        (production * 20) +
        (environmental * 15) +
        (reputation * 10)
    )
    criticality_data["risk_score"] = round(risk_score, 2)
    
    await db.equipment_nodes.update_one(
        {"id": node_id},
        {"$set": {
            "criticality": criticality_data,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Recalculate risk scores for all threats linked to this asset
    asset_name = node.get("name")
    updated_threats = 0
    if asset_name:
        updated_threats = await recalculate_threat_scores_for_asset(
            asset_name, 
            current_user["id"], 
            criticality_data,
            node_id  # Pass node_id to also find threats by linked_equipment_id
        )
        logger.info(f"Updated {updated_threats} threat scores after criticality change for {asset_name}")
    
    updated = await db.equipment_nodes.find_one({"id": node_id}, {"_id": 0})
    # Include count of updated threats in response
    updated["threats_updated"] = updated_threats if asset_name else 0
    return updated


# =============================================================================
# SECTION 5: DISCIPLINE & CRITICALITY ASSIGNMENT
# =============================================================================

@router.post("/equipment-hierarchy/nodes/{node_id}/discipline")
async def assign_discipline(
    node_id: str,
    discipline: str,
    current_user: dict = Depends(get_current_user)
):
    """Assign discipline to an equipment node."""
    # Equipment is shared - no created_by filter
    node = await db.equipment_nodes.find_one(
        {"id": node_id}
    )
    if not node:
        raise HTTPException(status_code=404, detail="Equipment node not found")
    
    # Validate discipline
    try:
        Discipline(discipline)  # Validate discipline enum
    except ValueError:
        raise HTTPException(
            status_code=400, 
            detail=f"Invalid discipline. Valid options: {[d.value for d in Discipline]}"
        )
    
    await db.equipment_nodes.update_one(
        {"id": node_id},
        {"$set": {
            "discipline": discipline,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    updated = await db.equipment_nodes.find_one({"id": node_id}, {"_id": 0})
    return updated


# =============================================================================
# SECTION 6: STATS & UNSTRUCTURED ITEMS
# =============================================================================

@router.get("/equipment-hierarchy/stats")
async def get_hierarchy_stats(
    current_user: dict = Depends(get_current_user)
):
    """Get statistics about the equipment hierarchy, filtered by assigned installations."""
    user_id = current_user["id"]
    
    # Get user's installation filter data
    installation_ids = await installation_filter.get_user_installation_ids(current_user)
    
    # If no installations assigned, return zeros
    if not installation_ids:
        level_counts = {level.value: 0 for level in ISO_LEVEL_ORDER}
        return {
            "total_nodes": 0,
            "by_level": level_counts,
            "by_criticality": {
                "safety_critical": 0,
                "production_critical": 0,
                "medium": 0,
                "low": 0,
                "unassigned": 0
            }
        }
    
    # Get all equipment IDs under assigned installations (shared - no created_by)
    equipment_ids = await installation_filter.get_all_equipment_ids_for_installations(
        installation_ids, user_id
    )
    
    if not equipment_ids:
        level_counts = {level.value: 0 for level in ISO_LEVEL_ORDER}
        return {
            "total_nodes": 0,
            "by_level": level_counts,
            "by_criticality": {
                "safety_critical": 0,
                "production_critical": 0,
                "medium": 0,
                "low": 0,
                "unassigned": 0
            }
        }
    
    # No created_by filter - equipment is shared
    base_query = {"id": {"$in": list(equipment_ids)}}
    
    total_nodes = await db.equipment_nodes.count_documents(base_query)
    
    # Count by level
    level_counts = {}
    for level in ISO_LEVEL_ORDER:
        count = await db.equipment_nodes.count_documents(
            {**base_query, "level": level.value}
        )
        level_counts[level.value] = count
    
    # Count by criticality
    criticality_counts = {
        "safety_critical": await db.equipment_nodes.count_documents(
            {**base_query, "criticality.level": "safety_critical"}
        ),
        "production_critical": await db.equipment_nodes.count_documents(
            {**base_query, "criticality.level": "production_critical"}
        ),
        "medium": await db.equipment_nodes.count_documents(
            {**base_query, "criticality.level": "medium"}
        ),
        "low": await db.equipment_nodes.count_documents(
            {**base_query, "criticality.level": "low"}
        ),
        "unassigned": await db.equipment_nodes.count_documents(
            {**base_query, "criticality": None}
        )
    }
    
    return {
        "total_nodes": total_nodes,
        "by_level": level_counts,
        "by_criticality": criticality_counts
    }

# ============= EQUIPMENT FAILURE MODES (EFM) ENDPOINTS =============

@router.get("/equipment-hierarchy/unstructured")
async def get_unstructured_items(
    current_user: dict = Depends(get_current_user)
):
    """Get all unstructured (unassigned) equipment items."""
    items = await db.unstructured_items.find(
        {"created_by": current_user["id"]},
        {"_id": 0}
    ).to_list(1000)
    return {"items": items}

@router.post("/equipment-hierarchy/unstructured")
async def create_unstructured_item(
    item_data: UnstructuredItemCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create a single unstructured equipment item."""
    # Detect equipment type if not provided
    detected = detect_equipment_type(item_data.name)
    
    item_id = str(uuid.uuid4())
    item_doc = {
        "id": item_id,
        "name": item_data.name,
        "detected_type_id": item_data.detected_type_id or (detected["id"] if detected else None),
        "detected_type_name": detected["name"] if detected else None,
        "detected_discipline": item_data.detected_discipline or (detected["discipline"] if detected else None),
        "detected_icon": detected["icon"] if detected else None,
        "source": item_data.source or "manual",
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.unstructured_items.insert_one(item_doc)
    item_doc.pop("_id", None)
    return item_doc

@router.post("/equipment-hierarchy/parse-list")
async def parse_equipment_list(
    request: ParseEquipmentListRequest,
    current_user: dict = Depends(get_current_user)
):
    """Parse a text list and create unstructured items with auto-detection."""
    import re
    
    content = request.content.strip()
    
    # Split by common delimiters: newlines, commas, semicolons, tabs
    items = re.split(r'[\n\r,;\t]+', content)
    
    # Clean and deduplicate
    seen = set()
    unique_items = []
    for item in items:
        cleaned = item.strip()
        # Remove common list prefixes like "1.", "- ", "• ", etc.
        cleaned = re.sub(r'^[\d]+[.\)]\s*', '', cleaned)
        cleaned = re.sub(r'^[-•*]\s*', '', cleaned)
        cleaned = cleaned.strip()
        
        if cleaned and len(cleaned) > 1 and cleaned.lower() not in seen:
            seen.add(cleaned.lower())
            unique_items.append(cleaned)
    
    # Create unstructured items
    created_items = []
    for name in unique_items:
        detected = detect_equipment_type(name)
        
        item_id = str(uuid.uuid4())
        item_doc = {
            "id": item_id,
            "name": name,
            "detected_type_id": detected["id"] if detected else None,
            "detected_type_name": detected["name"] if detected else None,
            "detected_discipline": detected["discipline"] if detected else None,
            "detected_icon": detected["icon"] if detected else None,
            "source": request.source,
            "created_by": current_user["id"],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.unstructured_items.insert_one(item_doc)
        item_doc.pop("_id", None)
        created_items.append(item_doc)
    
    return {
        "parsed_count": len(created_items),
        "items": created_items
    }

@router.post("/equipment-hierarchy/parse-file")
async def parse_equipment_file(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Parse an uploaded file (Excel, PDF, CSV, TXT) and extract equipment items."""
    import io
    
    filename = file.filename.lower()
    content = await file.read()
    
    extracted_items = []
    
    try:
        if filename.endswith('.csv') or filename.endswith('.txt'):
            # Plain text or CSV
            text_content = content.decode('utf-8', errors='ignore')
            items = text_content.strip().split('\n')
            extracted_items = [item.strip() for item in items if item.strip()]
            
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            # Excel file
            import pandas as pd
            df = pd.read_excel(io.BytesIO(content), header=None)
            # Get all non-empty cells from first column (or all columns)
            for col in df.columns:
                for val in df[col].dropna():
                    if isinstance(val, str) and val.strip():
                        extracted_items.append(val.strip())
                    elif not pd.isna(val):
                        extracted_items.append(str(val).strip())
                        
        elif filename.endswith('.pdf'):
            # PDF file - use PyPDF2 or pdfplumber
            try:
                import pdfplumber
                with pdfplumber.open(io.BytesIO(content)) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text()
                        if text:
                            lines = text.split('\n')
                            extracted_items.extend([l.strip() for l in lines if l.strip()])
            except ImportError:
                # Fallback to PyPDF2
                import PyPDF2
                reader = PyPDF2.PdfReader(io.BytesIO(content))
                for page in reader.pages:
                    text = page.extract_text()
                    if text:
                        lines = text.split('\n')
                        extracted_items.extend([l.strip() for l in lines if l.strip()])
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type. Use .txt, .csv, .xlsx, .xls, or .pdf")
    
    except Exception as e:
        logger.error(f"File parsing error: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")
    
    # Clean and deduplicate
    import re
    seen = set()
    unique_items = []
    for item in extracted_items:
        cleaned = item.strip()
        cleaned = re.sub(r'^[\d]+[.\)]\s*', '', cleaned)
        cleaned = re.sub(r'^[-•*]\s*', '', cleaned)
        cleaned = cleaned.strip()
        
        if cleaned and len(cleaned) > 1 and cleaned.lower() not in seen:
            seen.add(cleaned.lower())
            unique_items.append(cleaned)
    
    # Create unstructured items
    created_items = []
    for name in unique_items[:100]:  # Limit to 100 items per file
        detected = detect_equipment_type(name)
        
        item_id = str(uuid.uuid4())
        item_doc = {
            "id": item_id,
            "name": name,
            "detected_type_id": detected["id"] if detected else None,
            "detected_type_name": detected["name"] if detected else None,
            "detected_discipline": detected["discipline"] if detected else None,
            "detected_icon": detected["icon"] if detected else None,
            "source": "file",
            "created_by": current_user["id"],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.unstructured_items.insert_one(item_doc)
        item_doc.pop("_id", None)
        created_items.append(item_doc)
    
    return {
        "filename": file.filename,
        "parsed_count": len(created_items),
        "items": created_items
    }

@router.post("/equipment-hierarchy/unstructured/{item_id}/assign")
async def assign_unstructured_to_hierarchy(
    item_id: str,
    assignment: AssignToHierarchyRequest,
    current_user: dict = Depends(get_current_user)
):
    """Move an unstructured item into the ISO hierarchy."""
    # Get the unstructured item
    item = await db.unstructured_items.find_one(
        {"id": item_id, "created_by": current_user["id"]}
    )
    if not item:
        raise HTTPException(status_code=404, detail="Unstructured item not found")
    
    # Check for duplicate name under the same parent
    existing = await db.equipment_nodes.find_one({
        "name": item["name"],
        "parent_id": assignment.parent_id,
        "created_by": current_user["id"]
    })
    if existing:
        raise HTTPException(
            status_code=400, 
            detail=f"A node with name '{item['name']}' already exists under this parent"
        )
    
    # Validate parent exists (no created_by filter - equipment is shared)
    parent = await db.equipment_nodes.find_one(
        {"id": assignment.parent_id},
        {"_id": 0}
    )
    if not parent:
        raise HTTPException(status_code=400, detail="Parent node not found")
    
    # Validate ISO level relationship
    try:
        target_level = ISOLevel(assignment.level)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid ISO level: {assignment.level}")
    
    parent_level = ISOLevel(parent["level"])
    if not is_valid_parent_child(parent_level, target_level):
        valid_children = get_valid_child_levels(parent_level)
        raise HTTPException(
            status_code=400,
            detail=f"Cannot add {target_level.value} under {parent_level.value}. Valid: {[c.value for c in valid_children]}"
        )
    
    # Create the equipment node
    node_id = str(uuid.uuid4())
    node_doc = {
        "id": node_id,
        "name": item["name"],
        "level": target_level.value,
        "parent_id": assignment.parent_id,
        "equipment_type_id": item.get("detected_type_id"),
        "description": f"Imported from unstructured list (source: {item.get('source', 'unknown')})",
        "criticality": None,
        "discipline": item.get("detected_discipline"),
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.equipment_nodes.insert_one(node_doc)
    node_doc.pop("_id", None)
    
    # Delete the unstructured item
    await db.unstructured_items.delete_one({"id": item_id})
    
    return node_doc

@router.delete("/equipment-hierarchy/unstructured/{item_id}")
async def delete_unstructured_item(
    item_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete an unstructured item."""
    result = await db.unstructured_items.delete_one(
        {"id": item_id, "created_by": current_user["id"]}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Item deleted"}

@router.delete("/equipment-hierarchy/unstructured")
async def clear_unstructured_items(
    current_user: dict = Depends(get_current_user)
):
    """Delete all unstructured items for the current user."""
    result = await db.unstructured_items.delete_many(
        {"created_by": current_user["id"]}
    )
    return {"message": f"Deleted {result.deleted_count} items"}



# =============================================================================
# SECTION 7: EQUIPMENT HISTORY TIMELINE
# =============================================================================

@router.get("/equipment-hierarchy/nodes/{node_id}/history")
async def get_equipment_history(
    node_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Get the complete history timeline for an equipment node.
    Returns all observations, actions, and task instances related to this equipment.
    """
    # Verify the equipment node exists and belongs to user
    node = await db.equipment_nodes.find_one(
        {"id": node_id, "created_by": current_user["id"]},
        {"_id": 0}
    )
    if not node:
        raise HTTPException(status_code=404, detail="Equipment node not found")
    
    equipment_name = node.get("name", "")
    timeline_items = []
    
    # Get observations linked to this equipment
    observations = await db.threats.find(
        {
            "created_by": current_user["id"],
            "$or": [
                {"linked_equipment_id": node_id},
                {"asset": {"$regex": f"^{equipment_name}$", "$options": "i"}}
            ]
        },
        {"_id": 0}
    ).to_list(100)
    
    for obs in observations:
        timeline_items.append({
            "id": obs.get("id"),
            "type": "observation",
            "title": obs.get("title", "Untitled Observation"),
            "description": obs.get("description", ""),
            "failure_mode": obs.get("failure_mode", ""),
            "status": obs.get("status", "open"),
            "risk_level": obs.get("risk_level", "medium"),
            "risk_score": obs.get("risk_score", 0),
            "created_at": obs.get("created_at"),
            "updated_at": obs.get("updated_at"),
            "source": "threat"
        })
    
    # Get observation IDs for this equipment (to find related actions)
    observation_ids = [obs.get("id") for obs in observations if obs.get("id")]
    
    # Get actions linked to this equipment OR created from observations on this equipment
    action_query = {
        "created_by": current_user["id"],
        "$or": [
            {"linked_equipment_id": node_id},
            {"equipment_name": {"$regex": f"^{equipment_name}$", "$options": "i"}},
            {"source_id": {"$in": observation_ids}} if observation_ids else {"_id": None}
        ]
    }
    # Clean up the query if no observation_ids
    if not observation_ids:
        action_query["$or"] = action_query["$or"][:2]
    
    actions = await db.central_actions.find(
        action_query,
        {"_id": 0}
    ).to_list(100)
    
    for action in actions:
        timeline_items.append({
            "id": action.get("id"),
            "type": "action",
            "title": action.get("title", "Untitled Action"),
            "description": action.get("description", ""),
            "status": action.get("status", "open"),
            "priority": action.get("priority", "medium"),
            "due_date": action.get("due_date"),
            "created_at": action.get("created_at"),
            "updated_at": action.get("updated_at"),
            "source": "action"
        })
    
    # Get task instances linked to this equipment (including completed ones)
    task_instances = await db.task_instances.find(
        {
            "created_by": current_user["id"],
            "$or": [
                {"linked_equipment_id": node_id},
                {"equipment_name": {"$regex": f"^{equipment_name}$", "$options": "i"}},
                {"equipment_id": node_id}
            ]
        },
        {"_id": 0}
    ).to_list(100)
    
    for task in task_instances:
        timeline_items.append({
            "id": task.get("id"),
            "type": "task",
            "title": task.get("name", task.get("task_name", "Untitled Task")),
            "description": task.get("description", ""),
            "status": task.get("status", "pending"),
            "scheduled_date": task.get("scheduled_date"),
            "completed_at": task.get("completed_at"),
            "created_at": task.get("created_at"),
            "updated_at": task.get("updated_at"),
            "source": "task"
        })
    
    # Sort timeline by date (most recent first)
    def get_sort_date(item):
        date_str = item.get("created_at") or item.get("scheduled_date") or ""
        if isinstance(date_str, str):
            try:
                return datetime.fromisoformat(date_str.replace("Z", "+00:00"))
            except:
                return datetime.min.replace(tzinfo=timezone.utc)
        return datetime.min.replace(tzinfo=timezone.utc)
    
    timeline_items.sort(key=get_sort_date, reverse=True)
    
    return {
        "equipment_id": node_id,
        "equipment_name": equipment_name,
        "timeline": timeline_items,
        "total_items": len(timeline_items),
        "counts": {
            "observations": len([i for i in timeline_items if i["type"] == "observation"]),
            "actions": len([i for i in timeline_items if i["type"] == "action"]),
            "tasks": len([i for i in timeline_items if i["type"] == "task"])
        }
    }



# ============================================================================
# HIERARCHY IMPORT
# ============================================================================

class HierarchyImportRequest(BaseModel):
    """Request model for hierarchy import."""
    installation_id: str
    hierarchy: dict  # Nested hierarchy structure
    replace_existing: bool = True  # If True, delete existing equipment first


class ExcelHierarchyImportRequest(BaseModel):
    """Request model for Excel-based hierarchy import."""
    installation_id: str
    excel_url: str  # URL to Excel file
    replace_existing: bool = True


# Level mapping from Excel to ISO 14224
EXCEL_LEVEL_MAPPING = {
    "Plant/Unit": "plant",
    "Section/System": "section",
    "Equipment Unit": "unit",
    "Subunit": "subunit",
    "Maintainable Item": "maintainable_item"
}


def calculate_criticality_from_excel(safety: int, production: int, environmental: int, reputation: int):
    """Calculate criticality data including level, color, and risk score."""
    max_impact = max(safety, production, environmental, reputation)
    
    # Determine level based on max dimension
    if safety >= 4 or max_impact == 5:
        level = "safety_critical"
        color = "#EF4444"  # Red
    elif production >= 4 or max_impact >= 4:
        level = "production_critical"
        color = "#F97316"  # Orange
    elif max_impact >= 3:
        level = "medium"
        color = "#EAB308"  # Yellow
    else:
        level = "low"
        color = "#22C55E"  # Green
    
    # Calculate risk score weighted by dimensions
    risk_score = (
        (safety * 25) +  # Safety has highest weight
        (production * 20) +
        (environmental * 15) +
        (reputation * 10)
    )
    
    return {
        "safety_impact": safety,
        "production_impact": production,
        "environmental_impact": environmental,
        "reputation_impact": reputation,
        "level": level,
        "color": color,
        "max_impact": max_impact,
        "risk_score": round(risk_score, 2)
    }


@router.post("/equipment/import-hierarchy-excel")
async def import_hierarchy_from_excel(
    request: ExcelHierarchyImportRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Import equipment hierarchy from an Excel file URL.
    
    The Excel file should have columns:
    - Name: Equipment name
    - Level: One of "Plant/Unit", "Section/System", "Equipment Unit", "Subunit", "Maintainable Item"
    - Parent: Name of parent equipment
    - Equipment Type: Type of equipment
    - Discipline: Equipment discipline
    - Description: Equipment description
    - Safety: Safety criticality (0-5)
    - Production: Production criticality (0-5)
    - Environmental: Environmental criticality (0-5)
    - Reputation: Reputation criticality (0-5)
    
    Requires admin or owner role.
    """
    import requests as req_lib
    from openpyxl import load_workbook
    
    # Check permissions
    if current_user.get("role") not in ["admin", "owner"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Verify installation exists (equipment_nodes with level="installation")
    installation = await db.equipment_nodes.find_one({
        "id": request.installation_id,
        "level": "installation"
    })
    if not installation:
        raise HTTPException(status_code=404, detail="Installation not found")
    
    installation_id = request.installation_id
    installation_name = installation.get("name")
    
    # Download and parse Excel
    try:
        response = req_lib.get(request.excel_url, timeout=30)
        response.raise_for_status()
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
    except Exception as e:
        logger.error(f"Failed to load Excel file: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to load Excel file: {str(e)}")
    
    headers = [cell.value for cell in ws[1]]
    
    # Level order for hierarchy
    level_order = {"plant": 0, "section": 1, "unit": 2, "subunit": 3, "maintainable_item": 4}
    
    # Parse with FULL PATH tracking to uniquely identify items
    # Items with same name but different parents are DIFFERENT items
    current_path = []  # List of (level, name) tuples
    all_items = []
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = dict(zip(headers, row))
        name = data.get('Name')
        level_raw = data.get('Level')
        
        if not name or not level_raw:
            continue
        
        # Map level
        level = EXCEL_LEVEL_MAPPING.get(level_raw)
        if not level:
            logger.warning(f"Unknown level '{level_raw}' for '{name}', skipping...")
            continue
        
        name = name.strip() if isinstance(name, str) else name
        level_num = level_order[level]
        
        # Get criticality values
        safety = int(data.get('Safety') or 0)
        production = int(data.get('Production') or 0)
        environmental = int(data.get('Environmental') or 0)
        reputation = int(data.get('Reputation') or 0)
        
        # Calculate criticality if any dimension > 0
        criticality = None
        if safety > 0 or production > 0 or environmental > 0 or reputation > 0:
            criticality = calculate_criticality_from_excel(safety, production, environmental, reputation)
        
        # Trim path to correct level (clear deeper or same levels)
        while current_path and level_order[current_path[-1][0]] >= level_num:
            current_path.pop()
        
        # Get parent from path
        parent_name = current_path[-1][1] if current_path else None
        
        # Add to path
        current_path.append((level, name))
        
        # Create full path string for unique identification
        full_path = ' > '.join([p[1] for p in current_path])
        
        all_items.append({
            'name': name,
            'level': level,
            'parent_name': parent_name,
            'full_path': full_path,
            'equipment_type': data.get('Equipment Type'),
            'discipline': data.get('Discipline'),
            'description': data.get('Description'),
            'criticality': criticality
        })
    
    # Deduplicate by FULL PATH (preserves items with same name under different parents)
    unique_items = {}
    for item in all_items:
        key = item['full_path']
        if key not in unique_items:
            unique_items[key] = item
        elif item['criticality'] and not unique_items[key].get('criticality'):
            unique_items[key]['criticality'] = item['criticality']
    
    items_list = list(unique_items.values())
    logger.info(f"Parsed {len(items_list)} unique equipment items from Excel")
    
    # Delete existing equipment if requested
    deleted_count = 0
    if request.replace_existing:
        result = await db.equipment_nodes.delete_many({"installation_id": installation_id})
        deleted_count = result.deleted_count
    
    # Sort by level hierarchy for proper parent resolution
    sorted_items = sorted(items_list, key=lambda x: level_order.get(x['level'], 5))
    
    # Create items with proper parent IDs using full path
    path_to_id = {}  # Map full path to ID
    equipment_list = []
    sort_order = 0
    
    for item in sorted_items:
        sort_order += 1
        eq_id = str(uuid.uuid4())
        
        # Store ID by full path
        path_to_id[item['full_path']] = eq_id
        
        # Find parent ID using full path
        parent_id = installation_id  # Default to installation
        if item.get('parent_name'):
            # Find parent's full path by removing this item from the end
            parent_path_parts = item['full_path'].rsplit(' > ', 1)
            if len(parent_path_parts) > 1:
                parent_path = parent_path_parts[0]
                parent_id = path_to_id.get(parent_path, installation_id)
        
        eq = {
            "id": eq_id,
            "name": item['name'],
            "parent_id": parent_id,
            "installation_id": installation_id,
            "level": item['level'],
            "equipment_type": item.get('equipment_type'),
            "discipline": item.get('discipline'),
            "description": item.get('description'),
            "criticality": item.get('criticality'),
            "sort_order": sort_order,
            "created_by": current_user["id"],
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        equipment_list.append(eq)
    
    # Insert all equipment
    inserted_count = 0
    items_with_criticality = 0
    if equipment_list:
        result = await db.equipment_nodes.insert_many(equipment_list)
        inserted_count = len(result.inserted_ids)
        items_with_criticality = sum(1 for eq in equipment_list if eq.get('criticality'))
    
    # Count by level
    from collections import Counter
    level_counts = dict(Counter([eq['level'] for eq in equipment_list]))
    
    logger.info(f"Excel hierarchy import for installation {installation_id}: deleted={deleted_count}, inserted={inserted_count}, with_criticality={items_with_criticality}")
    
    return {
        "success": True,
        "installation_id": installation_id,
        "installation_name": installation.get("name"),
        "deleted_count": deleted_count,
        "inserted_count": inserted_count,
        "items_with_criticality": items_with_criticality,
        "by_level": level_counts,
        "message": f"Successfully imported {inserted_count} equipment items ({items_with_criticality} with criticality data)"
    }


@router.post("/equipment/import-hierarchy")
async def import_equipment_hierarchy(
    request: HierarchyImportRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Import a complete equipment hierarchy for an installation.
    
    This endpoint allows importing a nested hierarchy structure like:
    {
        "Line-90": {
            "type": "section",
            "children": {
                "Feedstock Prep Unit": {
                    "type": "unit",
                    "children": {
                        "Crane Subunit": {
                            "type": "subunit",
                            "children": ["Crane Motor", "Crane Chain"]
                        }
                    }
                }
            }
        }
    }
    
    Requires admin or owner role.
    """
    # Check permissions
    if current_user.get("role") not in ["admin", "owner"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Verify installation exists and user has access
    installation = await db.installations.find_one({"id": request.installation_id})
    if not installation:
        raise HTTPException(status_code=404, detail="Installation not found")
    
    installation_id = request.installation_id
    
    # Delete existing equipment if requested
    deleted_count = 0
    if request.replace_existing:
        result = await db.equipment.delete_many({"installation_id": installation_id})
        deleted_count = result.deleted_count
    
    # Process hierarchy and create equipment items
    equipment_list = []
    sort_order = 0
    
    def create_equipment(name, parent_id, eq_type, level):
        nonlocal sort_order
        sort_order += 1
        return {
            "id": str(uuid.uuid4()),
            "name": name,
            "parent_id": parent_id,
            "installation_id": installation_id,
            "type": eq_type,
            "level": level,
            "sort_order": sort_order,
            "created_by": current_user["id"],
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
    
    def process_hierarchy(data, parent_id=None, level=0):
        items = []
        for name, info in data.items():
            if isinstance(info, dict):
                eq_type = info.get("type", "equipment")
                eq = create_equipment(name, parent_id, eq_type, level)
                items.append(eq)
                
                children = info.get("children", {})
                if isinstance(children, dict):
                    items.extend(process_hierarchy(children, eq["id"], level + 1))
                elif isinstance(children, list):
                    for child_name in children:
                        child_eq = create_equipment(child_name, eq["id"], "maintainable_item", level + 1)
                        items.append(child_eq)
        return items
    
    equipment_list = process_hierarchy(request.hierarchy)
    
    # Insert all equipment
    inserted_count = 0
    if equipment_list:
        result = await db.equipment.insert_many(equipment_list)
        inserted_count = len(result.inserted_ids)
    
    logger.info(f"Hierarchy import for installation {installation_id}: deleted={deleted_count}, inserted={inserted_count}")
    
    return {
        "success": True,
        "installation_id": installation_id,
        "deleted_count": deleted_count,
        "inserted_count": inserted_count,
        "message": f"Successfully imported {inserted_count} equipment items"
    }
