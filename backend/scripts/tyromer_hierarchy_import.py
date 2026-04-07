"""
Tyromer Equipment Hierarchy Import Script

This script imports the Tyromer equipment hierarchy from the Excel file.
Run this script on the production database to update the hierarchy.

Usage:
    MONGO_URL="mongodb+srv://..." DB_NAME="assetiq" python3 tyromer_hierarchy_import.py

Or use the API endpoint:
    POST /api/equipment/import-hierarchy-excel
"""
import asyncio
from motor.motor_asyncio import AsyncIOMotorClient
import os
import uuid
from datetime import datetime, timezone
import requests
import io

try:
    from openpyxl import load_workbook
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.run(["pip", "install", "openpyxl", "-q"])
    from openpyxl import load_workbook

MONGO_URL = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.environ.get("DB_NAME", "assetiq")

# Excel file URL
EXCEL_URL = "https://customer-assets.emergentagent.com/job_d7857aee-556b-43c4-be83-cb625b13d6c2/artifacts/5nkh14hz_equipment_hierarchy_20260407_090659.xlsx"

# Level mapping from Excel to ISO 14224
LEVEL_MAPPING = {
    "Plant/Unit": "plant",
    "Section/System": "section",
    "Equipment Unit": "unit",
    "Subunit": "subunit",
    "Maintainable Item": "maintainable_item"
}

# Level order for hierarchy (lower = higher level)
LEVEL_ORDER = {
    "plant": 0,
    "section": 1,
    "unit": 2,
    "subunit": 3,
    "maintainable_item": 4
}


def calculate_criticality_data(safety, production, environmental, reputation):
    """Calculate criticality data including level, color, and risk score."""
    max_impact = max(safety, production, environmental, reputation)
    
    # Determine level based on max dimension
    if safety >= 4 or max_impact == 5:
        level = "safety_critical"
        color = "#EF4444"  # Red
    elif production >= 4 or max_impact >= 4:
        level = "production_critical"
        color = "#F97316"  # Orange
    elif max_impact >= 3:
        level = "medium"
        color = "#EAB308"  # Yellow
    else:
        level = "low"
        color = "#22C55E"  # Green
    
    # Calculate risk score weighted by dimensions
    risk_score = (
        (safety * 25) +  # Safety has highest weight
        (production * 20) +
        (environmental * 15) +
        (reputation * 10)
    )
    
    return {
        "safety_impact": safety,
        "production_impact": production,
        "environmental_impact": environmental,
        "reputation_impact": reputation,
        "level": level,
        "color": color,
        "max_impact": max_impact,
        "risk_score": round(risk_score, 2)
    }


def parse_excel_hierarchy(excel_url: str = EXCEL_URL):
    """Parse the Excel file and return a structured hierarchy using full-path tracking."""
    print(f"Downloading Excel file from {excel_url}")
    response = requests.get(excel_url)
    response.raise_for_status()
    
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    
    # Parse with FULL PATH tracking to uniquely identify items
    # Items with same name but different parents are DIFFERENT items
    current_path = []  # List of (level, name) tuples
    all_items = []
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = dict(zip(headers, row))
        name = data.get('Name')
        level_raw = data.get('Level')
        
        if not name or not level_raw:
            continue
        
        level = LEVEL_MAPPING.get(level_raw)
        if not level:
            print(f"Warning: Unknown level '{level_raw}' for '{name}', skipping...")
            continue
        
        name = name.strip() if isinstance(name, str) else name
        level_num = LEVEL_ORDER[level]
        
        # Get criticality values
        safety = int(data.get('Safety') or 0)
        production = int(data.get('Production') or 0)
        environmental = int(data.get('Environmental') or 0)
        reputation = int(data.get('Reputation') or 0)
        
        # Calculate criticality if any dimension > 0
        criticality = None
        if safety > 0 or production > 0 or environmental > 0 or reputation > 0:
            criticality = calculate_criticality_data(safety, production, environmental, reputation)
        
        # Trim path to correct level (clear deeper or same levels)
        while current_path and LEVEL_ORDER[current_path[-1][0]] >= level_num:
            current_path.pop()
        
        # Get parent from path
        parent_name = current_path[-1][1] if current_path else None
        
        # Add to path
        current_path.append((level, name))
        
        # Create full path string for unique identification
        full_path = ' > '.join([p[1] for p in current_path])
        
        all_items.append({
            'name': name,
            'level': level,
            'parent_name': parent_name,
            'full_path': full_path,
            'equipment_type': data.get('Equipment Type'),
            'discipline': data.get('Discipline'),
            'description': data.get('Description'),
            'criticality': criticality
        })
    
    # Deduplicate by FULL PATH (preserves items with same name under different parents)
    unique_items = {}
    for item in all_items:
        key = item['full_path']
        if key not in unique_items:
            unique_items[key] = item
        elif item['criticality'] and not unique_items[key].get('criticality'):
            unique_items[key]['criticality'] = item['criticality']
    
    items_list = list(unique_items.values())
    print(f"Parsed {len(items_list)} unique equipment items")
    return items_list


async def import_hierarchy(installation_name: str = "Tyromer", dry_run: bool = False, excel_url: str = EXCEL_URL):
    """
    Import the hierarchy from Excel into the specified installation.
    
    Args:
        installation_name: Name of the installation to update (case-insensitive)
        dry_run: If True, only show what would be done without making changes
        excel_url: URL to the Excel file
    """
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]
    
    # Find installation (equipment_nodes with level="installation")
    installation = await db.equipment_nodes.find_one(
        {
            "level": "installation",
            "name": {"$regex": f"^{installation_name}$", "$options": "i"}
        }
    )
    
    if not installation:
        # List available installations
        installations = await db.equipment_nodes.find(
            {"level": "installation"},
            {"_id": 0, "id": 1, "name": 1}
        ).to_list(100)
        print(f"Installation '{installation_name}' not found!")
        print("\nAvailable installations:")
        for inst in installations:
            print(f"  - {inst['name']}")
        client.close()
        return {"error": f"Installation '{installation_name}' not found"}
    
    installation_id = installation["id"]
    print(f"Found installation: {installation['name']} (id: {installation_id})")
    
    if dry_run:
        print("\n[DRY RUN] Would perform the following actions:")
    
    # Parse Excel
    items_list = parse_excel_hierarchy(excel_url)
    
    # Count existing equipment
    existing_count = await db.equipment_nodes.count_documents({"installation_id": installation_id})
    print(f"Existing equipment items: {existing_count}")
    
    if not dry_run:
        # Delete existing equipment for this installation
        result = await db.equipment_nodes.delete_many({"installation_id": installation_id})
        print(f"Deleted {result.deleted_count} existing equipment items")
    else:
        print(f"[DRY RUN] Would delete {existing_count} existing equipment items")
    
    # Sort by level for proper parent ID resolution
    sorted_items = sorted(items_list, key=lambda x: LEVEL_ORDER.get(x['level'], 5))
    
    # Create items with proper parent IDs
    path_to_id = {}  # Map full path to ID
    equipment_list = []
    sort_order = 0
    user_id = 'system-import'
    
    for item in sorted_items:
        sort_order += 1
        eq_id = str(uuid.uuid4())
        
        # Store ID by full path
        path_to_id[item['full_path']] = eq_id
        
        # Find parent ID using full path
        parent_id = installation_id  # Default to installation
        if item.get('parent_name'):
            # Find parent's full path by removing this item from the end
            parent_path_parts = item['full_path'].rsplit(' > ', 1)
            if len(parent_path_parts) > 1:
                parent_path = parent_path_parts[0]
                parent_id = path_to_id.get(parent_path, installation_id)
        
        eq = {
            "id": eq_id,
            "name": item['name'],
            "parent_id": parent_id,
            "installation_id": installation_id,
            "level": item['level'],
            "equipment_type": item.get('equipment_type'),
            "discipline": item.get('discipline'),
            "description": item.get('description'),
            "criticality": item.get('criticality'),
            "sort_order": sort_order,
            "created_by": user_id,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        equipment_list.append(eq)
    
    if not dry_run:
        # Insert all equipment
        if equipment_list:
            result = await db.equipment_nodes.insert_many(equipment_list)
            print(f"Inserted {len(result.inserted_ids)} equipment items")
    else:
        print(f"[DRY RUN] Would insert {len(equipment_list)} equipment items")
    
    # Print hierarchy summary
    print("\n=== Hierarchy Summary ===")
    
    # Count by level
    from collections import Counter
    level_counts = Counter([eq['level'] for eq in equipment_list])
    for level, count in sorted(level_counts.items()):
        print(f"  {level}: {count}")
    
    # Count with criticality
    with_crit = sum(1 for eq in equipment_list if eq.get('criticality'))
    print(f"\nEquipment with criticality data: {with_crit}")
    
    # Show sample items with criticality
    if with_crit > 0:
        print("\nSample items with criticality:")
        crit_items = [eq for eq in equipment_list if eq.get('criticality')][:5]
        for eq in crit_items:
            crit = eq['criticality']
            print(f"  - {eq['name']}: Safety={crit['safety_impact']}, Prod={crit['production_impact']}, Level={crit['level']}")
    
    client.close()
    print("\nDone!")
    
    return {
        "success": True,
        "installation": installation['name'],
        "total_items": len(equipment_list),
        "items_with_criticality": with_crit,
        "by_level": dict(level_counts)
    }


if __name__ == "__main__":
    import sys
    
    # Parse arguments
    installation_name = "Tyromer"
    dry_run = "--dry-run" in sys.argv
    
    for arg in sys.argv[1:]:
        if not arg.startswith("--"):
            installation_name = arg
    
    print(f"Importing hierarchy for: {installation_name}")
    if dry_run:
        print("Running in DRY RUN mode (no changes will be made)")
    
    asyncio.run(import_hierarchy(installation_name, dry_run))
