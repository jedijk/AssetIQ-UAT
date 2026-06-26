"""Production log parse template CRUD and batch ingest."""
import io
import logging
import uuid
from datetime import datetime, timezone
from typing import Dict, List, Optional

from fastapi import BackgroundTasks, Form, HTTPException
from pydantic import BaseModel

from database import db
from services.background_jobs import schedule_tracked_job
from services.production_logs_parsing import (
    ColumnMapping,
    ParseTemplate,
    _fuzzy_match_columns,
    _normalize_column_name,
)
from services.tenant_scope import scoped

logger = logging.getLogger(__name__)


class ColumnAlias(BaseModel):
    """Defines aliases for flexible column matching."""
    canonical_name: str  # The standard name used internally
    aliases: List[str] = []  # Alternative names that map to this column


class SavedTemplate(BaseModel):
    name: str
    description: Optional[str] = None
    delimiter: str = ","
    has_header: bool = True
    skip_rows: int = 0
    timestamp_format: Optional[str] = None
    column_mapping: ColumnMapping = ColumnMapping()
    # Column aliases for fuzzy matching
    column_aliases: Dict[str, List[str]] = {}  # {canonical_name: [alias1, alias2, ...]}


class SaveTemplateRequest(BaseModel):
    name: str
    description: Optional[str] = None
    template: dict
    column_aliases: Dict[str, List[str]] = {}


class BatchIngestWithTemplateRequest(BaseModel):
    job_ids: List[str]
    template_id: str


# ======================== Template Management ========================

async def save_template(user: dict,
    request: SaveTemplateRequest,
):
    """Save a parse template for reuse."""
    # Check for duplicate name
    existing = await db.log_parse_templates.find_one(scoped(user, {"name": request.name}))
    if existing:
        raise HTTPException(status_code=400, detail="Template with this name already exists")
    
    template_doc = {
        "id": str(uuid.uuid4()),
        "name": request.name,
        "description": request.description,
        "template": request.template,
        "column_aliases": request.column_aliases,
        "created_by": user.get("id"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "usage_count": 0,
    }
    
    await db.log_parse_templates.insert_one(template_doc)
    del template_doc["_id"]
    
    logger.info(f"[Templates] Created template '{request.name}' by user {user.get('id')}")
    return template_doc


async def list_templates(user: dict,
):
    """List all saved parse templates."""
    templates = await db.log_parse_templates.find(
        scoped(user, {}), {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    return {"templates": templates}


async def get_template(user: dict,
    template_id: str,
):
    """Get a specific template."""
    template = await db.log_parse_templates.find_one(scoped(user, {"id": template_id}), {"_id": 0})
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    return template


async def update_template(user: dict,
    template_id: str,
    request: SaveTemplateRequest,
):
    """Update an existing template."""
    template = await db.log_parse_templates.find_one(scoped(user, {"id": template_id}))
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    # Check for duplicate name (if changed)
    if request.name != template.get("name"):
        existing = await db.log_parse_templates.find_one(scoped(user, {"name": request.name, "id": {"$ne": template_id}}))
        if existing:
            raise HTTPException(status_code=400, detail="Template with this name already exists")
    
    await db.log_parse_templates.update_one(
        {"id": template_id},
        {"$set": {
            "name": request.name,
            "description": request.description,
            "template": request.template,
            "column_aliases": request.column_aliases,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }}
    )
    
    logger.info(f"[Templates] Updated template '{request.name}'")
    return {"message": "Template updated"}


async def delete_template(user: dict,
    template_id: str,
):
    """Delete a template."""
    result = await db.log_parse_templates.delete_one({"id": template_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    
    logger.info(f"[Templates] Deleted template {template_id}")
    return {"message": "Template deleted"}


async def batch_ingest_with_saved_template(user: dict,
    request: BatchIngestWithTemplateRequest,
    background_tasks: BackgroundTasks,
):
    """
    Batch ingest multiple jobs using a saved template with fuzzy column matching.
    This is the "train once, apply to many" workflow.
    """
    if not request.job_ids:
        raise HTTPException(status_code=400, detail="No jobs selected")
    
    # Get the saved template
    saved_template = await db.log_parse_templates.find_one(scoped(user, {"id": request.template_id}), {"_id": 0})
    if not saved_template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    try:
        base_template = ParseTemplate(**saved_template["template"])
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid template configuration: {e}")
    
    column_aliases = saved_template.get("column_aliases", {})
    started = []
    match_reports = []
    
    for job_id in request.job_ids:
        job = await db.log_ingestion_jobs.find_one(scoped(user, {"id": job_id}), {"_id": 0})
        if not job:
            continue
        if job.get("status") == "completed":
            continue
        
        # Get file columns to perform fuzzy matching
        from services.storage_service import get_object_async
        file_columns = []
        try:
            first_file = job["files"][0]
            data, _ = await get_object_async(first_file["storage_path"])
            file_ext = first_file.get("extension", "").lower()
            
            if file_ext in ("xlsx", "xls"):
                import openpyxl
                if file_ext == "xls":
                    import xlrd
                    wb = xlrd.open_workbook(file_contents=data)
                    ws = wb.sheet_by_index(0)
                    if ws.nrows > 0:
                        file_columns = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
                else:
                    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
                    ws = wb.active
                    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                    if first_row:
                        file_columns = [str(c).strip() if c else "" for c in first_row]
                    wb.close()
            else:
                content = data.decode("utf-8", errors="replace")
                lines = content.splitlines()
                if lines and base_template.has_header:
                    delimiter = base_template.delimiter
                    if delimiter == "\\t":
                        delimiter = "\t"
                    file_columns = [c.strip() for c in lines[0].split(delimiter)]
        except Exception as e:
            logger.warning(f"[BatchIngest] Could not read columns for job {job_id}: {e}")
            # Use original template columns if we can't read the file
            file_columns = []
        
        # Perform fuzzy column matching
        matched_columns = _fuzzy_match_columns(
            file_columns, 
            base_template.column_mapping,
            column_aliases
        )
        
        # Create adjusted template with matched columns
        adjusted_template = {
            "delimiter": base_template.delimiter,
            "has_header": base_template.has_header,
            "skip_rows": base_template.skip_rows,
            "timestamp_format": base_template.timestamp_format,
            "base_date_location": base_template.base_date_location,
            "header_metadata": base_template.header_metadata,
            "secondary_sheet": base_template.secondary_sheet,
            "column_mapping": {
                "timestamp": matched_columns.get("timestamp"),
                "asset_id": base_template.column_mapping.asset_id,  # Keep original (may be static)
                "status": matched_columns.get("status"),
                "event_type": matched_columns.get("event_type"),
                "metric_columns": matched_columns.get("metric_columns", []),
            }
        }
        
        # Log the matching for debugging
        match_reports.append({
            "job_id": job_id,
            "filename": job["files"][0]["filename"] if job["files"] else "unknown",
            "file_columns": file_columns[:10],  # First 10 columns
            "matched": matched_columns,
        })
        
        # Set template and mark as processing
        await db.log_ingestion_jobs.update_one(
            {"id": job_id},
            {"$set": {
                "parse_template": adjusted_template,
                "template_used": saved_template["name"],
                "status": "processing",
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }}
        )
        job["parse_template"] = adjusted_template
        from services.production_logs_service import _run_ingestion

        schedule_tracked_job(
            background_tasks,
            "production_logs_template_ingest",
            _run_ingestion,
            job_id,
            job,
            user_id=user.get("id"),
        )
        started.append(job_id)
    
    # Update template usage count
    await db.log_parse_templates.update_one(
        {"id": request.template_id},
        {"$inc": {"usage_count": len(started)}}
    )
    
    logger.info(f"[BatchIngest] Started {len(started)} jobs with template '{saved_template['name']}'")
    
    return {
        "started": len(started),
        "job_ids": started,
        "template_name": saved_template["name"],
        "match_reports": match_reports,
        "message": f"Batch ingestion started for {len(started)} job(s) using template '{saved_template['name']}'",
    }


async def preview_template_match(user: dict,
    job_id: str = Form(...),
    template_id: str = Form(...),
):
    """
    Preview how a saved template would match columns in a specific job.
    Useful for verifying the fuzzy matching before batch processing.
    """
    # Get job
    job = await db.log_ingestion_jobs.find_one(scoped(user, {"id": job_id}), {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    # Get template
    saved_template = await db.log_parse_templates.find_one(scoped(user, {"id": template_id}), {"_id": 0})
    if not saved_template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    try:
        base_template = ParseTemplate(**saved_template["template"])
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid template: {e}")
    
    column_aliases = saved_template.get("column_aliases", {})
    
    # Get file columns
    from services.storage_service import get_object_async
    file_columns = []
    sample_rows = []
    
    try:
        first_file = job["files"][0]
        data, _ = await get_object_async(first_file["storage_path"])
        file_ext = first_file.get("extension", "").lower()
        
        if file_ext in ("xlsx", "xls"):
            import openpyxl
            if file_ext == "xls":
                import xlrd
                wb = xlrd.open_workbook(file_contents=data)
                ws = wb.sheet_by_index(0)
                if ws.nrows > 0:
                    file_columns = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
                for r in range(1, min(ws.nrows, 6)):
                    row_data = {file_columns[c]: str(ws.cell_value(r, c)) for c in range(min(len(file_columns), ws.ncols))}
                    sample_rows.append(row_data)
            else:
                wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
                ws = wb.active
                rows_iter = ws.iter_rows(values_only=True)
                first_row = next(rows_iter, None)
                if first_row:
                    file_columns = [str(c).strip() if c else "" for c in first_row]
                for i, row in enumerate(rows_iter):
                    if i >= 5:
                        break
                    row_data = {file_columns[j]: str(row[j]) if row[j] is not None else "" for j in range(min(len(file_columns), len(row)))}
                    sample_rows.append(row_data)
                wb.close()
        else:
            content = data.decode("utf-8", errors="replace")
            lines = content.splitlines()
            delimiter = base_template.delimiter
            if delimiter == "\\t":
                delimiter = "\t"
            if lines and base_template.has_header:
                file_columns = [c.strip() for c in lines[0].split(delimiter)]
            for line in lines[1:6]:
                vals = line.split(delimiter)
                row_data = {file_columns[j]: vals[j].strip() if j < len(vals) else "" for j in range(len(file_columns))}
                sample_rows.append(row_data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read file: {e}")
    
    # Perform fuzzy matching
    matched_columns = _fuzzy_match_columns(
        file_columns,
        base_template.column_mapping,
        column_aliases
    )
    
    # Build match details showing what was matched
    match_details = []
    template_mapping = base_template.column_mapping
    
    if template_mapping.timestamp:
        match_details.append({
            "field": "timestamp",
            "template_column": template_mapping.timestamp,
            "matched_to": matched_columns.get("timestamp"),
            "match_type": "exact" if matched_columns.get("timestamp") == template_mapping.timestamp else "fuzzy",
            "success": matched_columns.get("timestamp") in file_columns,
        })
    
    if template_mapping.asset_id:
        match_details.append({
            "field": "asset_id",
            "template_column": template_mapping.asset_id,
            "matched_to": matched_columns.get("asset_id"),
            "match_type": "exact" if matched_columns.get("asset_id") == template_mapping.asset_id else "fuzzy",
            "success": matched_columns.get("asset_id") in file_columns,
        })
    
    if template_mapping.status:
        match_details.append({
            "field": "status",
            "template_column": template_mapping.status,
            "matched_to": matched_columns.get("status"),
            "match_type": "exact" if matched_columns.get("status") == template_mapping.status else "fuzzy",
            "success": matched_columns.get("status") in file_columns,
        })
    
    for metric in template_mapping.metric_columns:
        matched_metric = next((m for m in matched_columns.get("metric_columns", []) if _normalize_column_name(m) == _normalize_column_name(metric) or m == metric), None)
        match_details.append({
            "field": "metric",
            "template_column": metric,
            "matched_to": matched_metric,
            "match_type": "exact" if matched_metric == metric else "fuzzy",
            "success": matched_metric in file_columns if matched_metric else False,
        })
    
    return {
        "template_name": saved_template["name"],
        "file_columns": file_columns,
        "matched_columns": matched_columns,
        "match_details": match_details,
        "sample_rows": sample_rows,
        "all_matched": all(d["success"] for d in match_details if d["template_column"]),
    }


