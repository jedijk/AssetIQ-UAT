"""Failure modes library routes — read service module."""
from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
from datetime import datetime
import logging
import io
from database import db, failure_modes_service
from services.ai_gateway import user_context
from failure_modes import FAILURE_MODES_LIBRARY
from services.failure_modes.static_library_fallback import (
    filter_static_failure_modes,
    get_all_categories,
    get_all_equipment_types,
    list_uses_static_library_fallback,
)

logger = logging.getLogger(__name__)


class FindSimilarFailureModesScanRequest(BaseModel):
    equipment_type_id: Optional[str] = None
    jaccard_threshold: float = 0.45
    ratio_threshold: float = 0.72
    min_score: float = 52.0
    limit_groups: int = 200
    include_cross_equipment: bool = True
    only_cross_equipment: bool = False
    cross_equipment_ratio_threshold: float = 0.88
    use_ai: bool = True
    ai_max_clusters: int = 30
    ai_time_budget_seconds: float = 55.0


async def get_failure_modes(
    category: Optional[str] = None,
    equipment: Optional[str] = None,
    search: Optional[str] = None,
    min_rpn: Optional[int] = None,
    equipment_type_id: Optional[str] = None,
    mechanism: Optional[str] = None,
    is_validated: Optional[bool] = None,
    failure_mode_type: Optional[str] = None,
    recently_added_days: Optional[int] = 30,
    skip: int = 0,
    limit: int = 500,
    *,
    current_user: dict,
):
    """Get failure modes from MongoDB with optional filters."""
    try:
        result = await failure_modes_service.get_all(
            category=category,
            equipment=equipment,
            search=search,
            min_rpn=min_rpn,
            equipment_type_id=equipment_type_id,
            mechanism=mechanism,
            is_validated=is_validated,
            failure_mode_type=failure_mode_type,
            recently_added_days=recently_added_days if failure_mode_type == "recently_added" else None,
            skip=skip,
            limit=limit,
            user=current_user,
        )

        if result["total"] == 0 or (
            result["total"] < 100
            and list_uses_static_library_fallback(
                category=category,
                equipment=equipment,
                search=search,
                min_rpn=min_rpn,
                equipment_type_id=equipment_type_id,
                mechanism=mechanism,
                is_validated=is_validated,
                failure_mode_type=failure_mode_type,
            )
        ):
            logger.info("MongoDB failure_modes sparse for filters, using static library fallback")
            results = filter_static_failure_modes(
                category=category,
                equipment=equipment,
                search=search,
                min_rpn=min_rpn,
            )
            return {"total": len(results), "failure_modes": results, "source": "static"}

        return result
    except Exception as e:
        logger.error(f"Error fetching from MongoDB, falling back to static: {e}")
        results = filter_static_failure_modes(
            category=category,
            equipment=equipment,
            search=search,
            min_rpn=min_rpn,
        )
        return {"total": len(results), "failure_modes": results, "source": "static"}


async def get_categories(current_user: dict):
    """Get all unique categories."""
    static_categories = get_all_categories()
    try:
        categories = await failure_modes_service.get_categories()
        if categories and len(categories) >= 8:
            return {"categories": categories}
    except Exception as e:
        logger.error(f"Error fetching categories from MongoDB: {e}")
    return {"categories": static_categories}


async def get_equipment_types(current_user: dict):
    """Get all unique equipment types."""
    try:
        types = await failure_modes_service.get_equipment_types()
        if types:
            return {"equipment_types": types}
    except Exception as e:
        logger.error(f"Error fetching equipment types from MongoDB: {e}")
    return {"equipment_types": get_all_equipment_types()}


async def get_mechanisms(current_user: dict):
    """Get all unique ISO 14224 mechanisms."""
    try:
        mechanisms = await failure_modes_service.get_mechanisms()
        return {"mechanisms": mechanisms}
    except Exception as e:
        logger.error(f"Error fetching mechanisms: {e}")
        return {"mechanisms": []}


async def get_failure_mode_counts_by_equipment_type(*, current_user: dict):
    """
    Get failure mode counts grouped by equipment_type_id.
    Returns a dict mapping equipment_type_id to count of failure modes.
    """
    try:
        pipeline = [
            {"$unwind": "$equipment_type_ids"},
            {"$group": {"_id": "$equipment_type_ids", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
        ]

        results = await db.failure_modes.aggregate(pipeline).to_list(500)
        counts = {item["_id"]: item["count"] for item in results if item["_id"]}
        total = await db.failure_modes.count_documents({})

        return {"counts_by_type": counts, "total_failure_modes": total}
    except Exception as e:
        logger.error(f"Error fetching failure mode counts: {e}")
        return {"counts_by_type": {}, "total_failure_modes": 0}


async def export_failure_modes_excel(*, current_user: dict):
    """Export all failure modes to an Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    try:
        result = await failure_modes_service.get_all(skip=0, limit=10000)
        failure_modes = result.get("failure_modes", []) if isinstance(result, dict) else []
        if not failure_modes:
            failure_modes = FAILURE_MODES_LIBRARY.copy()
    except Exception as e:
        logger.error(f"Error fetching failure modes for export: {e}")
        failure_modes = FAILURE_MODES_LIBRARY.copy()

    wb = Workbook()
    ws = wb.active
    ws.title = "Failure Modes"

    headers = [
        "ID",
        "Category",
        "Equipment",
        "Failure Mode",
        "Process",
        "Potential Effects",
        "Potential Causes",
        "ISO 14224 Mechanism",
        "Severity",
        "Occurrence",
        "Detectability",
        "RPN",
        "Keywords",
        "Recommended Actions",
        "Validated",
        "Source",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, fm in enumerate(failure_modes, 2):
        keywords = fm.get("keywords", [])
        keywords_str = ", ".join(keywords) if isinstance(keywords, list) else str(keywords or "")

        actions = fm.get("recommended_actions", [])
        if isinstance(actions, list):
            if actions and isinstance(actions[0], dict):
                actions_str = "\n".join(
                    [
                        (
                            f"• {a.get('action') or a.get('description') or ''}".strip()
                            + (f" ({a.get('action_type', '')})" if a.get("action_type") else "")
                            + (
                                f" — {a.get('estimated_minutes')} min"
                                if a.get("estimated_minutes") not in (None, "", 0)
                                else ""
                            )
                        )
                        for a in actions
                    ]
                )
            else:
                actions_str = "\n".join([f"• {a}" for a in actions])
        else:
            actions_str = str(actions or "")

        potential_effects = fm.get("potential_effects", "")
        if isinstance(potential_effects, list):
            potential_effects_str = ", ".join(potential_effects)
        else:
            potential_effects_str = str(potential_effects or "")

        potential_causes = fm.get("potential_causes", "")
        if isinstance(potential_causes, list):
            potential_causes_str = ", ".join(potential_causes)
        else:
            potential_causes_str = str(potential_causes or "")

        row_data = [
            str(fm.get("id", "")),
            fm.get("category", ""),
            fm.get("equipment", ""),
            fm.get("failure_mode", ""),
            fm.get("process", ""),
            potential_effects_str,
            potential_causes_str,
            fm.get("iso14224_mechanism", ""),
            fm.get("severity", 0),
            fm.get("occurrence", 0),
            fm.get("detectability", 0),
            fm.get("rpn", 0),
            keywords_str,
            actions_str,
            "Yes" if fm.get("is_validated") else "No",
            fm.get("source", "library"),
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    column_widths = [10, 15, 18, 30, 20, 30, 30, 20, 10, 10, 12, 8, 25, 40, 10, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"failure_modes_{timestamp}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


async def get_high_risk_modes(threshold: int = 150, *, current_user: dict):
    """Get failure modes with RPN above threshold."""
    try:
        high_risk = await failure_modes_service.get_high_risk(threshold)
        return {"threshold": threshold, "total": len(high_risk), "failure_modes": high_risk}
    except Exception as e:
        logger.error(f"Error fetching high-risk modes: {e}")
        high_risk = [fm for fm in FAILURE_MODES_LIBRARY if fm["rpn"] >= threshold]
        high_risk.sort(key=lambda x: -x["rpn"])
        return {"threshold": threshold, "total": len(high_risk), "failure_modes": high_risk}


async def get_failure_mode_by_id(mode_id: str, *, current_user: dict):
    """Get a specific failure mode by ID (MongoDB _id or legacy_id)."""
    try:
        fm = await failure_modes_service.get_by_id(mode_id)
        if fm:
            return fm
    except Exception as e:
        logger.error(f"Error fetching failure mode {mode_id}: {e}")

    try:
        legacy_id = int(mode_id)
        for fm in FAILURE_MODES_LIBRARY:
            if fm["id"] == legacy_id:
                result = dict(fm)
                result["legacy_id"] = legacy_id
                result.setdefault("version", 1)
                result["id"] = str(legacy_id)
                return result
    except ValueError:
        pass

    raise HTTPException(status_code=404, detail="Failure mode not found")


async def get_similar_failure_modes(
    mode_id: str,
    threshold: float = 55.0,
    limit: int = 20,
    require_shared_equipment_type: bool = False,
    *,
    current_user: dict,
):
    """Lexical similarity search for near-duplicate failure modes (no AI)."""
    try:
        return await failure_modes_service.find_similar(
            mode_id,
            threshold=threshold,
            limit=limit,
            require_shared_equipment_type=require_shared_equipment_type,
        )
    except Exception as e:
        logger.error(f"Error finding similar failure modes for {mode_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))


async def scan_similar_failure_modes(request: FindSimilarFailureModesScanRequest, *, current_user: dict):
    """Batch scan for near-duplicate failure modes across the full library (lexical, no AI)."""
    uid, cid = user_context(current_user)
    try:
        return await failure_modes_service.scan_similar_groups(
            equipment_type_id=request.equipment_type_id,
            jaccard_threshold=request.jaccard_threshold,
            ratio_threshold=request.ratio_threshold,
            min_score=request.min_score,
            limit_groups=request.limit_groups,
            include_cross_equipment=request.include_cross_equipment,
            only_cross_equipment=request.only_cross_equipment,
            cross_equipment_ratio_threshold=request.cross_equipment_ratio_threshold,
            use_ai=request.use_ai,
            ai_max_clusters=request.ai_max_clusters,
            ai_time_budget_seconds=request.ai_time_budget_seconds,
            user_id=uid,
            company_id=cid,
        )
    except Exception as e:
        logger.error(f"Error scanning similar failure modes: {e}")
        raise HTTPException(status_code=500, detail=str(e))
