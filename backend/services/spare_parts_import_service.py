"""Spare parts Excel import — validate and bulk import."""
from __future__ import annotations

import io
import re
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlparse

from database import db
from models.spare_parts import SparePartCreate, SparePartEquipmentLink
from services.spare_categories_service import list_categories
from services.spare_parts_service import _find_by_duplicate_key, create_spare_part
from services.tenant_schema import merge_tenant_filter

HEADER_ALIASES = {
    "equipment": {"equipment", "equipment tag", "equipment name", "tag"},
    "description": {"spare part description", "description", "spare description"},
    "type_model": {"type / model", "type/model", "type_model", "type model", "model"},
    "manufacturer": {"manufacturer", "mfr", "vendor"},
    "category": {"category"},
    "component_position": {"component position", "position", "location"},
    "notes": {"notes", "note"},
    "document_url": {"document url", "document_url", "url", "document link"},
}


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _map_headers(header_row: List[Any]) -> Dict[str, int]:
    mapped: Dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        norm = _normalize_header(cell)
        if not norm:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if norm in aliases and field not in mapped:
                mapped[field] = idx
    return mapped


def parse_spare_parts_workbook(content: bytes) -> Tuple[List[dict], List[str]]:
    """Parse uploaded Excel into row dicts. Returns (rows, parse_errors)."""
    import openpyxl

    errors: List[str] = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        return [], [f"Could not read Excel file: {exc}"]

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], ["Worksheet is empty"]

    col_map = _map_headers(list(header_row))
    missing = [f for f in ("equipment", "description", "type_model") if f not in col_map]
    if missing:
        return [], [f"Missing required column(s): {', '.join(missing)}"]

    parsed: List[dict] = []
    for row_idx, row in enumerate(rows_iter, start=2):
        if not row or not any(cell is not None and str(cell).strip() for cell in row):
            continue

        def cell(field: str) -> str:
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return ""
            val = row[idx]
            return "" if val is None else str(val).strip()

        parsed.append({
            "row_number": row_idx,
            "equipment": cell("equipment"),
            "description": cell("description"),
            "type_model": cell("type_model"),
            "manufacturer": cell("manufacturer") or None,
            "category": cell("category") or None,
            "component_position": cell("component_position") or None,
            "notes": cell("notes") or None,
            "document_url": cell("document_url") or None,
        })
    return parsed, errors


async def _build_equipment_lookup(user: dict) -> Dict[str, dict]:
    """Map lowercase name/tag/id → equipment node."""
    nodes = await db.equipment_nodes.find(
        merge_tenant_filter({}, user),
        {"_id": 0, "id": 1, "name": 1, "tag": 1},
    ).to_list(5000)
    lookup: Dict[str, dict] = {}
    for node in nodes:
        for key in (node.get("id"), node.get("name"), node.get("tag")):
            if key:
                lookup[str(key).strip().lower()] = node
    return lookup


async def _build_category_lookup(user: dict) -> Dict[str, str]:
    data = await list_categories(user)
    lookup: Dict[str, str] = {}
    for cat in data.get("categories") or []:
        lookup[cat["label"].lower()] = cat["id"]
        lookup[cat["value"].lower()] = cat["id"]
    return lookup


def _valid_url(url: str) -> bool:
    try:
        parsed = urlparse(url)
        return parsed.scheme in ("http", "https") and bool(parsed.netloc)
    except Exception:
        return False


async def validate_import_rows(user: dict, rows: List[dict]) -> dict:
    equipment_lookup = await _build_equipment_lookup(user)
    category_lookup = await _build_category_lookup(user)
    validated: List[dict] = []
    error_count = 0
    warning_count = 0

    for row in rows:
        messages: List[str] = []
        status = "ok"
        equipment_ref = (row.get("equipment") or "").strip()
        description = (row.get("description") or "").strip()
        type_model = (row.get("type_model") or "").strip()

        equipment_node = equipment_lookup.get(equipment_ref.lower()) if equipment_ref else None
        if not equipment_ref:
            messages.append("Equipment is required")
            status = "error"
        elif not equipment_node:
            messages.append(f"Equipment not found: {equipment_ref}")
            status = "error"

        if not description:
            messages.append("Spare part description is required")
            status = "error"
        if not type_model:
            messages.append("Type / Model is required")
            status = "error"

        category_id = None
        category_label = row.get("category")
        if category_label:
            category_id = category_lookup.get(category_label.lower())
            if not category_id:
                messages.append(f"Unknown category: {category_label}")
                if status != "error":
                    status = "warning"

        doc_url = row.get("document_url")
        if doc_url and not _valid_url(doc_url):
            messages.append("Invalid document URL")
            if status != "error":
                status = "warning"

        existing = None
        duplicate_link = False
        if description and type_model and status != "error":
            existing = await _find_by_duplicate_key(user, description, type_model)
            if existing:
                messages.append("Spare part already exists — will update and add equipment link")
                if status == "ok":
                    status = "warning"
                eq_id = equipment_node.get("id") if equipment_node else None
                if eq_id and any(l.get("equipment_id") == eq_id for l in existing.get("equipment_links") or []):
                    duplicate_link = True
                    messages.append("Equipment link already exists — row will be skipped on import")
                    status = "warning"

        if status == "error":
            error_count += 1
        elif status == "warning":
            warning_count += 1

        validated.append({
            **row,
            "equipment_id": equipment_node.get("id") if equipment_node else None,
            "equipment_name": equipment_node.get("name") if equipment_node else equipment_ref,
            "category_id": category_id,
            "status": status,
            "messages": messages,
            "existing_spare_part_id": existing.get("id") if existing else None,
            "skip_import": duplicate_link,
        })

    ok_count = len(validated) - error_count - warning_count
    return {
        "rows": validated,
        "summary": {
            "total": len(validated),
            "ok": ok_count,
            "warnings": warning_count,
            "errors": error_count,
            "importable": sum(1 for r in validated if r["status"] != "error" and not r.get("skip_import")),
        },
    }


async def execute_import(user: dict, rows: List[dict]) -> dict:
    created = 0
    updated = 0
    links_created = 0
    skipped = 0
    errors: List[dict] = []

    for row in rows:
        if row.get("status") == "error" or row.get("skip_import"):
            skipped += 1
            continue
        equipment_id = row.get("equipment_id")
        if not equipment_id:
            skipped += 1
            continue
        try:
            payload = SparePartCreate(
                description=row["description"],
                type_model=row["type_model"],
                manufacturer=row.get("manufacturer"),
                category_id=row.get("category_id"),
                notes=row.get("notes"),
                document_url=row.get("document_url"),
                equipment_links=[SparePartEquipmentLink(
                    equipment_id=equipment_id,
                    component_position=row.get("component_position"),
                )],
            )
            result = await create_spare_part(user, payload)
            if result.get("merged"):
                updated += 1
            else:
                created += 1
            links_created += 1
        except Exception as exc:
            errors.append({
                "row_number": row.get("row_number"),
                "message": str(exc),
            })

    return {
        "created": created,
        "updated": updated,
        "equipment_links_created": links_created,
        "skipped": skipped,
        "errors": errors,
        "warnings": len([r for r in rows if r.get("status") == "warning"]),
    }


def build_import_template_bytes() -> bytes:
    """Generate Excel import template in memory."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Spare Parts"
    headers = [
        "Equipment",
        "Spare Part Description",
        "Type / Model",
        "Manufacturer",
        "Category",
        "Component Position",
        "Notes",
        "Document URL",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="B45309", end_color="B45309", fill_type="solid")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    example = [
        "P-101",
        "Bearing Block",
        "PCJ30-N",
        "INA",
        "Bearing",
        "Drive End",
        "Example row — delete before import",
        "",
    ]
    example_font = Font(italic=True, color="64748B")
    for col, value in enumerate(example, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.font = example_font

    instructions = wb.create_sheet("Instructions")
    instructions["A1"] = "Required columns: Equipment, Spare Part Description, Type / Model"
    instructions["A2"] = "Equipment must match an existing equipment name or tag in AssetIQ."
    instructions["A3"] = "Duplicate spare parts (same description + type/model) update the existing record."

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
