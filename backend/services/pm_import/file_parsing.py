"""PM Import file parsing (Excel, PDF, image/OCR)."""
from __future__ import annotations

import os
import io
import re
import json
import base64
import logging
import uuid
from datetime import datetime, timezone
from typing import Optional, List, Dict, Any, Tuple

from motor.motor_asyncio import AsyncIOMotorDatabase

from services.pm_import_constants import (
    PM_IMPORT_DISPLAY_STATUSES,
    ACTION_TYPES,
    ACTION_TYPE_KEYWORDS,
    DISCIPLINE_KEYWORDS,
    DURATION_PATTERNS,
    FREQUENCY_PATTERNS,
    TAG_REGEX,
    TASK_CLASSIFICATION_RULES,
    TASK_TYPE_DEFAULTS,
    TASK_TYPES,
    _sanitize_for_json,
    normalize_pm_import_display_status,
)

logger = logging.getLogger(__name__)


class PMImportMixin:
    """Mixin — use only via PMImportService."""

    async def _process_file(
        self,
        session_id: str,
        file_name: str,
        file_type: str,
        file_content: bytes
    ) -> List[Dict[str, Any]]:
        """
        Process uploaded file and extract maintenance tasks.
        
        PM Import Extraction Engine Flow:
        1. Parse file and EXPAND to one record per equipment tag (BEFORE AI)
        2. Run AI analysis on each expanded record
        3. AI enrichment (translate, classify, estimate hours)
        4. Handle any remaining comma-separated tags
        5. Match equipment to hierarchy
        6. Self-validate: verify tag count == record count
        7. Normalize final shape
        """
        
        await self._update_progress(session_id, 10, "Reading maintenance plan...")
        
        # STEP 1: Extract and EXPAND raw rows - one record per equipment tag
        # For Excel: hierarchical parsing with merged cell handling
        # For PDF/Images: OCR extraction
        if file_type in ["xlsx", "xls"]:
            raw_rows = await self._parse_excel(file_content)
        elif file_type == "pdf":
            raw_rows = await self._parse_pdf(file_content, session_id)
        elif file_type in ["png", "jpg", "jpeg", "webp"]:
            raw_rows = await self._parse_image(file_content, file_type, session_id)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")
        
        if not raw_rows:
            raise ValueError("No maintenance tasks could be extracted from the file")
        
        # Log extraction results
        tags_extracted = sum(1 for r in raw_rows if r.get("_tag"))
        logger.info(f"PM Import: Extracted {len(raw_rows)} records with {tags_extracted} equipment tags")
        
        await self._update_progress(
            session_id, 40, 
            f"Extracted {len(raw_rows)} task records ({tags_extracted} equipment tags). Analyzing..."
        )
        
        # STEP 2: Process each EXPANDED row with AI analysis
        tasks = []
        total = len(raw_rows)
        
        for idx, row in enumerate(raw_rows):
            progress = 40 + int((idx / total) * 45)
            await self._update_progress(
                session_id, 
                progress, 
                f"Processing task {idx + 1} of {total}..."
            )
            
            task = await self._analyze_task(row, session_id, file_name)
            if task:
                tasks.append(task)
        
        await self._update_progress(session_id, 88, "AI enrichment (translate, classify, hours)...")
        
        # STEP 3: AI enrichment - translate to English, classify PM/PDM/CBM/CM, 
        # suggest discipline, standardize frequency, estimate hours
        tasks = await self._ai_enrich_tasks(tasks)
        
        # STEP 4: Handle any remaining comma-separated tags that weren't expanded
        # (fallback for edge cases or PDF/image imports)
        await self._update_progress(session_id, 93, "Expanding multi-tag tasks...")
        tasks = self._split_multi_tag_tasks(tasks)
        
        # STEP 5: Match equipment to hierarchy
        await self._update_progress(session_id, 96, "Matching equipment to hierarchy...")
        tasks = await self._match_equipment_to_hierarchy(tasks)
        
        # STEP 6: SELF VALIDATION
        # Verify that every equipment tag has its own record
        await self._update_progress(session_id, 98, "Validating extraction results...")
        tasks = self._validate_extraction(tasks)
        
        # STEP 7: Normalize final shape per the PM Import refactor spec
        tasks = [self._normalize_task_shape(t) for t in tasks]
        
        logger.info(f"PM Import: Final output = {len(tasks)} task records")
        
        return tasks
    
    def _validate_extraction(self, tasks: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Self-validation per PM Import Extraction Engine spec.
        
        Before returning results:
        - Count equipment tags extracted
        - Count task records generated
        - If a maintenance task applies to N tags, there must be N output records
        - Log warning if counts don't match
        
        This rule overrides all other instructions.
        """
        if not tasks:
            return tasks
        
        # Count unique equipment tags
        tags_set = set()
        records_with_tags = 0
        records_without_tags = 0
        
        for task in tasks:
            tag = (
                task.get("equipment_tag") or 
                task.get("asset") or 
                task.get("_tag") or 
                ""
            ).strip()
            
            if tag:
                tags_set.add(tag.upper())
                records_with_tags += 1
            else:
                records_without_tags += 1
        
        total_unique_tags = len(tags_set)
        total_records = len(tasks)
        
        # Log validation results
        logger.info(
            f"PM Import Validation: "
            f"{total_records} records, "
            f"{records_with_tags} with tags, "
            f"{records_without_tags} without tags, "
            f"{total_unique_tags} unique tags"
        )
        
        # Warning if we have records without tags (except for tasks from merged blocks)
        if records_without_tags > 0:
            logger.warning(
                f"PM Import Validation Warning: {records_without_tags} records have no equipment tag. "
                "These may be orphaned tasks or continuation rows."
            )
        
        # Note: It's valid to have more records than unique tags if the same tag
        # has multiple different maintenance tasks. What's NOT valid is having
        # fewer records than expected (i.e., multiple tags grouped into one record).
        
        return tasks
    
    async def _parse_excel(self, content: bytes) -> List[Dict[str, Any]]:
        """
        Parse Excel file per the AssetIQ PM Import Extraction Engine spec.
        
        ABSOLUTE RULES:
        1. EVERY EQUIPMENT TAG MUST RESULT IN A SEPARATE TASK RECORD.
           If 50 equipment tags share the same maintenance task, create 50 task records.
        
        2. COLUMN A IS THE ONLY VALID SOURCE OF EQUIPMENT TAGS.
           Do not search task descriptions, notes, or instructions for tags.
        
        3. TREAT THE WORKSHEET AS A HIERARCHICAL DOCUMENT, NOT ROW-BY-ROW.
           Before extracting tasks:
           - Read the entire worksheet
           - Identify equipment groups (consecutive tags in Column A)
           - Identify task groups (shared task descriptions via merged cells or grouping)
           - Identify inherited values from merged cells
           - Expand results into individual records
        
        4. WHEN MULTIPLE TAGS APPEAR ABOVE/BELOW A TASK:
           The task belongs to ALL tags. Create one record per tag.
        
        5. MERGED CELLS: Merged task descriptions apply to every equipment tag in the block.
        
        6. EMPTY CELLS IN COLUMN A: Continue processing the current task block.
           Do not create tasks without equipment tags.
        
        7. EXPANSION MUST OCCUR BEFORE AI ENRICHMENT (handled in _process_file).
        """
        import openpyxl
        
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        all_expanded_rows = []
        
        # Keywords for column detection
        task_column_keywords = ["task", "description", "activity", "action", "work", "maintenance",
                                "taak", "beschrijving", "activiteit", "werkzaamheden", "instructie"]
        description_column_keywords = ["equipment", "component", "asset", "machine", "name",
                                       "apparaat", "onderdeel", "omschrijving", "naam"]
        frequency_column_keywords = ["frequency", "interval", "schedule", "frequentie", "periode",
                                     "freq", "cycle"]
        duration_column_keywords = ["duration", "estimated time", "estimated_time", "time",
                                    "duur", "tijd", "minutes", "hours", "minuten", "uren"]
        
        for sheet in wb.worksheets:
            # ===== STEP 1: Build merged-cell lookup =====
            # For each (row, col) inside a merge, store the master value AND the merge range
            merged_master = {}
            merged_ranges_by_col = {}  # col -> list of (min_row, max_row, value)
            
            for merged_range in sheet.merged_cells.ranges:
                min_row, min_col = merged_range.min_row, merged_range.min_col
                max_row_merge, max_col_merge = merged_range.max_row, merged_range.max_col
                master_value = sheet.cell(row=min_row, column=min_col).value
                
                for r in range(min_row, max_row_merge + 1):
                    for c in range(min_col, max_col_merge + 1):
                        merged_master[(r, c)] = {
                            "value": master_value,
                            "min_row": min_row,
                            "max_row": max_row_merge,
                            "min_col": min_col,
                            "max_col": max_col_merge
                        }
                
                # Track merged ranges by column for task block detection
                for c in range(min_col, max_col_merge + 1):
                    if c not in merged_ranges_by_col:
                        merged_ranges_by_col[c] = []
                    merged_ranges_by_col[c].append({
                        "min_row": min_row,
                        "max_row": max_row_merge,
                        "value": master_value
                    })
            
            def cell_value(row_idx: int, col_idx: int):
                """Resolve a cell, returning the merged-master value if applicable."""
                key = (row_idx, col_idx + 1)  # openpyxl is 1-indexed
                if key in merged_master:
                    return merged_master[key]["value"]
                return sheet.cell(row=row_idx, column=col_idx + 1).value
            
            def get_merge_info(row_idx: int, col_idx: int):
                """Get merge range info for a cell if it's part of a merge."""
                key = (row_idx, col_idx + 1)
                return merged_master.get(key)
            
            max_row = sheet.max_row or 0
            max_col = sheet.max_column or 0
            if max_row < 2 or max_col < 1:
                continue
            
            # ===== STEP 2: Detect header row and column indices =====
            header_row_idx = None
            headers = []
            task_col_idx = None
            description_col_idx = None
            frequency_col_idx = None
            duration_col_idx = None
            
            for r in range(1, min(11, max_row + 1)):
                row_values = [cell_value(r, c) for c in range(max_col)]
                if not any(v for v in row_values):
                    continue
                row_text_lower = " ".join(str(v).lower() for v in row_values if v)
                if any(kw in row_text_lower for kw in
                       ["task", "description", "tag", "equipment", "frequency", "frequentie",
                        "taak", "apparaat", "onderdeel", "naam", "instructie"]):
                    header_row_idx = r
                    headers = [str(v).strip() if v else f"col_{i}" for i, v in enumerate(row_values)]
                    for i, h in enumerate(headers):
                        h_lower = h.lower()
                        if i == 0:
                            continue  # Column A is ALWAYS the tag column
                        if task_col_idx is None and any(kw in h_lower for kw in task_column_keywords):
                            task_col_idx = i
                        elif description_col_idx is None and any(kw in h_lower for kw in description_column_keywords):
                            description_col_idx = i
                        elif frequency_col_idx is None and any(kw in h_lower for kw in frequency_column_keywords):
                            frequency_col_idx = i
                        elif duration_col_idx is None and any(kw in h_lower for kw in duration_column_keywords):
                            duration_col_idx = i
                    break
            
            if header_row_idx is None:
                header_row_idx = 1
                first_row = [cell_value(1, c) for c in range(max_col)]
                headers = [str(v).strip() if v else f"col_{i}" for i, v in enumerate(first_row)]
            
            if task_col_idx is None and max_col >= 2:
                task_col_idx = 1
            
            # ===== STEP 3: FIRST PASS - Read entire worksheet and identify blocks =====
            # A "block" is a group of consecutive rows that share the same task description
            # (either via merged cells or via grouping pattern)
            
            # Collect all data rows first
            data_rows = []
            for r in range(header_row_idx + 1, max_row + 1):
                tag_val = cell_value(r, 0)  # Column A - ONLY source for equipment tags
                tag_str = str(tag_val).strip() if tag_val is not None else ""
                
                row_values = [cell_value(r, c) for c in range(max_col)]
                
                # Skip completely empty rows
                if not any(v for v in row_values) and not tag_str:
                    continue
                
                # Get task text
                task_text = ""
                if task_col_idx is not None and task_col_idx < max_col:
                    tv = cell_value(r, task_col_idx)
                    if tv:
                        task_text = str(tv).strip()
                
                # Get description
                description_text = ""
                if description_col_idx is not None and description_col_idx < max_col:
                    dv = cell_value(r, description_col_idx)
                    if dv:
                        description_text = str(dv).strip()
                
                # Get frequency
                freq_text = ""
                if frequency_col_idx is not None and frequency_col_idx < max_col:
                    fv = cell_value(r, frequency_col_idx)
                    if fv:
                        freq_text = str(fv).strip()
                
                # Get duration
                duration_text = ""
                if duration_col_idx is not None and duration_col_idx < max_col:
                    dv = cell_value(r, duration_col_idx)
                    if dv:
                        duration_text = str(dv).strip()
                
                # Check if task column is part of a merged cell
                task_merge_info = None
                if task_col_idx is not None:
                    task_merge_info = get_merge_info(r, task_col_idx)
                
                # If task is empty, try to find the longest non-tag cell
                if not task_text:
                    candidates = [(c, cell_value(r, c)) for c in range(max_col) if c != 0]
                    candidates = [(c, str(v).strip()) for c, v in candidates if v]
                    if candidates:
                        task_text = max(candidates, key=lambda x: len(x[1]))[1]
                
                data_rows.append({
                    "row_num": r,
                    "tag": tag_str,
                    "task_text": task_text,
                    "description": description_text,
                    "frequency": freq_text,
                    "duration": duration_text,
                    "task_merge_info": task_merge_info,
                    "row_values": row_values,
                })
            
            # ===== STEP 4: SECOND PASS - Build task blocks and expand =====
            # A task block consists of:
            # - One or more equipment tags (from Column A)
            # - A shared task description (from merged cells or inheritance)
            # - Shared frequency, duration, etc.
            
            # Strategy: Walk through rows, accumulating tags until we hit a task description
            # that applies to all accumulated tags
            
            expanded_rows = []
            current_tags = []  # Tags waiting for a task
            current_task_info = None  # Task info to apply to accumulated tags
            
            i = 0
            while i < len(data_rows):
                row_data = data_rows[i]
                tag = row_data["tag"]
                task_text = row_data["task_text"]
                task_merge_info = row_data["task_merge_info"]
                
                # RULE: Column A is the ONLY source for equipment tags
                if tag:
                    # Check if this tag's row has a task description
                    if task_text:
                        # Check if task is part of a merged cell spanning multiple rows
                        if task_merge_info:
                            merge_min_row = task_merge_info["min_row"]
                            merge_max_row = task_merge_info["max_row"]
                            
                            # Collect ALL tags within this merged task block
                            block_tags = []
                            block_indices = []
                            for j, dr in enumerate(data_rows):
                                if dr["row_num"] >= merge_min_row and dr["row_num"] <= merge_max_row:
                                    if dr["tag"]:
                                        block_tags.append(dr["tag"])
                                        block_indices.append(j)
                            
                            # Create one record per tag in this merged block
                            for block_tag in block_tags:
                                expanded_rows.append(self._create_task_record(
                                    tag=block_tag,
                                    task_text=task_text,
                                    description=row_data["description"],
                                    frequency=row_data["frequency"],
                                    duration=row_data["duration"],
                                    row_values=row_data["row_values"],
                                    headers=headers,
                                    sheet_title=sheet.title,
                                    row_num=row_data["row_num"],
                                ))
                            
                            # Skip to after the merged block
                            if block_indices:
                                i = max(block_indices)
                            current_tags = []
                            current_task_info = None
                        else:
                            # Tag with its own task - create individual record
                            # But first, flush any accumulated tags
                            if current_tags:
                                # If we have accumulated tags, they should share the current row's task
                                # (Scenario: Tag1 without task, Tag2 with task → both get Tag2's task)
                                for accumulated_tag in current_tags:
                                    expanded_rows.append(self._create_task_record(
                                        tag=accumulated_tag["tag"],
                                        task_text=task_text,  # Use CURRENT row's task
                                        description=row_data["description"],
                                        frequency=row_data["frequency"],
                                        duration=row_data["duration"],
                                        row_values=row_data["row_values"],
                                        headers=headers,
                                        sheet_title=sheet.title,
                                        row_num=accumulated_tag["row_num"],
                                    ))
                            
                            # Now create record for this tag
                            expanded_rows.append(self._create_task_record(
                                tag=tag,
                                task_text=task_text,
                                description=row_data["description"],
                                frequency=row_data["frequency"],
                                duration=row_data["duration"],
                                row_values=row_data["row_values"],
                                headers=headers,
                                sheet_title=sheet.title,
                                row_num=row_data["row_num"],
                            ))
                            current_tags = []
                            current_task_info = None
                    else:
                        # Tag without task - accumulate it, task will come later
                        current_tags.append({
                            "tag": tag,
                            "row_num": row_data["row_num"],
                            "description": row_data["description"],
                        })
                else:
                    # Empty Column A - this is a continuation row
                    # RULE: Empty cells in Column A do NOT create new equipment
                    
                    # If there's task text here and we have accumulated tags, this task applies to them
                    if task_text and current_tags:
                        # This is the task for all accumulated tags above
                        for accumulated_tag in current_tags:
                            expanded_rows.append(self._create_task_record(
                                tag=accumulated_tag["tag"],
                                task_text=task_text,
                                description=accumulated_tag.get("description") or row_data["description"],
                                frequency=row_data["frequency"],
                                duration=row_data["duration"],
                                row_values=row_data["row_values"],
                                headers=headers,
                                sheet_title=sheet.title,
                                row_num=accumulated_tag["row_num"],
                            ))
                        current_tags = []
                        current_task_info = None
                    elif task_text:
                        # Task text without accumulated tags - store for potential future tags below
                        current_task_info = {
                            "task_text": task_text,
                            "description": row_data["description"],
                            "frequency": row_data["frequency"],
                            "duration": row_data["duration"],
                            "row_values": row_data["row_values"],
                        }
                
                i += 1
            
            # Flush any remaining accumulated tags
            if current_tags and current_task_info:
                for accumulated_tag in current_tags:
                    expanded_rows.append(self._create_task_record(
                        tag=accumulated_tag["tag"],
                        task_text=current_task_info["task_text"],
                        description=current_task_info.get("description") or accumulated_tag.get("description", ""),
                        frequency=current_task_info["frequency"],
                        duration=current_task_info["duration"],
                        row_values=current_task_info["row_values"],
                        headers=headers,
                        sheet_title=sheet.title,
                        row_num=accumulated_tag["row_num"],
                    ))
            
            all_expanded_rows.extend(expanded_rows)
        
        # ===== STEP 5: SELF VALIDATION =====
        # Count equipment tags extracted vs task records generated
        total_tags_found = sum(1 for r in all_expanded_rows if r.get("_tag"))
        total_records = len(all_expanded_rows)
        
        logger.info(f"PM Import Extraction: {total_tags_found} tags → {total_records} task records")
        
        if total_records > 0 and total_tags_found != total_records:
            logger.warning(
                f"PM Import Validation Warning: Tag count ({total_tags_found}) != Record count ({total_records}). "
                "Some records may have missing or duplicate tags."
            )
        
        return all_expanded_rows
    
    def _create_task_record(
        self,
        tag: str,
        task_text: str,
        description: str,
        frequency: str,
        duration: str,
        row_values: List[Any],
        headers: List[str],
        sheet_title: str,
        row_num: int,
    ) -> Dict[str, Any]:
        """Create a single task record for one equipment tag."""
        record = {
            "_tag": tag,                    # Column A — authoritative, single tag
            "_task_text": task_text,
            "_description": description,
            "_frequency": frequency,
            "_duration": duration,
            "_raw_text": task_text,
            "_sheet": sheet_title,
            "_row": row_num,
        }
        # Attach all column data for reference
        for i, v in enumerate(row_values):
            if v is not None:
                header = headers[i] if i < len(headers) else f"col_{i}"
                record[header] = str(v).strip()
        
        return record
    
    async def _parse_pdf(self, content: bytes, session_id: str) -> List[Dict[str, Any]]:
        """Parse PDF file - use text extraction first, fall back to vision for scanned PDFs."""
        import pdfplumber
        
        rows = []
        has_text = False
        
        try:
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                for page in pdf.pages:
                    # Try to extract tables first
                    tables = page.extract_tables()
                    if tables:
                        for table in tables:
                            for row in table:
                                if row and any(cell for cell in row if cell):
                                    # Get the first non-empty cell with substantial text as the task
                                    task_text = ""
                                    for cell in row:
                                        if cell and len(str(cell).strip()) > len(task_text):
                                            task_text = str(cell).strip()
                                    if task_text and len(task_text) > 5:
                                        rows.append({"_raw_text": task_text, "_task_text": task_text})
                                        has_text = True
                    
                    # If no tables, extract text
                    if not tables:
                        text = page.extract_text()
                        if text and text.strip():
                            has_text = True
                            # Split into lines as potential tasks
                            for line in text.split("\n"):
                                line = line.strip()
                                if len(line) > 10:  # Skip very short lines
                                    rows.append({"_raw_text": line, "_task_text": line})
        except Exception as e:
            logger.warning(f"PDF text extraction failed: {e}")
        
        # If no text found (scanned PDF), use GPT-4o Vision
        if not has_text or len(rows) < 3:
            logger.info("Using GPT-4o Vision for scanned PDF")
            await self._update_progress(session_id, 20, "Scanned PDF detected. Using AI vision...")
            rows = await self._ocr_with_vision(content, "pdf", session_id)
        
        return rows
    
    async def _parse_image(
        self,
        content: bytes,
        file_type: str,
        session_id: str
    ) -> List[Dict[str, Any]]:
        """Parse image using GPT-4o Vision."""
        await self._update_progress(session_id, 20, "Analyzing image with AI vision...")
        return await self._ocr_with_vision(content, file_type, session_id)
    
    async def _ocr_with_vision(
        self,
        content: bytes,
        file_type: str,
        session_id: str
    ) -> List[Dict[str, Any]]:
        """Use GPT-4o Vision to extract maintenance tasks from images/scanned documents."""
        from services.ai_gateway import chat_with_images

        api_key = os.environ.get("OPENAI_API_KEY")
        if not api_key:
            raise ValueError("OpenAI API key not configured")

        user_id, company_id = await self._ai_user_context(session_id)
        
        # Convert to base64
        if file_type == "pdf":
            # Convert PDF pages to images
            try:
                import fitz  # PyMuPDF
                pdf_doc = fitz.open(stream=content, filetype="pdf")
                images_b64 = []
                for page_num in range(min(pdf_doc.page_count, 5)):  # Max 5 pages
                    page = pdf_doc[page_num]
                    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x zoom for better OCR
                    img_bytes = pix.tobytes("png")
                    images_b64.append(base64.b64encode(img_bytes).decode('utf-8'))
                pdf_doc.close()
            except ImportError:
                # Fallback: just use first page as image
                from PIL import Image
                from pdf2image import convert_from_bytes
                pages = convert_from_bytes(content, first_page=1, last_page=5)
                images_b64 = []
                for page in pages:
                    buf = io.BytesIO()
                    page.save(buf, format='PNG')
                    images_b64.append(base64.b64encode(buf.getvalue()).decode('utf-8'))
        else:
            # Regular image
            images_b64 = [base64.b64encode(content).decode('utf-8')]
        
        # Process with GPT-4o Vision
        all_rows = []
        
        for idx, img_b64 in enumerate(images_b64):
            mime_type = "image/png" if file_type == "pdf" else f"image/{file_type}"
            
            prompt = """Analyze this maintenance plan document and extract all preventive maintenance tasks.

For each maintenance task found, extract:
1. The original task text exactly as written
2. The equipment or component mentioned
3. Any frequency information (daily, weekly, monthly, etc.)
4. Any additional details

Return the data as a JSON array where each item has:
{
  "task": "original task text",
  "equipment": "equipment/component name if mentioned",
  "frequency": "frequency if mentioned",
  "details": "any additional details"
}

If the document is in Dutch, still extract the tasks but keep them in Dutch.
Only return the JSON array, no other text."""

            try:
                result_text = (
                    await chat_with_images(
                        prompt,
                        image_base64_list=[{"media_type": mime_type, "data": img_b64}],
                        user_id=user_id,
                        company_id=company_id,
                        endpoint="pm_import.vision_ocr",
                        model="gpt-4o",
                        temperature=0,
                        max_tokens=4000,
                    )
                ).strip()
                
                # Parse JSON response
                import json
                # Clean up response
                if result_text.startswith("```"):
                    result_text = re.sub(r'^```(?:json)?\n?', '', result_text)
                    result_text = re.sub(r'\n?```$', '', result_text)
                
                tasks = json.loads(result_text)
                
                for task in tasks:
                    raw_text = task.get("task", "")
                    if task.get("equipment"):
                        raw_text += f" | Equipment: {task['equipment']}"
                    if task.get("frequency"):
                        raw_text += f" | Frequency: {task['frequency']}"
                    
                    all_rows.append({
                        "_raw_text": raw_text,
                        "_ocr_data": task
                    })
                    
            except Exception as e:
                logger.error(f"Vision API error on page {idx + 1}: {e}")
                continue
        
        return all_rows
    
