"""Production log file parsing helpers — extracted from production_logs_service."""
from __future__ import annotations

import csv
import io
import logging
import re
from datetime import datetime
from typing import Dict, List, Optional

from pydantic import BaseModel

logger = logging.getLogger(__name__)


class ColumnMapping(BaseModel):
    timestamp: Optional[str] = None
    asset_id: Optional[str] = None
    status: Optional[str] = None
    event_type: Optional[str] = None
    metric_columns: List[str] = []


class ParseTemplate(BaseModel):
    delimiter: str = ","
    has_header: bool = True
    skip_rows: int = 0
    timestamp_format: Optional[str] = None
    column_mapping: ColumnMapping = ColumnMapping()
    base_date_location: Optional[dict] = None
    header_metadata: Optional[List[dict]] = None
    secondary_sheet: Optional[dict] = None

def _normalize_column_name(name: str) -> str:
    """Normalize column name for fuzzy matching."""
    # Lowercase, remove special chars, collapse whitespace
    normalized = name.lower().strip()
    normalized = re.sub(r'[^a-z0-9]', '', normalized)
    return normalized


def _tokenize_column(name: str) -> set:
    """Split column name into lowercase word tokens for overlap matching."""
    return set(re.findall(r'[a-z0-9]+', name.lower()))


def _similarity_ratio(a: str, b: str) -> float:
    """Simple character-level similarity ratio (0-1)."""
    if not a or not b:
        return 0.0
    a_lower, b_lower = a.lower(), b.lower()
    if a_lower == b_lower:
        return 1.0
    # Longest common subsequence ratio
    from difflib import SequenceMatcher
    return SequenceMatcher(None, a_lower, b_lower).ratio()


def _fuzzy_match_columns(
    file_columns: List[str], 
    template_mapping: ColumnMapping,
    column_aliases: Dict[str, List[str]] = None
) -> Dict[str, str]:
    """
    Match file columns to template columns using relaxed fuzzy matching.
    Returns a mapping of {template_column: matched_file_column}.
    """
    column_aliases = column_aliases or {}
    
    # Build lookups
    file_col_lookup = {}       # normalized -> original
    file_col_tokens = {}       # original -> token set
    for col in file_columns:
        norm = _normalize_column_name(col)
        file_col_lookup[norm] = col
        file_col_tokens[col] = _tokenize_column(col)
        # Also try without common prefixes/suffixes
        for prefix in ['col_', 'field_', 'column_']:
            if norm.startswith(prefix):
                file_col_lookup[norm[len(prefix):]] = col
    
    def find_match(template_col: str) -> Optional[str]:
        """Find matching file column for a template column."""
        if not template_col:
            return None
            
        # 1. Exact match
        if template_col in file_columns:
            return template_col
            
        # 2. Normalized exact match
        norm_template = _normalize_column_name(template_col)
        if norm_template in file_col_lookup:
            return file_col_lookup[norm_template]
        
        # 3. Check aliases
        aliases = column_aliases.get(template_col, [])
        for alias in aliases:
            if alias in file_columns:
                return alias
            norm_alias = _normalize_column_name(alias)
            if norm_alias in file_col_lookup:
                return file_col_lookup[norm_alias]
        
        # 4. Substring match (template col contains or is contained in file col)
        for norm_file, orig_file in file_col_lookup.items():
            if len(norm_template) >= 2 and len(norm_file) >= 2:
                if norm_template in norm_file or norm_file in norm_template:
                    return orig_file

        # 5. Token overlap — if most words match (e.g. "T Product IR" vs "T Prod IR")
        template_tokens = _tokenize_column(template_col)
        if template_tokens:
            best_overlap = 0
            best_col = None
            for col in file_columns:
                col_tokens = file_col_tokens.get(col, set())
                if not col_tokens:
                    continue
                overlap = len(template_tokens & col_tokens)
                min_len = min(len(template_tokens), len(col_tokens))
                if min_len > 0 and overlap / min_len >= 0.5 and overlap > best_overlap:
                    best_overlap = overlap
                    best_col = col
            if best_col and best_overlap >= 1:
                return best_col

        # 6. Similarity ratio — catch typos and minor variations (threshold 0.6)
        best_ratio = 0
        best_col = None
        for norm_file, orig_file in file_col_lookup.items():
            ratio = _similarity_ratio(norm_template, norm_file)
            if ratio > best_ratio:
                best_ratio = ratio
                best_col = orig_file
        if best_ratio >= 0.6 and best_col:
            return best_col
                
        return None
    
    # Match all template columns
    matched = {}
    
    if template_mapping.timestamp:
        match = find_match(template_mapping.timestamp)
        matched['timestamp'] = match or template_mapping.timestamp
        
    if template_mapping.asset_id:
        match = find_match(template_mapping.asset_id)
        matched['asset_id'] = match or template_mapping.asset_id
        
    if template_mapping.status:
        match = find_match(template_mapping.status)
        matched['status'] = match or template_mapping.status
        
    if template_mapping.event_type:
        match = find_match(template_mapping.event_type)
        matched['event_type'] = match or template_mapping.event_type
    
    # Match metric columns
    matched_metrics = []
    for metric_col in template_mapping.metric_columns:
        match = find_match(metric_col)
        if match:
            matched_metrics.append(match)
        else:
            matched_metrics.append(metric_col)
    matched['metric_columns'] = matched_metrics
    
    return matched


def _get_ext(filename: str) -> str:
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def _detect_delimiter(sample: str) -> str:
    """Auto-detect delimiter from first few lines."""
    for delim in [",", ";", "\t", "|"]:
        if delim in sample:
            return delim
    return ","


def _parse_timestamp(val: str, fmt: Optional[str] = None, base_date: Optional[str] = None) -> Optional[str]:
    """Try to parse a timestamp string into ISO format.
    
    Args:
        val: The timestamp string to parse
        fmt: Optional format string
        base_date: Optional base date (YYYY-MM-DD) to combine with time-only values
    """
    if not val or not val.strip():
        return None
    val = str(val).strip()
    
    # Handle time-only values (HH:MM:SS or HH:MM)
    time_only_formats = ["%H:%M:%S", "%H:%M"]
    for tf in time_only_formats:
        try:
            time_obj = datetime.strptime(val, tf)
            # If we have a base date, combine them
            if base_date:
                try:
                    date_obj = datetime.strptime(base_date, "%Y-%m-%d")
                    combined = datetime.combine(date_obj.date(), time_obj.time())
                    return combined.isoformat()
                except:
                    pass
            # Otherwise, use today's date as fallback
            today = datetime.now().date()
            combined = datetime.combine(today, time_obj.time())
            return combined.isoformat()
        except ValueError:
            continue
    
    formats = []
    if fmt:
        formats.append(fmt)
    formats.extend([
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%m/%d/%Y %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d",
        "%d-%m-%Y",
    ])
    for f in formats:
        try:
            dt = datetime.strptime(val, f)
            return dt.isoformat()
        except ValueError:
            continue
    # Try ISO parse as last resort
    try:
        dt = datetime.fromisoformat(val.replace("Z", "+00:00"))
        return dt.isoformat()
    except Exception:
        return None


def _classify_event(status: Optional[str], metrics: dict) -> str:
    """Basic event classification rules."""
    if not status:
        return "normal"
    s = str(status).lower().strip()
    if any(w in s for w in ["stop", "down", "off", "shutdown", "error", "fault", "trip"]):
        return "downtime"
    if any(w in s for w in ["waste", "reject", "scrap", "discard"]):
        return "waste"
    if any(w in s for w in ["alarm", "alert", "warning", "abnormal", "high", "low", "critical"]):
        return "alarm"
    return "normal"


def _parse_csv_content(content: str, template: ParseTemplate) -> List[dict]:
    """Parse CSV/TXT content using a template configuration."""
    lines = content.splitlines()
    if template.skip_rows > 0:
        lines = lines[template.skip_rows:]
    if not lines:
        return []

    delimiter = template.delimiter
    if delimiter == "\\t":
        delimiter = "\t"

    reader = csv.reader(lines, delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return []

    headers = None
    data_start = 0
    if template.has_header:
        headers = [h.strip() for h in rows[0]]
        data_start = 1
    else:
        headers = [f"col_{i}" for i in range(len(rows[0]))]

    mapping = template.column_mapping
    records = []
    errors = []

    for row_idx, row in enumerate(rows[data_start:], start=data_start + 1):
        if not row or all(not cell.strip() for cell in row):
            continue  # Skip empty rows

        record = {"_row": row_idx}

        # Build a dict from row
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = val.strip()

        # Map timestamp
        ts_val = row_dict.get(mapping.timestamp) if mapping.timestamp else None
        if ts_val:
            parsed_ts = _parse_timestamp(ts_val, template.timestamp_format)
            record["timestamp"] = parsed_ts
            if not parsed_ts:
                record["_errors"] = record.get("_errors", []) + [f"Invalid timestamp: {ts_val}"]
        else:
            record["timestamp"] = None
            if mapping.timestamp:
                record["_errors"] = record.get("_errors", []) + ["Missing timestamp"]

        # Map asset_id - support static values with "__static:" prefix
        if mapping.asset_id:
            if mapping.asset_id.startswith("__static:"):
                record["asset_id"] = mapping.asset_id.replace("__static:", "")
            else:
                record["asset_id"] = row_dict.get(mapping.asset_id, "").strip()
        else:
            record["asset_id"] = None

        # Map status
        record["status"] = row_dict.get(mapping.status, "").strip() if mapping.status else None

        # Map event_type (auto-classify if not mapped)
        if mapping.event_type and row_dict.get(mapping.event_type):
            record["event_type"] = row_dict[mapping.event_type].strip().lower()
        else:
            record["event_type"] = _classify_event(record.get("status"), {})

        # Map metrics (all metric_columns)
        metrics = {}
        for col in mapping.metric_columns:
            val = row_dict.get(col, "").strip()
            if val:
                try:
                    metrics[col] = float(val.replace(",", "."))
                except ValueError:
                    metrics[col] = val
        record["metrics"] = metrics

        # Also classify based on metrics
        if record["event_type"] == "normal":
            record["event_type"] = _classify_event(record.get("status"), metrics)

        records.append(record)

    return records


def _parse_excel_content(file_bytes: bytes, ext: str, template: ParseTemplate) -> List[dict]:
    """Parse XLSX/XLS content using openpyxl or xlrd."""
    import openpyxl

    all_rows = []  # Keep all rows including header section for date extraction
    secondary_data = {}  # Data from secondary sheet keyed by time
    
    if ext == "xls":
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        ws = wb.sheet_by_index(0)
        for r in range(ws.nrows):
            all_rows.append([str(ws.cell_value(r, c)) for c in range(ws.ncols)])
        
        # Handle secondary sheet for xls
        if template.secondary_sheet and len(wb.sheet_names()) > 1:
            sheet_name = template.secondary_sheet.get("sheet", "")
            if sheet_name in wb.sheet_names():
                ws2 = wb.sheet_by_name(sheet_name)
                start_row = template.secondary_sheet.get("start_row", 1) - 1
                time_col = template.secondary_sheet.get("time_col", 0)
                data_cols = template.secondary_sheet.get("data_cols", [])
                
                for r in range(start_row, ws2.nrows):
                    time_val = str(ws2.cell_value(r, time_col)).strip()
                    if time_val and time_val != "None":
                        row_data = {}
                        for dc in data_cols:
                            col_idx = dc.get("col", 0)
                            col_name = dc.get("name", f"col_{col_idx}")
                            if col_idx < ws2.ncols:
                                val = str(ws2.cell_value(r, col_idx)).strip()
                                if val and val != "None":
                                    row_data[col_name] = val
                        if row_data:
                            # Normalize time for matching (handle both HH:MM:SS and datetime formats)
                            time_key = time_val.split()[-1] if " " in time_val else time_val
                            secondary_data[time_key] = row_data
                            logger.debug(f"[Excel Parse] Secondary sheet time {time_key}: {row_data}")
    else:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        # Use first sheet by index (not active, which may be different)
        ws = wb.worksheets[0]
        for row in ws.iter_rows(values_only=True):
            all_rows.append([str(c) if c is not None else "" for c in row])
        
        # Handle secondary sheet for xlsx
        if template.secondary_sheet:
            sheet_name = template.secondary_sheet.get("sheet", "")
            if sheet_name in wb.sheetnames:
                ws2 = wb[sheet_name]
                start_row = template.secondary_sheet.get("start_row", 1)
                time_col = template.secondary_sheet.get("time_col", 0)
                data_cols = template.secondary_sheet.get("data_cols", [])
                
                for i, row in enumerate(ws2.iter_rows(min_row=start_row, values_only=True), start=start_row):
                    time_val = str(row[time_col]).strip() if time_col < len(row) and row[time_col] else ""
                    if time_val and time_val != "None":
                        row_data = {}
                        for dc in data_cols:
                            col_idx = dc.get("col", 0)
                            col_name = dc.get("name", f"col_{col_idx}")
                            if col_idx < len(row) and row[col_idx]:
                                val = str(row[col_idx]).strip()
                                if val and val != "None":
                                    row_data[col_name] = val
                        if row_data:
                            # Normalize time for matching
                            time_key = time_val.split()[-1] if " " in time_val else time_val
                            secondary_data[time_key] = row_data
                            logger.debug(f"[Excel Parse] Secondary sheet time {time_key}: {row_data}")
        
        wb.close()

    if not all_rows:
        return []

    # Extract base date from specific location or auto-detect from header section
    base_date = None
    
    # First try specific location if provided
    if template.base_date_location:
        try:
            row_idx = template.base_date_location.get("row", 1) - 1  # Convert to 0-indexed
            col_idx = template.base_date_location.get("col", 0)
            if row_idx < len(all_rows) and col_idx < len(all_rows[row_idx]):
                cell_str = str(all_rows[row_idx][col_idx]).strip()
                # Try common date formats
                for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"]:
                    try:
                        dt = datetime.strptime(cell_str.split()[0] if ' ' in cell_str else cell_str, fmt.split()[0])
                        base_date = dt.strftime("%Y-%m-%d")
                        logger.info(f"[Excel Parse] Extracted base date from row {row_idx+1}, col {col_idx}: {base_date}")
                        break
                    except ValueError:
                        continue
        except Exception as e:
            logger.warning(f"[Excel Parse] Failed to extract date from specified location: {e}")
    
    # Fallback: auto-detect from header section
    if not base_date and template.skip_rows > 0:
        header_section = all_rows[:template.skip_rows]
        for row in header_section:
            for cell in row:
                if cell:
                    cell_str = str(cell).strip()
                    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"]:
                        try:
                            dt = datetime.strptime(cell_str.split()[0] if ' ' in cell_str else cell_str, fmt.split()[0])
                            base_date = dt.strftime("%Y-%m-%d")
                            logger.debug(f"[Excel Parse] Auto-detected base date in header: {base_date}")
                            break
                        except ValueError:
                            continue
                    if base_date:
                        break
            if base_date:
                break

    # Fallback: scan ALL rows for any date-like cell
    if not base_date:
        for row in all_rows:
            for cell in row:
                if cell:
                    cell_str = str(cell).strip()
                    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"]:
                        try:
                            dt = datetime.strptime(cell_str.split()[0] if ' ' in cell_str else cell_str, fmt.split()[0])
                            if 2019 <= dt.year <= 2030:
                                base_date = dt.strftime("%Y-%m-%d")
                                logger.info(f"[Excel Parse] Found base date in row scan: {base_date}")
                                break
                        except ValueError:
                            continue
                    if base_date:
                        break
            if base_date:
                break

    # Skip rows to get to data
    rows = all_rows[template.skip_rows:] if template.skip_rows > 0 else all_rows

    mapping = template.column_mapping

    # Fallback: if skip_rows lands on empty/wrong rows, auto-detect header row by looking for TIME column
    if rows and template.has_header:
        first_row_vals = [h.strip().upper() for h in rows[0] if h.strip()]
        time_col_name = (mapping.timestamp or "TIME").upper()
        if time_col_name not in [v.upper() for v in first_row_vals]:
            # Header not found at skip_rows — scan all rows for the header
            logger.info(f"[Excel Parse] Header '{time_col_name}' not found at skip_rows={template.skip_rows}, scanning...")
            for scan_idx, scan_row in enumerate(all_rows):
                scan_vals = [h.strip() for h in scan_row if h.strip()]
                if any(v.upper() == time_col_name for v in scan_vals):
                    logger.info(f"[Excel Parse] Found header at row {scan_idx + 1} (original skip_rows={template.skip_rows})")
                    rows = all_rows[scan_idx:]
                    break

    if not rows:
        return []

    # Extract header metadata if specified
    header_metadata = {}
    if template.header_metadata:
        for meta in template.header_metadata:
            try:
                name = meta.get("name", "")
                row_idx = meta.get("row", 1) - 1  # Convert to 0-indexed
                col_idx = meta.get("col", 0)
                if name and row_idx < len(all_rows) and col_idx < len(all_rows[row_idx]):
                    value = str(all_rows[row_idx][col_idx]).strip()
                    if value and value.lower() != "none":
                        header_metadata[name] = value
                        logger.debug(f"[Excel Parse] Extracted header metadata: {name} = {value}")
            except Exception as e:
                logger.warning(f"[Excel Parse] Failed to extract header metadata {meta}: {e}")

    headers = None
    data_start = 0
    if template.has_header:
        headers = [h.strip() for h in rows[0]]
        data_start = 1
    else:
        headers = [f"col_{i}" for i in range(len(rows[0]))]

    records = []
    for row_idx, row in enumerate(rows[data_start:], start=data_start + 1):
        if all(not cell.strip() for cell in row):
            continue

        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = val.strip()

        # Start with header metadata attached to every record
        record = {"_row": row_idx, **header_metadata}

        ts_val = row_dict.get(mapping.timestamp) if mapping.timestamp else None
        if ts_val:
            # Pass base_date for time-only values
            parsed_ts = _parse_timestamp(ts_val, template.timestamp_format, base_date)
            record["timestamp"] = parsed_ts
            if not parsed_ts:
                record["_errors"] = record.get("_errors", []) + [f"Invalid timestamp: {ts_val}"]
            
            # Join secondary sheet data by time
            if secondary_data and ts_val:
                time_key = ts_val.split()[-1] if " " in ts_val else ts_val
                if time_key in secondary_data:
                    record.update(secondary_data[time_key])
                    logger.debug(f"[Excel Parse] Joined secondary data for time {time_key}")
        else:
            record["timestamp"] = None
            if mapping.timestamp:
                record["_errors"] = record.get("_errors", []) + ["Missing timestamp"]

        # Handle asset_id - support static values with "__static:" prefix
        if mapping.asset_id:
            if mapping.asset_id.startswith("__static:"):
                # Static asset ID for all records
                record["asset_id"] = mapping.asset_id.replace("__static:", "")
            else:
                # Dynamic asset ID from column
                record["asset_id"] = row_dict.get(mapping.asset_id, "").strip()
        else:
            record["asset_id"] = None
            
        record["status"] = row_dict.get(mapping.status, "").strip() if mapping.status else None

        if mapping.event_type and row_dict.get(mapping.event_type):
            record["event_type"] = row_dict[mapping.event_type].strip().lower()
        else:
            record["event_type"] = _classify_event(record.get("status"), {})

        metrics = {}
        for col in mapping.metric_columns:
            val = row_dict.get(col, "").strip()
            if val:
                try:
                    metrics[col] = float(val.replace(",", "."))
                except ValueError:
                    metrics[col] = val
        record["metrics"] = metrics

        if record["event_type"] == "normal":
            record["event_type"] = _classify_event(record.get("status"), metrics)

        records.append(record)

    return records
