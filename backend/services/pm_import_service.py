"""
PM Intelligence Import Service - Converts maintenance plans to failure mode intelligence.

This service handles:
1. File parsing (Excel, PDF, Images via GPT-4o Vision)
2. AI-powered task classification
3. Failure mode extraction and mapping
4. Library matching
5. Confidence scoring
"""

import os
import io
import re
import json
import base64
import logging
import uuid
from datetime import datetime, timezone
from typing import Optional, List, Dict, Any, Tuple
from motor.motor_asyncio import AsyncIOMotorDatabase

logger = logging.getLogger(__name__)

# Tag-like equipment identifier extraction (best-effort).
# Examples: P-101, HX-201, 1F-3001, 1F-3001-0129
TAG_REGEX = re.compile(
    r"\b(?:"
    r"[A-Z]{1,5}\s*-\s*\d{2,6}(?:-\d{2,6})?"          # P-101, 1F-3001, 1F-3001-0129
    r"|"
    r"\d{1,3}[A-Z]{1,5}\s*-\s*\d{2,6}(?:-\d{2,6})?"   # 1F-3001 (number+letters prefix)
    r")\b"
)

# Task type classifications
TASK_TYPES = [
    "Inspection",
    "Lubrication",
    "Calibration",
    "Replacement",
    "Cleaning",
    "Adjustment",
    "Monitoring",
    "Unknown"
]

# Maintenance action type (CM / PM / PDM)
ACTION_TYPES = ["PM", "PDM", "CM"]

# Map task type → default action_type and discipline
# PM = Preventive (time-based), PDM = Predictive (condition-based), CM = Corrective
TASK_TYPE_DEFAULTS = {
    "Inspection":   {"action_type": "PM",  "discipline": "Inspection"},
    "Lubrication":  {"action_type": "PM",  "discipline": "Mechanical"},
    "Calibration":  {"action_type": "PM",  "discipline": "Instrumentation"},
    "Replacement":  {"action_type": "PM",  "discipline": "Mechanical"},
    "Cleaning":     {"action_type": "PM",  "discipline": "Process"},
    "Adjustment":   {"action_type": "PM",  "discipline": "Mechanical"},
    "Monitoring":   {"action_type": "PDM", "discipline": "Reliability"},
    "Unknown":      {"action_type": "PM",  "discipline": "Maintenance"},
}

# Discipline detection via component keywords (overrides task-type defaults)
DISCIPLINE_KEYWORDS = {
    "Electrical": [
        "cable", "wire", "switchgear", "transformer", "breaker", "fuse",
        "relay", "circuit", "panel", "motor winding", "vfd", "drive",
        "schakelaar", "kabel"
    ],
    "Instrumentation": [
        "sensor", "transmitter", "controller", "plc", "gauge", "indicator",
        "level switch", "flow meter", "pressure transmitter", "temperature transmitter",
        "calibrate", "calibration", "kalibreer", "control valve", "positioner"
    ],
    "Process": [
        "tank", "vessel", "column", "reactor", "exchanger", "heat exchanger",
        "filter", "strainer", "separator", "scrubber", "drum",
        "process line", "piping", "line", "leiding"
    ],
    "Mechanical": [
        "pump", "compressor", "motor", "fan", "blower", "turbine",
        "bearing", "seal", "gear", "gearbox", "coupling", "shaft",
        "belt", "chain", "roller", "conveyor", "valve body",
        "pomp", "lager", "tandwiel", "ventilator"
    ],
}

# Action type detection via task text keywords (overrides task-type defaults)
ACTION_TYPE_KEYWORDS = {
    "PDM": [
        "vibration analysis", "vibration monitoring", "thermography",
        "thermal imaging", "oil analysis", "trillingsanalyse",
        "ultrasonic", "condition monitor", "trend", "predictive",
        "motor current signature", "infrared", "acoustic emission"
    ],
    "CM": [
        "repair", "fix", "rebuild", "restore after failure", "corrective",
        "herstel", "repareer"
    ],
}

# Keyword-based pre-classification rules
TASK_CLASSIFICATION_RULES = {
    "lubrication": {
        "keywords": ["grease", "lubricate", "oil", "smeer", "vet", "lub", "lube"],
        "failure_modes": ["Lubrication starvation", "Bearing wear", "Bearing seizure", "Overheating"],
        "failure_mechanisms": ["Insufficient lubrication", "Grease contamination", "Grease degradation"],
        "detection_methods": ["Temperature monitoring", "Vibration analysis", "Noise detection", "Visual inspection"]
    },
    "inspection": {
        "keywords": ["inspect", "visual check", "controleer", "listen", "leakage", "check pressure", "check", "examine", "verify", "control"],
        "failure_modes": ["Leak", "Abnormal noise", "Loose part", "Crack", "Pressure loss", "Overheating", "Corrosion"],
        "failure_mechanisms": ["Wear", "Fatigue", "Corrosion", "Loosening"],
        "detection_methods": ["Visual inspection", "Auditory check", "Pressure measurement", "Temperature monitoring"]
    },
    "calibration": {
        "keywords": ["calibrate", "kalibreren", "adjust", "afstellen", "tune", "zero", "set point"],
        "failure_modes": ["Sensor drift", "False reading", "Instrument failure", "Measurement error"],
        "failure_mechanisms": ["Sensor degradation", "Electronic drift", "Environmental effects"],
        "detection_methods": ["Reference comparison", "Trend analysis", "Deviation monitoring"]
    },
    "replacement": {
        "keywords": ["replace", "vervangen", "overhaul", "flush", "change", "renew", "swap"],
        "failure_modes": ["Consumable degradation", "Contamination", "Wear-out", "Fluid degradation"],
        "failure_mechanisms": ["Age-related degradation", "Contamination buildup", "Material fatigue"],
        "detection_methods": ["Time-based schedule", "Condition monitoring", "Sample analysis"]
    },
    "cleaning": {
        "keywords": ["clean", "reinig", "cooling", "blow out", "spoelen", "wash", "purge", "filter"],
        "failure_modes": ["Fouling", "Overheating", "Blockage", "Contamination", "Reduced efficiency"],
        "failure_mechanisms": ["Buildup accumulation", "Debris collection", "Scale formation"],
        "detection_methods": ["Visual inspection", "Pressure drop measurement", "Temperature monitoring"]
    },
    "adjustment": {
        "keywords": ["adjust", "tension", "align", "tighten", "torque", "set", "position"],
        "failure_modes": ["Misalignment", "Loose components", "Belt slip", "Coupling failure"],
        "failure_mechanisms": ["Vibration loosening", "Thermal expansion", "Wear"],
        "detection_methods": ["Alignment check", "Torque verification", "Vibration analysis"]
    },
    "monitoring": {
        "keywords": ["monitor", "record", "log", "measure", "track", "observe", "trend"],
        "failure_modes": ["Performance degradation", "Abnormal operation", "Efficiency loss"],
        "failure_mechanisms": ["Progressive wear", "System degradation"],
        "detection_methods": ["Trend analysis", "Threshold monitoring", "Data logging"]
    }
}

# Frequency pattern matching
FREQUENCY_PATTERNS = [
    (r'\b(daily|dagelijks)\b', 'Daily'),
    (r'\b(weekly|wekelijks)\b', 'Weekly'),
    (r'\b(bi-weekly|tweewekelijks)\b', 'Bi-weekly'),
    (r'\b(monthly|maandelijks)\b', 'Monthly'),
    (r'\b(quarterly|driemaandelijks|kwartaal)\b', 'Quarterly'),
    (r'\b(semi-annual|halfjaarlijks)\b', 'Semi-annual'),
    (r'\b(annual|yearly|jaarlijks)\b', 'Annual'),
    (r'\bevery\s+(\d+)\s*hours?\b', 'Every {0} hours'),
    (r'\bevery\s+(\d+)\s*days?\b', 'Every {0} days'),
    (r'\bevery\s+(\d+)\s*weeks?\b', 'Every {0} weeks'),
    (r'\bevery\s+(\d+)\s*months?\b', 'Every {0} months'),
    (r'\b(\d+)\s*hours?\b', 'Every {0} hours'),
    (r'\b(\d+)\s*h\b', 'Every {0} hours'),
]

# Estimated duration pattern matching (time required to execute the task)
# Captures "30 min", "2 hours", "1.5 hrs", "45 minutes", Dutch "minuten", "uur"
DURATION_PATTERNS = [
    (r'\b(\d+(?:[.,]\d+)?)\s*(?:hours?|hrs?|uur|uren|hr)\b', 'hours'),
    (r'\b(\d+)\s*(?:minutes?|mins?|min|minuten)\b', 'minutes'),
    (r'\b(\d+)\s*m\b(?!o)', 'minutes'),  # "30m" but not "month"
]


class PMImportService:
    """Service class for PM Import operations."""
    
    def __init__(self, db: AsyncIOMotorDatabase):
        self.db = db
        self.sessions_collection = db["pm_import_sessions"]
        self.failure_modes_collection = db["failure_modes"]
    
    async def create_session_placeholder(
        self,
        file_name: str,
        file_type: str,
        created_by: str
    ) -> str:
        """Create a session placeholder for background processing. Returns session_id."""
        
        session_id = str(uuid.uuid4())
        now = datetime.now(timezone.utc)
        
        session = {
            "session_id": session_id,
            "file_name": file_name,
            "file_type": file_type,
            "status": "processing",
            "progress": 0,
            "progress_message": "Initializing...",
            "tasks_extracted": [],
            "stats": {
                "total_tasks": 0,
                "failure_modes_identified": 0,
                "existing_matches": 0,
                "new_proposed": 0,
                "low_confidence_items": 0,
                "manual_review_required": 0
            },
            "created_by": created_by,
            "created_at": now,
            "updated_at": now
        }
        
        await self.sessions_collection.insert_one(session)
        return session_id
    
    async def process_session(
        self,
        session_id: str,
        file_name: str,
        file_type: str,
        file_content: bytes
    ) -> None:
        """Process a PM import session (called as background task)."""
        
        try:
            tasks = await self._process_file(
                session_id, file_name, file_type, file_content
            )
            
            # Update session with results
            stats = self._calculate_stats(tasks)
            
            await self.sessions_collection.update_one(
                {"session_id": session_id},
                {"$set": {
                    "status": "ready_for_review",
                    "progress": 100,
                    "progress_message": "Processing complete",
                    "tasks_extracted": tasks,
                    "stats": stats,
                    "updated_at": datetime.now(timezone.utc)
                }}
            )
            logger.info(f"PM Import session {session_id} completed with {len(tasks)} tasks")
            
        except Exception as e:
            logger.error(f"Error processing session {session_id}: {e}", exc_info=True)
            await self.sessions_collection.update_one(
                {"session_id": session_id},
                {"$set": {
                    "status": "error",
                    "error_message": str(e),
                    "progress_message": f"Error: {str(e)[:200]}",
                    "updated_at": datetime.now(timezone.utc)
                }}
            )
    
    async def create_session(
        self,
        file_name: str,
        file_type: str,
        file_content: bytes,
        created_by: str
    ) -> Dict[str, Any]:
        """Create a new PM import session and start processing (legacy sync method)."""
        
        session_id = str(uuid.uuid4())
        now = datetime.now(timezone.utc)
        
        session = {
            "session_id": session_id,
            "file_name": file_name,
            "file_type": file_type,
            "status": "processing",
            "progress": 0,
            "progress_message": "Initializing...",
            "tasks_extracted": [],
            "stats": {
                "total_tasks": 0,
                "failure_modes_identified": 0,
                "existing_matches": 0,
                "new_proposed": 0,
                "low_confidence_items": 0,
                "manual_review_required": 0
            },
            "created_by": created_by,
            "created_at": now,
            "updated_at": now
        }
        
        await self.sessions_collection.insert_one(session)
        
        # Process the file
        try:
            tasks = await self._process_file(
                session_id, file_name, file_type, file_content
            )
            
            # Update session with results
            stats = self._calculate_stats(tasks)
            
            await self.sessions_collection.update_one(
                {"session_id": session_id},
                {"$set": {
                    "status": "ready_for_review",
                    "progress": 100,
                    "progress_message": "Processing complete",
                    "tasks_extracted": tasks,
                    "stats": stats,
                    "updated_at": datetime.now(timezone.utc)
                }}
            )
            
            session["status"] = "ready_for_review"
            session["progress"] = 100
            session["tasks_extracted"] = tasks
            session["stats"] = stats
            
        except Exception as e:
            logger.error(f"Error processing file: {e}", exc_info=True)
            await self.sessions_collection.update_one(
                {"session_id": session_id},
                {"$set": {
                    "status": "error",
                    "error_message": str(e),
                    "progress_message": f"Error: {str(e)[:200]}",
                    "updated_at": datetime.now(timezone.utc)
                }}
            )
            session["status"] = "error"
            session["error_message"] = str(e)
        
        return session
    
    async def get_session(self, session_id: str) -> Optional[Dict[str, Any]]:
        """Get a PM import session by ID."""
        session = await self.sessions_collection.find_one({"session_id": session_id})
        if session:
            session["_id"] = str(session["_id"])
        return session

    async def ensure_equipment_impacts(
        self,
        session_id: str,
        current_user: dict,
    ) -> Optional[Dict[str, Any]]:
        """
        Enrich tasks with equipment/tag impact information.

        Runs the same matchers used during upload (`_match_equipment_to_hierarchy`,
        `_match_equipment_types`) so the data shown in the review screen is consistent
        with the data the new PM Import table reads.

        Adds session fields:
        - equipment_impact_summary
        - equipment_impact_updated_at
        """
        session = await self.sessions_collection.find_one({"session_id": session_id})
        if not session:
            return None

        tasks = session.get("tasks_extracted") or []
        if not tasks:
            session["_id"] = str(session["_id"])
            return session

        # Only recompute if we haven't done it yet, or if tasks changed since last compute.
        prev = session.get("equipment_impact_summary") or {}
        if prev.get("tasks_count") == len(tasks) and session.get("equipment_impact_updated_at"):
            session["_id"] = str(session["_id"])
            return session

        # Use the SAME matchers as the upload pipeline so review and listing stay in sync.
        tasks = await self._match_equipment_to_hierarchy(tasks)

        matched_task_count = 0
        for task in tasks:
            if task.get("equipment_match"):
                matched_task_count += 1

        summary = {
            "tasks_count": len(tasks),
            "tasks_with_matches": matched_task_count,
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }

        await self.sessions_collection.update_one(
            {"session_id": session_id},
            {"$set": {
                "tasks_extracted": tasks,
                "equipment_impact_summary": summary,
                "equipment_impact_updated_at": datetime.now(timezone.utc),
                "updated_at": datetime.now(timezone.utc),
            }},
        )

        session["tasks_extracted"] = tasks
        session["equipment_impact_summary"] = summary
        session["equipment_impact_updated_at"] = datetime.now(timezone.utc)
        session["_id"] = str(session["_id"])
        return session

    async def _load_accessible_equipment_nodes(self, current_user: dict) -> List[Dict[str, Any]]:
        """
        Load equipment nodes the user can access (based on installation assignments).
        """
        try:
            from services.installation_filter_service import installation_filter

            installation_ids = await installation_filter.get_user_installation_ids(current_user)
            if not installation_ids:
                return []
            equipment_ids = await installation_filter.get_all_equipment_ids_for_installations(
                installation_ids, current_user.get("id")
            )
            if not equipment_ids:
                return []
            return await self.db.equipment_nodes.find(
                {"id": {"$in": list(equipment_ids)}},
                {"_id": 0, "id": 1, "tag": 1, "name": 1, "level": 1},
            ).to_list(20000)
        except Exception as e:
            logger.error(f"Failed to load accessible equipment nodes for PM impact: {e}")
            return []

    def _extract_tag_refs(self, text: str) -> List[str]:
        if not text:
            return []
        refs = []
        for m in TAG_REGEX.finditer(text.upper()):
            val = m.group(0).replace(" ", "")
            if val and val not in refs:
                refs.append(val)
        return refs

    async def _compute_import_impact_preview(self, task: Dict[str, Any], fm_service) -> Dict[str, Any]:
        """
        Compute a best-effort preview of what the PM import would change for this task.

        Returns:
          {
            "action": "link_existing" | "create_new" | "skip" | "rejected",
            "target_failure_mode": {id, failure_mode, equipment, category} | null,
            "changes": [{field, from, to, note}]  // human readable change intents
            "action_preview": {...} | null         // recommended action object we would add
            "reason": string | null
          }
        """
        status = task.get("review_status")
        if status == "rejected":
            return {"action": "rejected", "target_failure_mode": None, "changes": [], "action_preview": None, "reason": "Rejected"}

        if status not in ["accepted", "edited"]:
            return {"action": "skip", "target_failure_mode": None, "changes": [], "action_preview": None, "reason": "Not accepted"}

        library_match = task.get("library_match") or {}
        match_status = library_match.get("status", "")

        # Determine scenario target id
        target_id = None
        scenario = None
        if task.get("selected_match_id"):
            target_id = task.get("selected_match_id")
            scenario = "selected_match"
        elif match_status == "existing_match" and library_match.get("matched_id"):
            target_id = library_match.get("matched_id")
            scenario = "existing_match"

        action_type = task.get("action_type") or "PM"
        discipline = task.get("discipline") or "Maintenance"
        estimated_time = task.get("estimated_time") or ""
        desc = task.get("existing_control") or task.get("original_task") or ""
        freq = task.get("frequency")

        if target_id:
            existing_fm = await fm_service.get_by_id(target_id)
            if not existing_fm:
                return {"action": "skip", "target_failure_mode": None, "changes": [], "action_preview": None, "reason": "Matched failure mode not found"}

            actions = existing_fm.get("recommended_actions", []) or []
            action_exists = any(
                isinstance(a, dict) and (a.get("description") or "").lower() == desc.lower()
                for a in actions
            )

            new_action = {
                "description": desc,
                "action_type": action_type,
                "discipline": discipline,
                "source": "PM Import",
                "frequency": freq,
                "estimated_time": estimated_time,
            }

            changes = []
            if not action_exists:
                changes.append({"field": "recommended_actions", "from": "unchanged", "to": "add 1 action", "note": desc[:120]})
            else:
                changes.append({"field": "recommended_actions", "from": "already contains action", "to": "unchanged", "note": desc[:120]})

            fm_type = existing_fm.get("failure_mode_type")
            if fm_type != "customer_specific":
                changes.append({"field": "failure_mode_type", "from": fm_type or "", "to": "customer_specific", "note": "Marked as customized by PM import"})

            return {
                "action": "link_existing",
                "target_failure_mode": {
                    "id": str(existing_fm.get("id") or target_id),
                    "failure_mode": existing_fm.get("failure_mode", ""),
                    "equipment": existing_fm.get("equipment", ""),
                    "category": existing_fm.get("category", ""),
                },
                "changes": changes,
                "action_preview": new_action,
                "reason": None if scenario else None,
            }

        # Scenario C: approved new FM
        if task.get("approved_new_fm"):
            approved = task.get("approved_new_fm") or {}
            fm_name = approved.get("failure_mode") or approved.get("name") or ""
            if not fm_name:
                return {"action": "skip", "target_failure_mode": None, "changes": [], "action_preview": None, "reason": "No failure mode name approved"}

            # If name already exists, we will link instead of create
            existing = await fm_service.find_by_name(fm_name)
            if existing:
                # best-effort: existing may have _id not id
                existing_id = str(existing.get("id") or existing.get("_id"))
                # Recursively treat as link
                task_copy = dict(task)
                task_copy["selected_match_id"] = existing_id
                return await self._compute_import_impact_preview(task_copy, fm_service)

            equipment = approved.get("equipment") or task.get("component") or "General Equipment"
            category = approved.get("category") or self._determine_category(task)
            new_action = {
                "description": desc,
                "action_type": action_type,
                "discipline": discipline,
                "source": "PM Import",
                "frequency": freq,
                "estimated_time": estimated_time,
            }

            return {
                "action": "create_new",
                "target_failure_mode": {
                    "id": None,
                    "failure_mode": fm_name,
                    "equipment": equipment,
                    "category": category,
                },
                "changes": [
                    {"field": "failure_mode", "from": "", "to": fm_name, "note": "New failure mode created"},
                    {"field": "recommended_actions", "from": "", "to": "1 action", "note": desc[:120]},
                    {"field": "failure_mode_type", "from": "", "to": "customer_specific", "note": "Created as customer-specific"},
                ],
                "action_preview": new_action,
                "reason": None,
            }

        # Otherwise skip reasons based on match status
        reason = "No failure mode mapping confirmed"
        if match_status == "multiple_possible":
            reason = "Multiple matches — no selection"
        elif match_status == "new_proposed":
            reason = "New failure mode proposed — not approved"
        elif match_status == "weak_match":
            reason = "Weak match — not confirmed"

        return {"action": "skip", "target_failure_mode": None, "changes": [], "action_preview": None, "reason": reason}
    
    async def update_task(
        self,
        session_id: str,
        task_id: str,
        updates: Dict[str, Any]
    ) -> Optional[Dict[str, Any]]:
        """Update a specific task in a session."""
        
        session = await self.sessions_collection.find_one({"session_id": session_id})
        if not session:
            return None
        
        tasks = session.get("tasks_extracted", [])
        for task in tasks:
            if task.get("task_id") == task_id:
                task.update(updates)
                # Only set to "edited" if review_status is not explicitly being updated
                if "review_status" not in updates:
                    task["review_status"] = "edited"
                break
        
        # Recalculate stats
        stats = self._calculate_stats(tasks)
        
        await self.sessions_collection.update_one(
            {"session_id": session_id},
            {"$set": {
                "tasks_extracted": tasks,
                "stats": stats,
                "updated_at": datetime.now(timezone.utc)
            }}
        )
        
        return {"tasks": tasks, "stats": stats}
    
    async def accept_task(self, session_id: str, task_id: str) -> Optional[Dict[str, Any]]:
        """Mark a task as accepted for import."""
        return await self.update_task(session_id, task_id, {"review_status": "accepted"})
    
    async def reject_task(self, session_id: str, task_id: str) -> Optional[Dict[str, Any]]:
        """Mark a task as rejected (won't be imported)."""
        return await self.update_task(session_id, task_id, {"review_status": "rejected"})
    
    async def delete_task(self, session_id: str, task_id: str) -> Optional[Dict[str, Any]]:
        """Remove a task from a session entirely."""
        session = await self.sessions_collection.find_one({"session_id": session_id})
        if not session:
            return None
        
        tasks = session.get("tasks_extracted", [])
        new_tasks = [t for t in tasks if t.get("task_id") != task_id]
        if len(new_tasks) == len(tasks):
            return None
        
        stats = self._calculate_stats(new_tasks)
        
        await self.sessions_collection.update_one(
            {"session_id": session_id},
            {"$set": {
                "tasks_extracted": new_tasks,
                "stats": stats,
                "updated_at": datetime.now(timezone.utc)
            }}
        )
        
        return {"tasks": new_tasks, "stats": stats}
    
    async def accept_all_high_confidence(self, session_id: str) -> Optional[Dict[str, Any]]:
        """Accept all tasks with confidence >= 70."""
        session = await self.sessions_collection.find_one({"session_id": session_id})
        if not session:
            return None
        
        tasks = session.get("tasks_extracted", [])
        accepted_count = 0
        
        for task in tasks:
            if task.get("confidence_score", 0) >= 70 and task.get("review_status") == "pending":
                task["review_status"] = "accepted"
                accepted_count += 1
        
        stats = self._calculate_stats(tasks)
        
        await self.sessions_collection.update_one(
            {"session_id": session_id},
            {"$set": {
                "tasks_extracted": tasks,
                "stats": stats,
                "updated_at": datetime.now(timezone.utc)
            }}
        )
        
        return {"accepted_count": accepted_count, "tasks": tasks, "stats": stats}
    
    async def import_to_library(
        self,
        session_id: str,
        created_by: str
    ) -> Dict[str, Any]:
        """Import accepted tasks to the Failure Mode Library.
        
        Three scenarios:
        A) existing_match: Auto-link as preventive control to existing failure mode
        B) selected_match_id set: User selected from multiple matches - link to selected
        C) approved_new_fm set: User approved new failure mode - create it
        
        Tasks without proper mapping (no existing match, no user selection, no approval) are skipped.
        """
        
        session = await self.sessions_collection.find_one({"session_id": session_id})
        if not session:
            raise ValueError("Session not found")
        
        tasks = session.get("tasks_extracted", [])
        now = datetime.now(timezone.utc)
        
        imported_count = 0
        linked_count = 0
        new_count = 0
        skipped_count = 0
        low_confidence_imported = 0
        
        # Track detailed results for UI display
        linked_details = []  # Tasks linked to existing failure modes
        created_details = []  # New failure modes created
        skipped_details = []  # Skipped tasks
        
        from services.failure_modes_service import FailureModesService
        fm_service = FailureModesService(self.db)
        
        for task in tasks:
            # Skip rejected tasks
            if task.get("review_status") == "rejected":
                skipped_count += 1
                skipped_details.append({
                    "task": task.get("original_task", "")[:100],
                    "reason": "Rejected by user"
                })
                continue
            
            # Skip tasks not accepted/edited
            if task.get("review_status") not in ["accepted", "edited"]:
                continue
            
            confidence = task.get("confidence_score", 0)
            if confidence < 70:
                low_confidence_imported += 1
            
            library_match = task.get("library_match", {})
            match_status = library_match.get("status", "")
            
            # SCENARIO B (priority): User selected/overrode a match — always wins
            if task.get("selected_match_id"):
                matched_id = task["selected_match_id"]
                result = await self._link_task_to_failure_mode(
                    task, matched_id, fm_service, session.get("file_name"), now
                )
                if result:
                    linked_count += 1
                    imported_count += 1
                    linked_details.append(result)
                continue
            
            # SCENARIO A: Existing high-confidence match - auto link
            if match_status == "existing_match" and library_match.get("matched_id"):
                matched_id = library_match["matched_id"]
                result = await self._link_task_to_failure_mode(
                    task, matched_id, fm_service, session.get("file_name"), now
                )
                if result:
                    linked_count += 1
                    imported_count += 1
                    linked_details.append(result)
                continue
            
            # SCENARIO C: User approved new failure mode creation
            if task.get("approved_new_fm"):
                approved_fm = task["approved_new_fm"]
                fm_name = approved_fm.get("failure_mode") or approved_fm.get("name")
                
                if not fm_name:
                    skipped_count += 1
                    skipped_details.append({
                        "task": task.get("original_task", "")[:100],
                        "reason": "No failure mode name in approval"
                    })
                    continue
                
                # Check if already exists
                existing = await fm_service.find_by_name(fm_name)
                if existing:
                    # Link to existing instead
                    result = await self._link_task_to_failure_mode(
                        task, str(existing.get("_id")), fm_service, session.get("file_name"), now
                    )
                    if result:
                        linked_count += 1
                        imported_count += 1
                        linked_details.append(result)
                    continue
                
                # Create the approved new failure mode
                action_type = task.get("action_type") or "PM"
                discipline = task.get("discipline") or "Maintenance"
                estimated_time = task.get("estimated_time") or ""
                
                new_fm_data = {
                    "category": approved_fm.get("category") or self._determine_category(task),
                    "equipment": approved_fm.get("equipment") or task.get("component") or "General Equipment",
                    "failure_mode": fm_name,
                    "keywords": self._extract_keywords(task),
                    "severity": approved_fm.get("severity", 5),
                    "occurrence": approved_fm.get("occurrence", 5),
                    "detectability": approved_fm.get("detectability", 5),
                    "recommended_actions": [{
                        "description": task.get("existing_control") or task.get("original_task"),
                        "action_type": action_type,
                        "discipline": discipline,
                        "source": "PM Import",
                        "frequency": task.get("frequency"),
                        "estimated_time": estimated_time,
                        "imported_from": session.get("file_name"),
                        "imported_at": now.isoformat()
                    }],
                    "failure_mode_type": "customer_specific",
                    "source": "pm_import",
                    "potential_causes": task.get("failure_mechanisms", []),
                    "process": task.get("task_type")
                }
                
                try:
                    created_fm = await fm_service.create(new_fm_data, created_by=created_by)
                    new_count += 1
                    imported_count += 1
                    created_details.append({
                        "task": task.get("original_task", "")[:100],
                        "task_type": task.get("task_type", ""),
                        "action_type": action_type,
                        "discipline": discipline,
                        "estimated_time": estimated_time,
                        "component": task.get("component", ""),
                        "frequency": task.get("frequency", ""),
                        "failure_modes_created": [{
                            "failure_mode_id": created_fm.get("id"),
                            "failure_mode_name": fm_name,
                            "equipment": new_fm_data["equipment"],
                            "category": new_fm_data["category"]
                        }]
                    })
                except Exception as e:
                    logger.error(f"Failed to create failure mode: {e}")
                    skipped_details.append({
                        "task": task.get("original_task", "")[:100],
                        "reason": f"Failed to create: {str(e)[:50]}"
                    })
                continue
            
            # No valid mapping - skip the task
            skipped_count += 1
            reason = "No failure mode mapping confirmed"
            if match_status == "multiple_possible":
                reason = "Multiple matches - user did not select one"
            elif match_status == "new_proposed":
                reason = "New failure mode proposed - user did not approve"
            elif match_status == "weak_match":
                reason = "Weak match - user did not confirm"
            
            skipped_details.append({
                "task": task.get("original_task", "")[:100],
                "reason": reason
            })
        
        # Build import result with details
        import_result = {
            "total_imported": imported_count,
            "linked_to_existing": linked_count,
            "new_created": new_count,
            "skipped": skipped_count,
            "low_confidence_imported": low_confidence_imported,
            "imported_at": now.isoformat(),
            "imported_by": created_by,
            "linked_details": linked_details,
            "created_details": created_details,
            "skipped_details": skipped_details
        }
        
        # Update session status
        await self.sessions_collection.update_one(
            {"session_id": session_id},
            {"$set": {
                "status": "imported",
                "import_result": import_result,
                "updated_at": now
            }}
        )
        
        return {
            "success": True,
            "total_imported": imported_count,
            "linked_to_existing": linked_count,
            "new_created": new_count,
            "skipped": skipped_count,
            "low_confidence_imported": low_confidence_imported,
            "linked_details": linked_details,
            "created_details": created_details,
            "skipped_details": skipped_details
        }
    
    async def _link_task_to_failure_mode(
        self,
        task: Dict[str, Any],
        failure_mode_id: str,
        fm_service,
        file_name: str,
        now
    ) -> Optional[Dict[str, Any]]:
        """Link a PM task to an existing failure mode as a preventive control.
        
        Updates the failure mode with:
        - The PM/CM/PDM action (with discipline)
        - failure_mode_type → "customer_specific" (customized with user's PM)
        """
        
        existing_fm = await fm_service.get_by_id(failure_mode_id)
        if not existing_fm:
            return None
        
        action_type = task.get("action_type") or "PM"
        discipline = task.get("discipline") or "Maintenance"
        estimated_time = task.get("estimated_time") or ""
        
        # Add the PM task as a recommended action if not already present
        actions = existing_fm.get("recommended_actions", [])
        new_action = {
            "description": task.get("existing_control") or task.get("original_task"),
            "action_type": action_type,
            "discipline": discipline,
            "source": "PM Import",
            "frequency": task.get("frequency"),
            "estimated_time": estimated_time,
            "imported_from": file_name,
            "imported_at": now.isoformat()
        }
        
        # Check if similar action exists
        action_exists = any(
            a.get("description", "").lower() == new_action["description"].lower()
            for a in actions if isinstance(a, dict)
        )
        
        if not action_exists:
            actions.append(new_action)
        
        # Mark failure mode as customer_specific since user has customized it with their PM
        update_payload = {
            "recommended_actions": actions,
            "failure_mode_type": "customer_specific",
        }
        await fm_service.update(
            failure_mode_id,
            update_payload,
            updated_by="PM Import",
            change_reason=f"PM Import: linked task '{(task.get('original_task') or '')[:80]}'"
        )
        
        return {
            "task": task.get("original_task", "")[:100],
            "task_type": task.get("task_type", ""),
            "action_type": action_type,
            "discipline": discipline,
            "estimated_time": estimated_time,
            "component": task.get("component", ""),
            "frequency": task.get("frequency", ""),
            "failure_mode_id": failure_mode_id,
            "failure_mode_name": existing_fm.get("failure_mode", ""),
            "equipment": existing_fm.get("equipment", ""),
            "category": existing_fm.get("category", ""),
            "action_added": new_action["description"][:80] if not action_exists else None,
            "already_existed": action_exists,
            "marked_customer_specific": existing_fm.get("failure_mode_type") != "customer_specific",
        }
    
    # ==================== FILE PROCESSING ====================
    
    async def _process_file(
        self,
        session_id: str,
        file_name: str,
        file_type: str,
        file_content: bytes
    ) -> List[Dict[str, Any]]:
        """Process uploaded file and extract maintenance tasks."""
        
        await self._update_progress(session_id, 10, "Reading maintenance plan...")
        
        # Extract raw text/rows based on file type
        if file_type in ["xlsx", "xls"]:
            raw_rows = await self._parse_excel(file_content)
        elif file_type == "pdf":
            raw_rows = await self._parse_pdf(file_content, session_id)
        elif file_type in ["png", "jpg", "jpeg", "webp"]:
            raw_rows = await self._parse_image(file_content, file_type, session_id)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")
        
        if not raw_rows:
            raise ValueError("No maintenance tasks could be extracted from the file")
        
        await self._update_progress(session_id, 40, f"Extracted {len(raw_rows)} tasks. Analyzing...")
        
        # Process each row with AI
        tasks = []
        total = len(raw_rows)
        
        for idx, row in enumerate(raw_rows):
            progress = 40 + int((idx / total) * 50)
            await self._update_progress(
                session_id, 
                progress, 
                f"Processing task {idx + 1} of {total}..."
            )
            
            task = await self._analyze_task(row, session_id, file_name)
            if task:
                tasks.append(task)
        
        await self._update_progress(session_id, 90, "AI enrichment (translate, classify, hours)...")
        
        # AI enrichment: translate to English, classify PM/PDM/CBM/CM, suggest discipline,
        # standardize frequency (+ frequency_days), estimate hours.
        tasks = await self._ai_enrich_tasks(tasks)
        
        # Split tasks whose Tag column contains multiple tags into one task per tag.
        # E.g. asset="P-1001, P-1002, P-1003" → three tasks with identical content
        # but distinct equipment_tag values.
        await self._update_progress(session_id, 95, "Splitting multi-tag tasks...")
        tasks = self._split_multi_tag_tasks(tasks)
        
        # Match equipment to hierarchy (Priority 1: tag exact, Priority 2: description fuzzy)
        await self._update_progress(session_id, 99, "Matching equipment to hierarchy...")
        tasks = await self._match_equipment_to_hierarchy(tasks)
        
        # Normalize final shape per the PM Import refactor spec
        tasks = [self._normalize_task_shape(t) for t in tasks]
        
        return tasks
    
    async def _parse_excel(self, content: bytes) -> List[Dict[str, Any]]:
        """
        Parse Excel file per the PM Import Extraction Engine spec.
        
        Rules:
        - Column A is the ONLY authoritative source for equipment tags.
        - Merged cells are resolved (master value applied to every cell in the merge).
        - Worksheet is analyzed as a complete table, not row-by-row.
        - Blank Column A rows are treated as continuations of the previous active task block.
        - When N tags share the same task block (merged or grouped), one record per tag is emitted.
        """
        import openpyxl
        
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        rows = []
        
        # Keywords used to auto-detect non-tag columns by header name. Column A is ALWAYS the tag column.
        task_column_keywords = ["task", "description", "activity", "action", "work", "maintenance",
                                "taak", "beschrijving", "activiteit", "werkzaamheden", "instructie"]
        description_column_keywords = ["equipment", "component", "asset", "machine", "name",
                                       "apparaat", "onderdeel", "omschrijving", "naam"]
        frequency_column_keywords = ["frequency", "interval", "schedule", "frequentie", "periode",
                                     "freq", "cycle"]
        duration_column_keywords = ["duration", "estimated time", "estimated_time", "time",
                                    "duur", "tijd", "minutes", "hours", "minuten", "uren"]
        
        for sheet in wb.worksheets:
            # Build merged-cell lookup: for each (row, col) inside a merge, store the master value
            merged_master = {}
            for merged_range in sheet.merged_cells.ranges:
                min_row, min_col = merged_range.min_row, merged_range.min_col
                master_value = sheet.cell(row=min_row, column=min_col).value
                for r in range(merged_range.min_row, merged_range.max_row + 1):
                    for c in range(merged_range.min_col, merged_range.max_col + 1):
                        merged_master[(r, c)] = master_value
            
            def cell_value(row_idx: int, col_idx: int):
                """Resolve a cell, returning the merged-master value if applicable."""
                key = (row_idx, col_idx + 1)  # openpyxl is 1-indexed
                if key in merged_master:
                    return merged_master[key]
                # Direct value (col_idx is 0-based input → openpyxl is 1-based)
                return sheet.cell(row=row_idx, column=col_idx + 1).value
            
            max_row = sheet.max_row or 0
            max_col = sheet.max_column or 0
            if max_row < 2 or max_col < 1:
                continue
            
            # Detect header row (first 10 rows): the row where Column A AND any other column have header-like text
            header_row_idx = None
            headers = []
            task_col_idx = None
            description_col_idx = None
            frequency_col_idx = None
            duration_col_idx = None
            
            for r in range(1, min(11, max_row + 1)):
                row_values = [cell_value(r, c) for c in range(max_col)]
                if not any(v for v in row_values):
                    continue
                row_text_lower = " ".join(str(v).lower() for v in row_values if v)
                if any(kw in row_text_lower for kw in
                       ["task", "description", "tag", "equipment", "frequency", "frequentie",
                        "taak", "apparaat", "onderdeel", "naam", "instructie"]):
                    header_row_idx = r
                    headers = [str(v).strip() if v else f"col_{i}" for i, v in enumerate(row_values)]
                    for i, h in enumerate(headers):
                        h_lower = h.lower()
                        if i == 0:
                            continue  # Column A is always the tag column
                        if task_col_idx is None and any(kw in h_lower for kw in task_column_keywords):
                            task_col_idx = i
                        elif description_col_idx is None and any(kw in h_lower for kw in description_column_keywords):
                            description_col_idx = i
                        elif frequency_col_idx is None and any(kw in h_lower for kw in frequency_column_keywords):
                            frequency_col_idx = i
                        elif duration_col_idx is None and any(kw in h_lower for kw in duration_column_keywords):
                            duration_col_idx = i
                    break
            
            if header_row_idx is None:
                header_row_idx = 1
                first_row = [cell_value(1, c) for c in range(max_col)]
                headers = [str(v).strip() if v else f"col_{i}" for i, v in enumerate(first_row)]
            
            # Fallback: if no task column was detected, pick column B (index 1) as task description
            if task_col_idx is None and max_col >= 2:
                task_col_idx = 1
            
            # ------ Walk the data rows building one record per non-empty Column A ------
            current_record = None
            for r in range(header_row_idx + 1, max_row + 1):
                tag_val = cell_value(r, 0)  # Column A
                tag_str = str(tag_val).strip() if tag_val is not None else ""
                
                # Skip totally empty rows
                row_values = [cell_value(r, c) for c in range(max_col)]
                if not any(v for v in row_values) and not tag_str:
                    continue
                
                if tag_str:
                    # New equipment tag → start a new record
                    task_text = ""
                    if task_col_idx is not None and task_col_idx < max_col:
                        tv = cell_value(r, task_col_idx)
                        if tv:
                            task_text = str(tv).strip()
                    
                    description_text = ""
                    if description_col_idx is not None and description_col_idx < max_col:
                        dv = cell_value(r, description_col_idx)
                        if dv:
                            description_text = str(dv).strip()
                    
                    freq_text = ""
                    if frequency_col_idx is not None and frequency_col_idx < max_col:
                        fv = cell_value(r, frequency_col_idx)
                        if fv:
                            freq_text = str(fv).strip()
                    
                    duration_text = ""
                    if duration_col_idx is not None and duration_col_idx < max_col:
                        dv = cell_value(r, duration_col_idx)
                        if dv:
                            duration_text = str(dv).strip()
                    
                    # If task description is still empty, pick the longest non-tag cell on this row
                    if not task_text:
                        candidates = [(c, cell_value(r, c)) for c in range(max_col) if c != 0]
                        candidates = [(c, str(v).strip()) for c, v in candidates if v]
                        if candidates:
                            task_text = max(candidates, key=lambda x: len(x[1]))[1]
                    
                    current_record = {
                        "_tag": tag_str,           # Column A — authoritative
                        "_task_text": task_text,
                        "_description": description_text,
                        "_frequency": freq_text,
                        "_duration": duration_text,
                        "_raw_text": task_text,
                        "_sheet": sheet.title,
                        "_row": r,
                    }
                    # Attach all column data for reference
                    for i, v in enumerate(row_values):
                        if v is not None:
                            header = headers[i] if i < len(headers) else f"col_{i}"
                            current_record[header] = str(v).strip()
                    
                    rows.append(current_record)
                else:
                    # Column A is empty → continuation of the active task block.
                    # Append any non-empty cell content to the current task description.
                    if current_record is None:
                        continue
                    continuation_parts = []
                    for c in range(max_col):
                        v = cell_value(r, c)
                        if v:
                            continuation_parts.append(str(v).strip())
                    if continuation_parts:
                        extra = " | ".join(continuation_parts)
                        current_record["_task_text"] = (
                            (current_record.get("_task_text") or "") + " — " + extra
                        ).strip(" —")
                        current_record["_raw_text"] = current_record["_task_text"]
        
        return rows
    
    async def _parse_pdf(self, content: bytes, session_id: str) -> List[Dict[str, Any]]:
        """Parse PDF file - use text extraction first, fall back to vision for scanned PDFs."""
        import pdfplumber
        
        rows = []
        has_text = False
        
        try:
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                for page in pdf.pages:
                    # Try to extract tables first
                    tables = page.extract_tables()
                    if tables:
                        for table in tables:
                            for row in table:
                                if row and any(cell for cell in row if cell):
                                    # Get the first non-empty cell with substantial text as the task
                                    task_text = ""
                                    for cell in row:
                                        if cell and len(str(cell).strip()) > len(task_text):
                                            task_text = str(cell).strip()
                                    if task_text and len(task_text) > 5:
                                        rows.append({"_raw_text": task_text, "_task_text": task_text})
                                        has_text = True
                    
                    # If no tables, extract text
                    if not tables:
                        text = page.extract_text()
                        if text and text.strip():
                            has_text = True
                            # Split into lines as potential tasks
                            for line in text.split("\n"):
                                line = line.strip()
                                if len(line) > 10:  # Skip very short lines
                                    rows.append({"_raw_text": line, "_task_text": line})
        except Exception as e:
            logger.warning(f"PDF text extraction failed: {e}")
        
        # If no text found (scanned PDF), use GPT-4o Vision
        if not has_text or len(rows) < 3:
            logger.info("Using GPT-4o Vision for scanned PDF")
            await self._update_progress(session_id, 20, "Scanned PDF detected. Using AI vision...")
            rows = await self._ocr_with_vision(content, "pdf", session_id)
        
        return rows
    
    async def _parse_image(
        self,
        content: bytes,
        file_type: str,
        session_id: str
    ) -> List[Dict[str, Any]]:
        """Parse image using GPT-4o Vision."""
        await self._update_progress(session_id, 20, "Analyzing image with AI vision...")
        return await self._ocr_with_vision(content, file_type, session_id)
    
    async def _ocr_with_vision(
        self,
        content: bytes,
        file_type: str,
        session_id: str
    ) -> List[Dict[str, Any]]:
        """Use GPT-4o Vision to extract maintenance tasks from images/scanned documents."""
        from openai import OpenAI
        
        api_key = os.environ.get("OPENAI_API_KEY")
        if not api_key:
            raise ValueError("OpenAI API key not configured")
        
        client = OpenAI(api_key=api_key)
        
        # Convert to base64
        if file_type == "pdf":
            # Convert PDF pages to images
            try:
                import fitz  # PyMuPDF
                pdf_doc = fitz.open(stream=content, filetype="pdf")
                images_b64 = []
                for page_num in range(min(pdf_doc.page_count, 5)):  # Max 5 pages
                    page = pdf_doc[page_num]
                    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x zoom for better OCR
                    img_bytes = pix.tobytes("png")
                    images_b64.append(base64.b64encode(img_bytes).decode('utf-8'))
                pdf_doc.close()
            except ImportError:
                # Fallback: just use first page as image
                from PIL import Image
                from pdf2image import convert_from_bytes
                pages = convert_from_bytes(content, first_page=1, last_page=5)
                images_b64 = []
                for page in pages:
                    buf = io.BytesIO()
                    page.save(buf, format='PNG')
                    images_b64.append(base64.b64encode(buf.getvalue()).decode('utf-8'))
        else:
            # Regular image
            images_b64 = [base64.b64encode(content).decode('utf-8')]
        
        # Process with GPT-4o Vision
        all_rows = []
        
        for idx, img_b64 in enumerate(images_b64):
            mime_type = "image/png" if file_type == "pdf" else f"image/{file_type}"
            
            prompt = """Analyze this maintenance plan document and extract all preventive maintenance tasks.

For each maintenance task found, extract:
1. The original task text exactly as written
2. The equipment or component mentioned
3. Any frequency information (daily, weekly, monthly, etc.)
4. Any additional details

Return the data as a JSON array where each item has:
{
  "task": "original task text",
  "equipment": "equipment/component name if mentioned",
  "frequency": "frequency if mentioned",
  "details": "any additional details"
}

If the document is in Dutch, still extract the tasks but keep them in Dutch.
Only return the JSON array, no other text."""

            try:
                response = client.chat.completions.create(
                    model="gpt-4o",
                    messages=[
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": prompt},
                                {
                                    "type": "image_url",
                                    "image_url": {
                                        "url": f"data:{mime_type};base64,{img_b64}",
                                        "detail": "high"
                                    }
                                }
                            ]
                        }
                    ],
                    max_tokens=4000,
                    temperature=0
                )
                
                result_text = response.choices[0].message.content.strip()
                
                # Parse JSON response
                import json
                # Clean up response
                if result_text.startswith("```"):
                    result_text = re.sub(r'^```(?:json)?\n?', '', result_text)
                    result_text = re.sub(r'\n?```$', '', result_text)
                
                tasks = json.loads(result_text)
                
                for task in tasks:
                    raw_text = task.get("task", "")
                    if task.get("equipment"):
                        raw_text += f" | Equipment: {task['equipment']}"
                    if task.get("frequency"):
                        raw_text += f" | Frequency: {task['frequency']}"
                    
                    all_rows.append({
                        "_raw_text": raw_text,
                        "_ocr_data": task
                    })
                    
            except Exception as e:
                logger.error(f"Vision API error on page {idx + 1}: {e}")
                continue
        
        return all_rows
    
    async def _analyze_task(
        self,
        row: Dict[str, Any],
        session_id: str,
        file_name: str
    ) -> Optional[Dict[str, Any]]:
        """Analyze a single maintenance task and extract failure mode intelligence."""
        
        # Use clean task text if available, otherwise use raw_text
        task_text = row.get("_task_text") or row.get("_raw_text", "")
        if not task_text or len(task_text) < 5:
            return None
        
        # Step 1: Rule-based pre-classification
        task_type, rule_based_data = self._classify_by_rules(task_text)
        
        # Step 2: Extract frequency (from parsed data or text)
        frequency = row.get("_frequency") or row.get("_ocr_data", {}).get("frequency", "")
        if not frequency:
            frequency = self._extract_frequency(task_text)
        
        # Step 2b: Extract estimated duration (task execution time)
        duration_field_text = ""
        # Check common duration column names from parsed data
        for key in ("duration", "estimated_time", "estimated time", "time", "estimated_duration", "tijd", "duur"):
            if row.get(key):
                duration_field_text = str(row.get(key))
                break
        ocr_duration = row.get("_ocr_data", {}).get("duration") or row.get("_ocr_data", {}).get("estimated_time")
        estimated_time = self._extract_duration(duration_field_text) \
            or (str(ocr_duration) if ocr_duration else "") \
            or self._extract_duration(task_text)
        
        # Step 3: Extract component (from parsed data, OCR data, or text)
        component = row.get("_equipment") or row.get("_ocr_data", {}).get("equipment", "")
        if not component:
            component = self._extract_component(task_text)
        
        # Step 4: AI enhancement
        ai_analysis = await self._ai_analyze_task(task_text, task_type, rule_based_data)
        
        # Build task object
        task_id = str(uuid.uuid4())
        
        # Calculate confidence score
        confidence = self._calculate_confidence(
            has_component=bool(component),
            has_frequency=bool(frequency),
            task_type_known=(task_type != "Unknown"),
            ai_confidence=ai_analysis.get("confidence", 70),
            failure_modes_count=len(ai_analysis.get("failure_modes", rule_based_data.get("failure_modes", [])))
        )
        
        return {
            "task_id": task_id,
            "original_task": task_text,  # Clean task text without delimiters
            "component": component or ai_analysis.get("component", ""),
            "asset": ai_analysis.get("asset", ""),
            "task_type": ai_analysis.get("task_type", task_type),
            "action_type": self._infer_action_type(
                task_text,
                ai_analysis.get("task_type", task_type),
                ai_analysis.get("action_type"),
            ),
            "discipline": self._infer_discipline(
                task_text,
                component or ai_analysis.get("component", ""),
                ai_analysis.get("task_type", task_type),
                ai_analysis.get("discipline"),
            ),
            "suggested_failure_modes": ai_analysis.get("failure_modes", rule_based_data.get("failure_modes", [])),
            "failure_mechanisms": ai_analysis.get("mechanisms", rule_based_data.get("failure_mechanisms", [])),
            "detection_methods": ai_analysis.get("detection_methods", rule_based_data.get("detection_methods", [])),
            "existing_control": task_text,
            "frequency": frequency or ai_analysis.get("frequency", ""),
            "estimated_time": estimated_time or ai_analysis.get("estimated_time", ""),
            "confidence_score": confidence,
            "ai_reasoning": ai_analysis.get("reasoning", ""),
            "review_status": "pending",
            "source_document": file_name,
            "source_row": row
        }
    
    def _classify_by_rules(self, text: str) -> Tuple[str, Dict[str, Any]]:
        """Classify task type and extract data using keyword rules."""
        text_lower = text.lower()
        
        for task_type, rules in TASK_CLASSIFICATION_RULES.items():
            for keyword in rules["keywords"]:
                if keyword in text_lower:
                    return (
                        task_type.capitalize(),
                        {
                            "failure_modes": rules["failure_modes"][:3],
                            "failure_mechanisms": rules["failure_mechanisms"][:2],
                            "detection_methods": rules["detection_methods"][:2]
                        }
                    )
        
        return ("Unknown", {})
    
    def _infer_action_type(
        self,
        task_text: str,
        task_type: str,
        ai_hint: Optional[str] = None,
    ) -> str:
        """Infer maintenance action type: PM (preventive), PDM (predictive), CM (corrective)."""
        text_lower = (task_text or "").lower()
        
        # 1) Honor explicit AI hint if it's a valid action type
        if ai_hint and ai_hint.upper() in ACTION_TYPES:
            return ai_hint.upper()
        
        # 2) Keyword override (predictive/corrective signals beat the default)
        for action_type, keywords in ACTION_TYPE_KEYWORDS.items():
            if any(kw in text_lower for kw in keywords):
                return action_type
        
        # 3) Fall back to task-type default
        default = TASK_TYPE_DEFAULTS.get(task_type, TASK_TYPE_DEFAULTS["Unknown"])
        return default["action_type"]
    
    def _infer_discipline(
        self,
        task_text: str,
        component: str,
        task_type: str,
        ai_hint: Optional[str] = None,
    ) -> str:
        """Infer execution discipline from task text, component, and task type."""
        from models.disciplines import normalize_discipline, DISCIPLINE_LIST
        
        # 1) Honor explicit AI hint if it normalizes to a known discipline
        if ai_hint:
            normalized = normalize_discipline(ai_hint)
            if normalized in DISCIPLINE_LIST:
                return normalized
        
        haystack = f"{task_text or ''} {component or ''}".lower()
        
        # 2) Keyword scoring across haystack — pick the discipline with most matches
        scores = {discipline: 0 for discipline in DISCIPLINE_KEYWORDS}
        for discipline, keywords in DISCIPLINE_KEYWORDS.items():
            for kw in keywords:
                if kw in haystack:
                    scores[discipline] += 1
        
        best = max(scores.items(), key=lambda kv: kv[1])
        if best[1] > 0:
            return best[0]
        
        # 3) Fall back to task-type default
        default = TASK_TYPE_DEFAULTS.get(task_type, TASK_TYPE_DEFAULTS["Unknown"])
        return default["discipline"]
    
    def _extract_frequency(self, text: str) -> str:
        """Extract frequency information from text."""
        text_lower = text.lower()
        
        for pattern, template in FREQUENCY_PATTERNS:
            match = re.search(pattern, text_lower)
            if match:
                if '{0}' in template:
                    return template.format(match.group(1))
                return template
        
        return ""
    
    def _extract_duration(self, text: str) -> str:
        """Extract estimated task duration from text (e.g., '30 min', '2 hours').
        
        Returns a clean, normalized duration string or empty if none found.
        """
        text_lower = (text or "").lower()
        
        for pattern, unit in DURATION_PATTERNS:
            match = re.search(pattern, text_lower)
            if not match:
                continue
            value = match.group(1).replace(",", ".")
            try:
                num = float(value)
            except ValueError:
                continue
            # Filter out values that are obviously frequency intervals, not durations
            if unit == "hours" and num > 24:
                continue
            if unit == "minutes" and num > 480:  # >8h likely not a duration
                continue
            # Normalize formatting
            if num.is_integer():
                num_str = str(int(num))
            else:
                num_str = f"{num:g}"
            label = "hour" if (unit == "hours" and num == 1) else unit
            return f"{num_str} {label}"
        
        return ""
    
    def _extract_component(self, text: str) -> str:
        """Extract component/equipment name from text."""
        # Common equipment keywords
        equipment_patterns = [
            r'\b(pump|compressor|motor|valve|bearing|seal|gear|fan|blower|turbine|exchanger|filter|sensor|conveyor|roller|belt|chain|coupling|shaft)\b',
            r'\b(pomp|compressor|motor|klep|lager|afdichting|tandwiel|ventilator|turbine|warmtewisselaar|filter|sensor)\b',  # Dutch
        ]
        
        for pattern in equipment_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                # Try to get more context
                start = max(0, match.start() - 20)
                end = min(len(text), match.end() + 20)
                context = text[start:end]
                
                # Clean up
                words = context.split()
                for i, word in enumerate(words):
                    if match.group(1).lower() in word.lower():
                        # Get the word and maybe adjacent words
                        component_words = []
                        if i > 0 and words[i-1][0].isupper():
                            component_words.append(words[i-1])
                        component_words.append(word)
                        if i < len(words) - 1 and words[i+1][0].islower():
                            component_words.append(words[i+1])
                        return " ".join(component_words)
                
                return match.group(1).capitalize()
        
        return ""
    
    async def _ai_analyze_task(
        self,
        task_text: str,
        pre_classified_type: str,
        rule_based_data: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Use AI to enhance task analysis."""
        from openai import OpenAI
        
        api_key = os.environ.get("OPENAI_API_KEY")
        if not api_key:
            return {
                "task_type": pre_classified_type,
                "failure_modes": rule_based_data.get("failure_modes", []),
                "mechanisms": rule_based_data.get("failure_mechanisms", []),
                "detection_methods": rule_based_data.get("detection_methods", []),
                "confidence": 60,
                "reasoning": "AI analysis unavailable - using rule-based classification"
            }
        
        client = OpenAI(api_key=api_key)
        
        prompt = f"""Analyze this preventive maintenance task and extract reliability intelligence.

MAINTENANCE TASK: "{task_text}"

Pre-classified type: {pre_classified_type}
Pre-identified failure modes: {rule_based_data.get('failure_modes', [])}

Please analyze and provide:
1. Component: What equipment/component is this task for?
2. Task Type: One of [Inspection, Lubrication, Calibration, Replacement, Cleaning, Adjustment, Monitoring, Unknown]
3. Action Type: Maintenance strategy — one of:
   - "PM"  (Preventive — time/usage-based scheduled work)
   - "PDM" (Predictive — condition-based monitoring like vibration, oil, thermography)
   - "CM"  (Corrective — repair on failure)
4. Discipline: Execution discipline — one of [Mechanical, Electrical, Instrumentation, Process, Inspection, Operations, Maintenance, Reliability]
5. Failure Modes: What failures is this task preventing? (list 2-4 specific failure modes)
6. Mechanisms: What failure mechanisms are being addressed? (list 1-3)
7. Detection Methods: How would failures be detected? (list 1-3)
8. Frequency: If mentioned, what is the task frequency?
9. Estimated Time: If mentioned, estimated duration to execute this task (e.g., "30 min", "2 hours"). Empty if not stated.
10. Confidence: How confident are you in this analysis? (0-100)
11. Reasoning: Brief explanation of your analysis

Respond in JSON format:
{{
  "component": "string",
  "asset": "broader asset category if identifiable",
  "task_type": "string",
  "action_type": "PM | PDM | CM",
  "discipline": "string",
  "failure_modes": ["string", "string"],
  "mechanisms": ["string"],
  "detection_methods": ["string"],
  "frequency": "string or empty",
  "estimated_time": "string or empty",
  "confidence": number,
  "reasoning": "string"
}}"""

        try:
            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": prompt}],
                max_tokens=800,
                temperature=0.2
            )
            
            result_text = response.choices[0].message.content.strip()
            
            # Parse JSON
            import json
            if result_text.startswith("```"):
                result_text = re.sub(r'^```(?:json)?\n?', '', result_text)
                result_text = re.sub(r'\n?```$', '', result_text)
            
            return json.loads(result_text)
            
        except Exception as e:
            logger.error(f"AI analysis error: {e}")
            return {
                "task_type": pre_classified_type,
                "failure_modes": rule_based_data.get("failure_modes", []),
                "mechanisms": rule_based_data.get("failure_mechanisms", []),
                "detection_methods": rule_based_data.get("detection_methods", []),
                "confidence": 50,
                "reasoning": f"AI analysis failed: {str(e)[:100]}"
            }
    
    def _calculate_confidence(
        self,
        has_component: bool,
        has_frequency: bool,
        task_type_known: bool,
        ai_confidence: int,
        failure_modes_count: int
    ) -> int:
        """Calculate overall confidence score."""
        score = ai_confidence * 0.4  # AI contributes 40%
        
        if has_component:
            score += 15
        if has_frequency:
            score += 10
        if task_type_known:
            score += 15
        if failure_modes_count >= 2:
            score += 10
        if failure_modes_count >= 3:
            score += 10
        
        return min(100, max(0, int(score)))
    
    # ============================================================
    # AI Enrichment (new — replaces FM/library matching per refactor spec)
    # ============================================================
    
    # Canonical frequency vocabulary + day mapping
    _FREQUENCY_DAYS = {
        "Daily": 1,
        "Weekly": 7,
        "Biweekly": 14,
        "Monthly": 30,
        "Quarterly": 90,
        "Semi-Annual": 180,
        "Annual": 365,
        "Every 2 Years": 730,
        "Every 3 Years": 1095,
        "Condition Based": None,
        "One Time": None,
    }
    _DISCIPLINES = [
        "Mechanical", "Electrical", "Instrumentation",
        "Process", "Civil", "Operations", "HVAC",
    ]
    _TASK_TYPES = ["PM", "PDM", "CBM", "CM"]
    
    async def _ai_enrich_tasks(self, tasks: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Synchronous AI enrichment per the PM Import refactor spec. For each task,
        the LLM is asked to:
          1. Translate `task_description` to English
          2. Classify task_type ∈ {PM, PDM, CBM, CM}
          3. Suggest a discipline from the configured list
          4. Standardize frequency to canonical vocabulary
          5. Estimate labor hours
        
        Each call is batched to keep cost down. On any failure the task keeps
        rule-derived defaults so the import never fully blocks.
        """
        if not tasks:
            return tasks
        
        from services.openai_service import chat_completion
        
        # Batch into groups of 10 to stay within token limits
        BATCH = 10
        for batch_start in range(0, len(tasks), BATCH):
            batch = tasks[batch_start: batch_start + BATCH]
            
            # Build a compact prompt
            payload = []
            for idx, t in enumerate(batch):
                payload.append({
                    "i": idx,
                    "raw_task": (t.get("original_task") or "")[:400],
                    "component": (t.get("component") or "")[:120],
                    "asset": (t.get("asset") or "")[:120],
                    "raw_frequency": (t.get("frequency") or "")[:80],
                })
            
            sys_prompt = (
                "You are a maintenance engineering assistant. For each task, return a JSON "
                "object with keys exactly: task_description (English), task_type, discipline, "
                "frequency, frequency_days, estimated_hours, confidence_score. "
                f"task_type MUST be one of {self._TASK_TYPES}. "
                f"discipline MUST be one of {self._DISCIPLINES}. "
                f"frequency MUST be one of {list(self._FREQUENCY_DAYS.keys())}. "
                "frequency_days is the integer day count (null for Condition Based / One Time). "
                "estimated_hours is a float labor estimate (0.1–24). "
                "confidence_score is your overall confidence (0–100). "
                "Translate the task to clear English regardless of source language. "
                "Return ONLY a JSON object with key 'results' as a list keyed by index 'i'."
            )
            user_prompt = (
                "Enrich these maintenance tasks. Reply with JSON only.\n\n"
                + json.dumps(payload, ensure_ascii=False)
            )
            
            try:
                raw = await chat_completion(
                    messages=[
                        {"role": "system", "content": sys_prompt},
                        {"role": "user", "content": user_prompt},
                    ],
                    model="gpt-4o",
                    temperature=0.1,
                    response_format={"type": "json_object"},
                )
                parsed = json.loads(raw)
                results = parsed.get("results") or parsed.get("tasks") or []
                by_idx = {int(r.get("i")): r for r in results if "i" in r}
            except Exception as e:
                logger.warning(f"AI enrichment batch failed: {e}")
                by_idx = {}
            
            # Merge back into tasks
            for idx, t in enumerate(batch):
                ai = by_idx.get(idx, {})
                
                # Task description (English)
                t["task_description"] = (
                    ai.get("task_description")
                    or t.get("original_task")
                    or ""
                ).strip()
                
                # Task type — clamp to allowed
                tt = (ai.get("task_type") or "").upper().strip()
                t["task_type"] = tt if tt in self._TASK_TYPES else "PM"
                
                # Discipline — clamp to allowed
                disc = (ai.get("discipline") or "").strip()
                t["discipline"] = disc if disc in self._DISCIPLINES else (
                    t.get("discipline") if t.get("discipline") in self._DISCIPLINES
                    else "Mechanical"
                )
                
                # Frequency — clamp to allowed
                freq = (ai.get("frequency") or "").strip()
                if freq not in self._FREQUENCY_DAYS:
                    # try title-case variant
                    freq_tc = freq.title()
                    freq = freq_tc if freq_tc in self._FREQUENCY_DAYS else "Monthly"
                t["frequency"] = freq
                
                # Frequency days
                fdays = ai.get("frequency_days")
                if fdays is None:
                    fdays = self._FREQUENCY_DAYS.get(freq)
                try:
                    t["frequency_days"] = int(fdays) if fdays is not None else None
                except (TypeError, ValueError):
                    t["frequency_days"] = self._FREQUENCY_DAYS.get(freq)
                
                # Estimated hours
                try:
                    eh = float(ai.get("estimated_hours") or 0)
                    t["estimated_hours"] = max(0.1, min(24.0, eh)) if eh else 0.5
                except (TypeError, ValueError):
                    t["estimated_hours"] = 0.5
                
                # Confidence score
                try:
                    cs = int(ai.get("confidence_score") or 0)
                    t["confidence_score"] = max(0, min(100, cs)) if cs else 50
                except (TypeError, ValueError):
                    t["confidence_score"] = 50
        
        return tasks
    
    def _normalize_task_shape(self, task: Dict[str, Any]) -> Dict[str, Any]:
        """
        Project a task into the canonical PM Import output shape per the refactor spec.
        Drops FM-related fields entirely.
        """
        return {
            "task_id": task.get("task_id"),
            # Equipment fields
            "equipment_tag": task.get("equipment_tag") or task.get("asset") or "",
            "equipment_description": task.get("equipment_description") or task.get("component") or "",
            # Task fields
            "task_description": task.get("task_description") or task.get("original_task") or "",
            "task_type": task.get("task_type") or "PM",
            "discipline": task.get("discipline") or "Mechanical",
            "frequency": task.get("frequency") or "Monthly",
            "frequency_days": task.get("frequency_days"),
            "estimated_hours": task.get("estimated_hours") or 0.5,
            "confidence_score": task.get("confidence_score") or 50,
            # Match + review
            "equipment_match": task.get("equipment_match"),
            "review_status": task.get("review_status") or "pending",
            # Raw source preserved for traceability
            "original_task": task.get("original_task") or "",
        }
    
    # ============================================================
    # Multi-tag splitting (one task per tag in the Tag column)
    # ============================================================
    
    # Recognized separators between tags inside a single Tag-column cell.
    _MULTI_TAG_SEPARATORS = re.compile(
        r"\s*(?:,|;|\||/|\\|\+|\bin\b|\band\b|\b&\b|\ben\b|\bof\b|\n|\r)\s*",
        flags=re.IGNORECASE,
    )
    
    def _split_multi_tag_tasks(self, tasks: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        If a task's Tag column (`equipment_tag` / `asset`) contains multiple tags
        (e.g. "P-1001, P-1002" or "MTR-01 / MTR-02"), duplicate the task into one
        row per tag. All other fields are copied; each child gets:
          - a fresh task_id
          - parent_task_id pointing at the original
          - equipment_tag set to the single tag
        """
        if not tasks:
            return tasks
        
        out = []
        for task in tasks:
            raw_tag = (task.get("equipment_tag") or task.get("asset") or "").strip()
            if not raw_tag:
                out.append(task)
                continue
            
            # Split & filter
            parts = [p.strip() for p in self._MULTI_TAG_SEPARATORS.split(raw_tag)]
            parts = [p for p in parts if p and len(p) >= 2]
            
            # Deduplicate (case-insensitive) while preserving order
            seen = set()
            unique_parts = []
            for p in parts:
                key = p.upper()
                if key not in seen:
                    seen.add(key)
                    unique_parts.append(p)
            
            # If only one tag (or splitting produced nothing useful), keep as-is
            if len(unique_parts) <= 1:
                out.append(task)
                continue
            
            parent_id = task.get("task_id")
            for idx, tag_value in enumerate(unique_parts):
                child = dict(task)
                child["task_id"] = str(uuid.uuid4())
                child["parent_task_id"] = parent_id
                child["multi_tag_index"] = idx
                child["multi_tag_total"] = len(unique_parts)
                child["equipment_tag"] = tag_value
                child["asset"] = tag_value
                # Clear any prior auto-match so it's recomputed per tag
                child["equipment_match"] = None
                out.append(child)
        
        return out
    
    async def _match_equipment_to_hierarchy(self, tasks: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Match each task to ONE equipment node per the refactor spec.
        
        Priority 1: tag exact match against `equipment_nodes.tag`
        Priority 2: description fuzzy match against `equipment_nodes.name`
        
        Stores a single `equipment_match` object (or None) per task.
        """
        nodes_cursor = self.db.equipment_nodes.find({})
        nodes = await nodes_cursor.to_list(20000)
        
        by_tag = {}
        by_name_exact = {}
        all_names = []  # for fuzzy
        for n in nodes:
            tag = (n.get("tag") or "").strip()
            if tag:
                by_tag[tag.upper()] = n
            name = (n.get("name") or "").strip()
            if name:
                by_name_exact[name.lower()] = n
                all_names.append((name.lower(), n))
        
        for task in tasks:
            # Priority 1: tag exact match — ONLY use the explicit Tag column from the
            # upload (`equipment_tag` / `asset`). We deliberately do NOT mine the
            # description or task text for tag-like patterns.
            tag_field = (
                str(task.get("equipment_tag") or task.get("asset") or "").strip()
            )
            
            match = None
            
            if tag_field:
                # Try exact tag match (single token or the whole field)
                hit = by_tag.get(tag_field.upper())
                if hit:
                    match = {
                        "equipment_id": hit.get("id"),
                        "tag": hit.get("tag"),
                        "name": hit.get("name"),
                        "match_type": "tag_exact",
                        "confidence": 100,
                    }
                else:
                    # If the tag column contains multiple tokens (e.g. "P-1001, P-1002"),
                    # try each token individually.
                    for token in self._extract_tag_refs(tag_field):
                        hit = by_tag.get(token.upper())
                        if hit:
                            match = {
                                "equipment_id": hit.get("id"),
                                "tag": hit.get("tag"),
                                "name": hit.get("name"),
                                "match_type": "tag_exact",
                                "confidence": 100,
                            }
                            break
            
            # Priority 2: description fuzzy match
            if not match:
                description_candidates = [
                    str(task.get("equipment_description") or task.get("component") or "").strip().lower(),
                    str(task.get("equipment_tag") or task.get("asset") or "").strip().lower(),
                ]
                description_candidates = [c for c in description_candidates if len(c) >= 3]
                
                # Exact name match
                for cand in description_candidates:
                    if cand in by_name_exact:
                        hit = by_name_exact[cand]
                        match = {
                            "equipment_id": hit.get("id"),
                            "tag": hit.get("tag"),
                            "name": hit.get("name"),
                            "match_type": "name_exact",
                            "confidence": 90,
                        }
                        break
                
                # Partial / fuzzy contains
                if not match:
                    for cand in description_candidates:
                        best = None
                        best_score = 0
                        for name_lower, n in all_names:
                            if len(name_lower) < 4:
                                continue
                            # simple substring score
                            score = 0
                            if cand == name_lower:
                                score = 90
                            elif cand in name_lower:
                                score = 80
                            elif name_lower in cand:
                                score = 70
                            elif self._token_overlap_score(cand, name_lower) >= 0.5:
                                score = 60
                            if score > best_score:
                                best_score = score
                                best = n
                        if best and best_score >= 60:
                            match = {
                                "equipment_id": best.get("id"),
                                "tag": best.get("tag"),
                                "name": best.get("name"),
                                "match_type": "name_partial" if best_score < 90 else "name_exact",
                                "confidence": best_score,
                            }
                            break
            
            task["equipment_match"] = match
            # Preserve canonical equipment_tag / equipment_description fields
            task["equipment_tag"] = task.get("equipment_tag") or task.get("asset") or ""
            task["equipment_description"] = task.get("equipment_description") or task.get("component") or ""
        
        return tasks
    
    @staticmethod
    def _token_overlap_score(a: str, b: str) -> float:
        """Cheap token-overlap score in [0,1] used as a fuzzy fallback."""
        ta = {w for w in a.split() if len(w) >= 3}
        tb = {w for w in b.split() if len(w) >= 3}
        if not ta or not tb:
            return 0.0
        inter = ta & tb
        return len(inter) / max(len(ta), len(tb))
    
    def _calculate_stats(self, tasks: List[Dict[str, Any]]) -> Dict[str, int]:
        """Calculate statistics for the session."""
        stats = {
            "total_tasks": len(tasks),
            "failure_modes_identified": 0,
            "existing_matches": 0,
            "new_proposed": 0,
            "low_confidence_items": 0,
            "manual_review_required": 0,
            "accepted": 0,
            "rejected": 0,
            "pending": 0
        }
        
        fm_set = set()
        
        for task in tasks:
            # Count failure modes
            for fm in task.get("suggested_failure_modes", []):
                fm_name = fm if isinstance(fm, str) else fm.get("name", str(fm))
                fm_set.add(fm_name)
            
            # Count by library match status
            match_status = task.get("library_match", {}).get("status", "")
            if match_status == "existing_match":
                stats["existing_matches"] += 1
            elif match_status in ["new_proposed", "weak_match"]:
                stats["new_proposed"] += 1
            elif match_status == "multiple_possible":
                stats["manual_review_required"] += 1
            
            # Count by confidence
            confidence = task.get("confidence_score", 0)
            if confidence < 70:
                stats["low_confidence_items"] += 1
                stats["manual_review_required"] += 1
            
            # Count by review status
            review_status = task.get("review_status", "pending")
            if review_status == "accepted":
                stats["accepted"] += 1
            elif review_status == "rejected":
                stats["rejected"] += 1
            else:
                stats["pending"] += 1
        
        stats["failure_modes_identified"] = len(fm_set)
        
        return stats
    
    def _determine_category(self, task: Dict[str, Any]) -> str:
        """Determine the failure mode category based on task data."""
        task_type = task.get("task_type", "").lower()
        component = (task.get("component") or "").lower()
        
        # Map based on component type
        if any(kw in component for kw in ["pump", "compressor", "motor", "fan", "turbine", "bearing", "gear"]):
            return "Rotating"
        elif any(kw in component for kw in ["valve", "pipe", "tank", "vessel", "exchanger"]):
            return "Static"
        elif any(kw in component for kw in ["sensor", "transmitter", "controller", "plc", "gauge"]):
            return "Instrumentation"
        elif any(kw in component for kw in ["cable", "switch", "transformer", "breaker"]):
            return "Electrical"
        elif "seal" in component or "gasket" in component:
            return "Piping"
        
        # Map based on task type
        if task_type == "calibration":
            return "Instrumentation"
        elif task_type == "lubrication":
            return "Rotating"
        
        return "Process"  # Default
    
    def _extract_keywords(self, task: Dict[str, Any]) -> List[str]:
        """Extract keywords from task for the failure mode."""
        keywords = []
        
        # Add task type
        task_type = task.get("task_type", "")
        if task_type and task_type != "Unknown":
            keywords.append(task_type.lower())
        
        # Add component words
        component = task.get("component", "")
        if component:
            keywords.extend(component.lower().split())
        
        # Add from detection methods
        for method in task.get("detection_methods", [])[:2]:
            keywords.append(method.lower().split()[0])
        
        # Dedupe and limit
        return list(dict.fromkeys(keywords))[:6]
    
    async def _update_progress(self, session_id: str, progress: int, message: str):
        """Update session progress."""
        await self.sessions_collection.update_one(
            {"session_id": session_id},
            {"$set": {
                "progress": progress,
                "progress_message": message,
                "updated_at": datetime.now(timezone.utc)
            }}
        )
